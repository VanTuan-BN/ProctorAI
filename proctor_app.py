import sys, csv, logging, requests, os, unicodedata, base64, tempfile
import mysql.connector
import cv2
import numpy as np
from openpyxl import load_workbook
from PyQt5.QtCore import Qt, QTimer, QThreadPool
from PyQt5.QtGui import QBrush, QColor, QImage, QPixmap
from PyQt5.QtWidgets import (
    QApplication,
    QAbstractItemView,
    QComboBox,
    QDialog,
    QDoubleSpinBox,
    QFileDialog,
    QFormLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QProgressBar,
    QSpinBox,
    QStackedWidget,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)
from ui_branding import build_login_logo_label
from ui_components import AppToolbar, EmptyState, EnterpriseDialog, FilterBar, SidebarNav, connect_sidebar_to_tabs, wire_filter_bar
from qt_async import BackgroundTask
from ui_theme import apply_theme, polish_table, set_page_margins, style_stat_label
from auth_security import verify_and_upgrade_password
from password_recovery_dialog import PasswordRecoveryDialog

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s: %(message)s")
logger = logging.getLogger(__name__)
ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
EVIDENCE_DIR = os.path.join(ROOT_DIR, "server_evidence")
DB_HOST = os.getenv("S_MONITOR_DB_HOST", "127.0.0.1")
DB_USER = os.getenv("S_MONITOR_DB_USER", "root")
DB_PASSWORD = os.getenv("S_MONITOR_DB_PASSWORD", "12345")
DB_NAME = os.getenv("S_MONITOR_DB_NAME", "exam_monitor_db")


def _normalize_import_key(value):
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    return "".join(ch for ch in text.lower() if ch.isalnum())

class ProctorDashboard(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("S-MONITOR: TRUNG TÂM QUẢN LÝ GIÁM THỊ")
        self.setMinimumSize(1200, 800)
        apply_theme(self, role="proctor")
        
        self.proctor_id = None
        self.proctor_name = ""
        self.proctor_token = None
        self.template_dict = {} # Lưu id và tên bộ đề
        self.api_url = "http://127.0.0.1:8000" 
        self.current_question_id = None
        self.monitor_entries = []
        self.current_monitor_session_token = None
        self.current_monitor_class_id = None
        self.current_monitor_detail = {}
        self.thread_pool = QThreadPool.globalInstance()
        self.proctor_login_inflight = False
        self.monitor_overview_inflight = False
        self.snapshot_request_inflight = False
        self.preview_review_inflight = False
        self.close_class_inflight = False
        self.reports_inflight = False
        self.report_rows_cache = []
        self.monitor_timer = QTimer()
        self.monitor_timer.timeout.connect(self.load_monitor_overview)
        self.init_ui()

    def get_db(self):
        return mysql.connector.connect(host=DB_HOST, user=DB_USER, password=DB_PASSWORD, database=DB_NAME)

    def _proctor_api_headers(self, token=None):
        active_token = token or self.proctor_token
        return {"X-Proctor-Token": active_token} if active_token else {}

    def _show_dialog(self, title, message, detail=""):
        EnterpriseDialog(title, message, detail=detail, role="proctor", parent=self).exec_()

    def request_proctor_password_reset(self):
        dialog = PasswordRecoveryDialog("Giám thị", "ID Giám thị", account_value=self.inp_pid.text().strip(), parent=self)
        if dialog.exec_() != QDialog.Accepted:
            return
        payload = dialog.get_payload()
        if not payload["account_id"] or not payload["full_name"]:
            return self._show_dialog("Thiếu thông tin", "Vui lòng nhập ID Giám thị và Họ tên để gửi yêu cầu.")
        try:
            response = requests.post(
                f"{self.api_url}/api/password-recovery/request",
                data={
                    "role": "proctor",
                    "account_id": payload["account_id"],
                    "full_name": payload["full_name"],
                    "note": payload["note"],
                },
                timeout=10,
            )
            response.raise_for_status()
            data = response.json()
            if data.get("status") == "success":
                self._show_dialog("Đã gửi yêu cầu", data.get("message", "Quản trị viên sẽ cấp lại mật khẩu tạm sau khi kiểm tra."))
            else:
                self._show_dialog("Không gửi được yêu cầu", data.get("message", "Hệ thống từ chối yêu cầu cấp lại mật khẩu."))
        except requests.RequestException as exc:
            logger.exception("Proctor password recovery request failed")
            self._show_dialog("Kết nối thất bại", str(exc))
        except ValueError:
            logger.exception("Proctor password recovery returned invalid JSON")
            self._show_dialog("Phản hồi không hợp lệ", "Server trả về dữ liệu không đọc được.")

    def _run_background_task(self, fn, on_success, on_error=None, on_finished=None):
        task = BackgroundTask(fn)
        task.signals.result.connect(on_success)
        task.signals.error.connect(on_error or self._log_background_error)
        if on_finished is not None:
            task.signals.finished.connect(on_finished)
        self.thread_pool.start(task)

    def _log_background_error(self, payload):
        logger.error("Background task failed: %s", payload.get("traceback") or payload.get("message"))

    def _set_login_busy(self, busy, button_text=None):
        self.proctor_login_inflight = busy
        self.inp_pid.setEnabled(not busy)
        self.inp_ppw.setEnabled(not busy)
        self.btn_login.setEnabled(not busy)
        self.btn_login.setText(button_text or ("ĐĂNG NHẬP" if not busy else "ĐANG ĐĂNG NHẬP..."))

    def _apply_proctor_stats(self, stats):
        self.lbl_my_templates.setText(f"Bộ đề của tôi\n{stats.get('templates', 0)}")
        self.lbl_my_classes.setText(f"Lớp thi quản lý\n{stats.get('classes', 0)}")
        self.lbl_my_students.setText(f"Tổng Sinh viên\n{stats.get('students', 0)}")
        self.lbl_my_violations.setText(f"Cảnh báo Gian lận\n{stats.get('violations', 0)}")

    def _apply_templates_rows(self, rows):
        current_template_id = self._selected_template_id()
        self.cb_templates_for_class.clear()
        self.cb_templates_for_question.clear()
        self.tb_templates.setRowCount(len(rows))
        self.template_dict.clear()
        for row_index, row in enumerate(rows):
            template_label = f"[{row['template_id']}] {row['template_name']}"
            self.template_dict[template_label] = row['template_id']
            self.cb_templates_for_class.addItem(template_label)
            self.cb_templates_for_question.addItem(template_label)
            self.tb_templates.setItem(row_index, 0, QTableWidgetItem(str(row['template_id'])))
            self.tb_templates.setItem(row_index, 1, QTableWidgetItem(str(row['template_name'])))
            self.tb_templates.setItem(row_index, 2, QTableWidgetItem(str(row.get('question_count', 0))))
        self.templates_empty.setVisible(len(rows) == 0)
        self.tb_templates.setVisible(len(rows) > 0)
        if current_template_id is not None:
            self._set_active_template(current_template_id)
        if self.cb_templates_for_question.count() > 0 and self.cb_templates_for_question.currentIndex() < 0:
            self.cb_templates_for_question.setCurrentIndex(0)
        if self.cb_templates_for_question.count() > 0:
            self.refresh_template_details()
        else:
            self.refresh_template_details()

    def _populate_reports_table(self, rows):
        self.report_rows_cache = rows or []
        self._sync_report_class_filter_options(self.report_rows_cache)

        selected_class_id = ""
        selected_review_status = ""
        if hasattr(self, "cb_reports_class_filter"):
            selected_class_id = str(self.cb_reports_class_filter.currentData() or "").strip()
        if hasattr(self, "cb_reports_status_filter"):
            selected_review_status = str(self.cb_reports_status_filter.currentData() or "").strip().lower()

        visible_rows = self.report_rows_cache
        if selected_class_id:
            visible_rows = [
                row for row in self.report_rows_cache
                if str(row.get("exam_id") or "").strip() == selected_class_id
            ]
        if selected_review_status:
            visible_rows = [
                row for row in visible_rows
                if str(row.get("review_status") or "pending").strip().lower() == selected_review_status
            ]

        self.table.setRowCount(len(visible_rows))
        self.reports_empty.setVisible(len(visible_rows) == 0)
        self.table.setVisible(len(visible_rows) > 0)
        for i, row in enumerate(visible_rows):
            self.table.setItem(i, 0, QTableWidgetItem(str(row['msv'])))
            self.table.setItem(i, 1, QTableWidgetItem(str(row['exam_id'])))
            self.table.setItem(i, 2, QTableWidgetItem(str(row['time_detected'])))
            self.table.setItem(i, 3, QTableWidgetItem(str(row['error_type'])))
            self.table.setItem(i, 4, QTableWidgetItem(str(row['evidence_path'])))

            review_status = str(row.get('review_status') or 'pending').strip().lower()
            review_label_map = {
                'confirm': 'CONFIRM',
                'reject': 'REJECT',
                'pending': 'PENDING',
            }
            review_text = review_label_map.get(review_status, review_status.upper() if review_status else 'PENDING')
            review_item = QTableWidgetItem(review_text)
            review_item.setTextAlignment(Qt.AlignCenter)
            review_item.setData(Qt.UserRole, str(row.get('review_note') or ''))
            review_item.setData(Qt.UserRole + 1, str(row.get('reviewed_by') or ''))
            review_item.setData(Qt.UserRole + 2, str(row.get('reviewed_at') or '--'))
            review_color_map = {
                'confirm': (QColor('#E8F7EA'), QColor('#1B5E20')),
                'reject': (QColor('#FDECEC'), QColor('#8E1C1C')),
                'pending': (QColor('#FFF7DB'), QColor('#7A4B00')),
            }
            background_color, foreground_color = review_color_map.get(review_status, (QColor('#EEF2F7'), QColor('#334155')))
            review_item.setBackground(QBrush(background_color))
            review_item.setForeground(QBrush(foreground_color))
            self.table.setItem(i, 5, review_item)

            # Cột clip: hiển thị dấu tick nếu đã có clip event
            has_clip = bool(str(row.get('clip_path', '')).strip())
            dur = float(row.get('duration_seconds') or 0)
            clip_text = f"✅ {dur:.1f}s" if has_clip else "❌"
            clip_item = QTableWidgetItem(clip_text)
            clip_item.setTextAlignment(Qt.AlignCenter)
            clip_item.setData(Qt.UserRole, str(row.get('clip_path', '') or '').strip())
            clip_item.setData(Qt.UserRole + 1, float(dur))
            self.table.setItem(i, 6, clip_item)

    def _sync_report_class_filter_options(self, rows):
        if not hasattr(self, "cb_reports_class_filter"):
            return
        selected_data = str(self.cb_reports_class_filter.currentData() or "").strip()
        class_ids = sorted({str(row.get("exam_id") or "").strip() for row in (rows or []) if str(row.get("exam_id") or "").strip()})

        self.cb_reports_class_filter.blockSignals(True)
        self.cb_reports_class_filter.clear()
        self.cb_reports_class_filter.addItem("Tất cả phòng", "")
        for class_id in class_ids:
            self.cb_reports_class_filter.addItem(f"Phòng {class_id}", class_id)

        idx = self.cb_reports_class_filter.findData(selected_data)
        if idx < 0:
            idx = 0
        self.cb_reports_class_filter.setCurrentIndex(idx)
        self.cb_reports_class_filter.blockSignals(False)

    def on_reports_class_filter_changed(self, _index):
        self._populate_reports_table(self.report_rows_cache)

    def on_reports_status_filter_changed(self, _index):
        self._populate_reports_table(self.report_rows_cache)

    def _populate_scores_table(self, rows):
        self.tb_scores.setRowCount(len(rows))
        self.scores_empty.setVisible(len(rows) == 0)
        self.tb_scores.setVisible(len(rows) > 0)
        for i, row in enumerate(rows):
            self.tb_scores.setItem(i, 0, QTableWidgetItem(str(row['msv'])))
            self.tb_scores.setItem(i, 1, QTableWidgetItem(str(row['full_name'])))
            self.tb_scores.setItem(i, 2, QTableWidgetItem(str(row['class_id'])))
            self.tb_scores.setItem(i, 3, QTableWidgetItem(str(row['class_name'])))
            score_item = QTableWidgetItem(f"{row['score']} Điểm")
            score_item.setTextAlignment(Qt.AlignCenter)
            self.tb_scores.setItem(i, 4, score_item)
            self.tb_scores.setItem(i, 5, QTableWidgetItem(str(row['submission_time'])))

    def _apply_monitor_overview_payload(self, rows):
        ordered_rows = sorted(
            rows or [],
            key=lambda row: (
                0 if bool(row.get("manual_review_required", False)) else 1,
                -float(row.get("seconds_since_update", 0.0) or 0.0),
            ),
        )
        self.monitor_entries = ordered_rows
        self.apply_monitor_filter()
        if self.current_monitor_session_token:
            for row_index in range(self.tb_monitor.rowCount()):
                item = self.tb_monitor.item(row_index, 0)
                if item and item.data(Qt.UserRole) == self.current_monitor_session_token:
                    self.tb_monitor.selectRow(row_index)
                    break
        elif self.tb_monitor.rowCount() > 0:
            self.tb_monitor.selectRow(0)

    def _apply_monitor_detail_payload(self, session_token, detail):
        if session_token != self.current_monitor_session_token:
            return
        self.current_monitor_detail = detail or {}
        self.current_monitor_class_id = detail.get("class_id")
        pixmap = self._load_pixmap_from_base64(detail.get("preview_b64"))
        if pixmap is not None and not pixmap.isNull():
            self.lbl_monitor_preview.setPixmap(pixmap.scaled(620, 360, Qt.KeepAspectRatio, Qt.SmoothTransformation))
            self.lbl_monitor_preview.setText("")
        else:
            self.lbl_monitor_preview.setPixmap(QPixmap())
            self.lbl_monitor_preview.setText("Chưa có khung hình gần nhất")
        self.lbl_monitor_student.setText(f"Sinh viên: {detail.get('msv', '--')} | {detail.get('full_name', '')}")
        self.lbl_monitor_class.setText(f"Lớp thi: {detail.get('class_id', '--')} | {detail.get('class_name', '')}")
        self.lbl_monitor_status.setText(
            f"Trạng thái AI: {detail.get('identity_status', '--')} | Risk {detail.get('risk_score', 0):.0f}% | Cập nhật {detail.get('seconds_since_update', 0)}s trước"
        )
        self.lbl_monitor_exam.setText(
            f"Bài làm: {detail.get('answered_count', 0)}/{detail.get('question_total', 0)} câu | Đúng {detail.get('correct_count', 0)} | Sai {detail.get('wrong_count', 0)}"
        )
        self.lbl_monitor_score.setText(
            f"Điểm tạm: {detail.get('current_score', 0.0):.2f} | Đã nộp: {'Có' if detail.get('submitted') else 'Chưa'} | Xác minh: {'OK' if detail.get('verified') else 'Chưa'}"
        )
        self.lbl_monitor_warnings.setText(
            f"Cảnh báo phiên: {detail.get('session_warning_count', 0)}/{detail.get('max_warnings', 0)} | Lịch sử: {detail.get('historical_warning_count', 0)} | Tổng: {detail.get('total_warning_count', 0)} | Khóa phiên sinh viên: {'Có' if detail.get('warning_locked') else 'Không'} | Khóa phòng thi: {'Có' if detail.get('class_locked') else 'Không'}"
        )
        manual_review_required = bool(detail.get("manual_review_required", False))
        if manual_review_required:
            violation_type = str(detail.get("manual_review_violation_type") or "hành vi nghi vấn").strip() or "hành vi nghi vấn"
            review_title = str(detail.get("manual_review_title") or "Tầng 2 chưa chắc chắn").strip()
            review_reason = str(detail.get("manual_review_reason") or "Cần giám thị đánh giá thủ công.").strip()
            review_source = str(detail.get("manual_review_source") or "Hệ thống AI").strip()
            self.lbl_manual_review_badge.setText(f"{review_title} | Chờ giám thị duyệt")
            self.lbl_manual_review_badge.setStyleSheet("background:#FFF4D6; color:#7A4B00; border:1px solid #E8B14C; border-radius:8px; padding:8px 10px; font-size:13px; font-weight:700;")
            self.lbl_manual_review_detail.setText(f"Vi phạm liên quan: {violation_type} | Nguồn: {review_source} | Lý do: {review_reason}")
        else:
            self.lbl_manual_review_badge.setText("Không có case cần review thủ công")
            self.lbl_manual_review_badge.setStyleSheet("background:#EEF4FF; color:#234A84; border:1px solid #9EB9E5; border-radius:8px; padding:8px 10px; font-size:13px; font-weight:600;")
            self.lbl_manual_review_detail.setText("Các vi phạm đã đủ chắc chắn hoặc chưa có tín hiệu cần giám thị đánh giá thủ công.")
        class_locked = bool(detail.get("class_locked", False))
        self.btn_request_snapshot.setEnabled(not self.snapshot_request_inflight and bool(self.current_monitor_session_token))
        has_preview = bool(detail.get("frame_available", False))
        self.btn_preview_confirm.setEnabled(bool(self.current_monitor_session_token) and has_preview and not self.preview_review_inflight)
        self.btn_preview_reject.setEnabled(bool(self.current_monitor_session_token) and has_preview and not self.preview_review_inflight)
        self.btn_close_class.setEnabled(not self.close_class_inflight and bool(self.current_monitor_class_id) and not class_locked)
        self.btn_open_class.setEnabled(not self.close_class_inflight and bool(self.current_monitor_class_id) and class_locked)
        yolo_world = detail.get("yolo_world") or {}
        earpiece_tier2 = detail.get("earpiece_tier2") or {}
        yw_thresholds = yolo_world.get("thresholds") or {}
        ep_thresholds = earpiece_tier2.get("thresholds") or {}
        ep_verdict_raw = str(earpiece_tier2.get("verdict") or "").strip().lower()
        ep_verdict_map = {
            "confirm": "CONFIRM",
            "reject": "REJECT",
            "review": "REVIEW",
        }
        ep_verdict = ep_verdict_map.get(ep_verdict_raw, "--")
        yw_threshold_text = (
            f"task_conf={yw_thresholds.get('task_conf', '--')} | "
            f"vote={yw_thresholds.get('temporal_min_votes', '--')}/{yw_thresholds.get('temporal_window_size', '--')}"
        )
        if yw_thresholds.get("phone_confirm_conf") is not None:
            yw_threshold_text += (
                f" | phone_confirm>={yw_thresholds.get('phone_confirm_conf')}"
                f" | area=[{yw_thresholds.get('phone_area_ratio_min', '--')},{yw_thresholds.get('phone_area_ratio_max', '--')}]"
                f" | ar=[{yw_thresholds.get('phone_aspect_ratio_min', '--')},{yw_thresholds.get('phone_aspect_ratio_max', '--')}]"
            )
        if ep_thresholds:
            yw_threshold_text += (
                f" | earpiece_specialist={ep_thresholds.get('specialist_conf', '--')}"
                f"/{ep_thresholds.get('specialist_strict_conf', '--')}"
            )
        summary_lines = [
            f"Sự kiện gần nhất: {detail.get('last_event', 'Không có')}",
            f"Snapshot gần nhất: {detail.get('snapshot_updated_at', '--')} | Trạng thái: {detail.get('snapshot_status', 'missing')} | Nguồn: {detail.get('snapshot_source', '--')}",
            f"Review thủ công: {'CẦN DUYỆT' if manual_review_required else 'KHÔNG'} | Vi phạm liên quan: {detail.get('manual_review_violation_type', '--') or '--'} | Nguồn: {detail.get('manual_review_source', '--') or '--'}",
            f"Đánh giá preview: {str(detail.get('preview_review_status', 'pending')).upper()} | Người duyệt: {detail.get('preview_reviewed_by', '--')} | Lúc: {detail.get('preview_reviewed_at', '--')}",
            f"Ghi chú preview: {detail.get('preview_review_note', '') or '(không có)'}",
            f"Yêu cầu snapshot chờ xử lý: {'Có' if detail.get('snapshot_requested') else 'Không'} | Lúc: {detail.get('snapshot_requested_at', '--')}",
            f"Model summary: people={detail.get('people_count', 0)}, phone={detail.get('phone_detected')}, head-pose audit={detail.get('head_pose_status', 'unavailable')}",
            f"Pose: pitch={detail.get('pitch', 0.0)}, yaw={detail.get('yaw', 0.0)}, reliable={detail.get('pose_reliable')}",
            f"EARPIECE Tier2: verdict={ep_verdict} | label={earpiece_tier2.get('top_label', '--')} | conf={float(earpiece_tier2.get('top_confidence', 0.0) or 0.0):.2f} | status={earpiece_tier2.get('status', 'idle')} | cập nhật={earpiece_tier2.get('updated_at', '--')}",
            f"YOLO-World: status={yolo_world.get('status', 'idle')} | label={yolo_world.get('top_label', '--')} | conf={float(yolo_world.get('top_confidence', 0.0) or 0.0):.2f} | verdict={yolo_world.get('verdict', '--')} | ms={yolo_world.get('inference_ms', '--')} | cập nhật={yolo_world.get('updated_at', '--')}",
            f"YOLO-World thresholds: {yw_threshold_text}",
            f"Trạng thái phiên: {'Đang hoạt động' if detail.get('active') else 'Không còn cập nhật mới'}",
        ]
        self.txt_monitor_predictions.setPlainText("\n".join(summary_lines))
        self._set_monitor_risk_bar(self.pb_monitor_phone, detail, "phone")
        self._set_monitor_risk_bar(self.pb_monitor_people, detail, "multiple_people")
        self._set_monitor_risk_bar(self.pb_monitor_away, detail, "away")
        self._set_monitor_risk_bar(self.pb_monitor_intruder, detail, "intruder")

    def _fetch_proctor_login_bundle(self, pid, pw):
        db = self.get_db()
        try:
            cur = db.cursor(dictionary=True)
            cur.execute("SELECT proctor_id, full_name, password FROM proctors WHERE proctor_id = %s", (pid,))
            res = cur.fetchone()
            if not res:
                return {"status": "error", "message": "Sai ID hoặc Mật khẩu Giám thị!"}

            verified, upgraded = verify_and_upgrade_password(cur, "proctors", "proctor_id", pid, pw, res.get("password"))
            if not verified:
                return {"status": "error", "message": "Sai ID hoặc Mật khẩu Giám thị!"}
            if upgraded:
                db.commit()

            cur.execute(
                """
                SELECT et.template_id, et.template_name, COUNT(q.q_id) AS question_count
                FROM exam_templates et
                LEFT JOIN question_bank q ON q.template_id = et.template_id
                WHERE et.proctor_id = %s
                GROUP BY et.template_id, et.template_name
                ORDER BY et.template_id DESC
                """,
                (pid,),
            )
            templates = cur.fetchall()
            cur.execute(
                """
                SELECT v.msv, v.exam_id, v.time_detected, v.error_type, v.evidence_path,
                       COALESCE(v.clip_path, '') AS clip_path, COALESCE(v.duration_seconds, 0) AS duration_seconds
                FROM violations v JOIN classes c ON v.exam_id = c.class_id
                JOIN exam_templates et ON c.template_id = et.template_id
                WHERE et.proctor_id = %s ORDER BY v.time_detected DESC
                """,
                (pid,),
            )
            reports = cur.fetchall()
            cur.execute("SELECT COUNT(*) AS value FROM exam_templates WHERE proctor_id = %s", (pid,))
            templates_count = int(cur.fetchone()["value"])
            cur.execute("SELECT COUNT(*) AS value FROM classes c JOIN exam_templates et ON c.template_id = et.template_id WHERE et.proctor_id = %s", (pid,))
            classes_count = int(cur.fetchone()["value"])
            cur.execute("SELECT COUNT(DISTINCT cs.msv) AS value FROM class_students cs JOIN classes c ON cs.class_id = c.class_id JOIN exam_templates et ON c.template_id = et.template_id WHERE et.proctor_id = %s", (pid,))
            students_count = int(cur.fetchone()["value"])
            cur.execute("SELECT COUNT(*) AS value FROM violations v JOIN classes c ON v.exam_id = c.class_id JOIN exam_templates et ON c.template_id = et.template_id WHERE et.proctor_id = %s", (pid,))
            violations_count = int(cur.fetchone()["value"])
            cur.execute(
                """
                SELECT r.msv, s.full_name, c.class_id, c.class_name, r.score, r.submission_time
                FROM exam_results r JOIN students s ON r.msv = s.msv
                JOIN classes c ON r.exam_id = c.class_id JOIN exam_templates et ON c.template_id = et.template_id
                WHERE et.proctor_id = %s ORDER BY r.submission_time DESC
                """,
                (pid,),
            )
            scores = cur.fetchall()
        finally:
            db.close()

        proctor_token = None
        try:
            login_response = requests.post(
                f"{self.api_url}/api/proctor/login",
                data={"proctor_id": pid, "password": pw},
                timeout=10,
            )
            login_response.raise_for_status()
            login_payload = login_response.json()
            if login_payload.get("status") == "success":
                proctor_token = login_payload.get("token")
                response = requests.get(
                    f"{self.api_url}/api/monitor/proctor/{pid}",
                    headers=self._proctor_api_headers(proctor_token),
                    timeout=5,
                )
                response.raise_for_status()
                monitor_payload = response.json()
                monitor_entries = monitor_payload.get("data", []) if monitor_payload.get("status") == "success" else []
            else:
                monitor_entries = []
        except Exception:
            logger.exception("Failed to preload proctor monitor overview for %s", pid)
            monitor_entries = []

        return {
            "status": "success",
            "proctor_id": pid,
            "proctor_name": res["full_name"],
            "proctor_token": proctor_token,
            "templates": templates,
            "reports": reports,
            "stats": {
                "templates": templates_count,
                "classes": classes_count,
                "students": students_count,
                "violations": violations_count,
            },
            "scores": scores,
            "monitor_entries": monitor_entries,
        }

    def _on_proctor_login_finished(self, payload):
        if payload.get("status") != "success":
            QMessageBox.warning(self, "Lỗi", payload.get("message", "Đăng nhập thất bại"))
            return
        self.proctor_id = payload["proctor_id"]
        self.proctor_name = payload["proctor_name"]
        self.proctor_token = payload.get("proctor_token")
        self.lbl_welcome.setText(f"QUẢN LÝ GIÁM THỊ: {self.proctor_name.upper()}")
        self.stack.setCurrentIndex(1)
        self._apply_templates_rows(payload["templates"])
        self._apply_monitor_overview_payload(payload["monitor_entries"])
        self._populate_reports_table(payload["reports"])
        self._apply_proctor_stats(payload["stats"])
        self._populate_scores_table(payload["scores"])
        self.monitor_timer.start(4000)

    def _on_proctor_login_error(self, payload):
        logger.error("Proctor login background task failed: %s", payload.get("traceback") or payload.get("message"))
        QMessageBox.critical(self, "Lỗi DB", payload.get("message", "Không thể đăng nhập."))
    def _fetch_proctor_monitor_overview(self):
        if not self.proctor_token:
            return []
        response = requests.get(
            f"{self.api_url}/api/monitor/proctor/{self.proctor_id}",
            headers=self._proctor_api_headers(),
            timeout=5,
        )
        response.raise_for_status()
        data = response.json()
        return data.get("data", []) if data.get("status") == "success" else []

    def _fetch_proctor_monitor_detail(self, session_token):
        if not self.proctor_token:
            return {"session_token": session_token, "detail": {}}
        response = requests.get(
            f"{self.api_url}/api/monitor/proctor/{self.proctor_id}/{session_token}",
            headers=self._proctor_api_headers(),
            timeout=5,
        )
        response.raise_for_status()
        payload = response.json()
        if payload.get("status") != "success":
            return {"session_token": session_token, "detail": {}}
        return {"session_token": session_token, "detail": payload.get("data", {})}

    def _fetch_proctor_snapshot_request(self, session_token):
        if not self.proctor_token:
            return {"status": "error", "message": "Chưa có phiên giám sát hợp lệ trên API Server."}
        response = requests.post(
            f"{self.api_url}/api/monitor/proctor/{self.proctor_id}/{session_token}/request_snapshot",
            headers=self._proctor_api_headers(),
            timeout=5,
        )
        response.raise_for_status()
        return response.json()

    def _fetch_proctor_preview_review(self, session_token, action, note):
        if not self.proctor_token:
            return {"status": "error", "message": "Chưa có phiên giám sát hợp lệ trên API Server."}
        yolo_world = self.current_monitor_detail.get("yolo_world") or {}
        earpiece_tier2 = self.current_monitor_detail.get("earpiece_tier2") or {}
        violation_id = yolo_world.get("violation_id") or earpiece_tier2.get("violation_id")
        response = requests.post(
            f"{self.api_url}/api/monitor/proctor/{self.proctor_id}/{session_token}/review_preview",
            headers=self._proctor_api_headers(),
            data={
                "action": action,
                "note": note,
                "violation_id": str(violation_id) if violation_id is not None else "",
            },
            timeout=5,
        )
        response.raise_for_status()
        return response.json()

    def _close_proctor_class(self, class_id):
        if not self.proctor_token:
            return {"status": "error", "message": "Chưa có phiên giám thị hợp lệ trên API Server."}
        response = requests.post(
            f"{self.api_url}/api/monitor/proctor/{self.proctor_id}/classes/{class_id}/lock",
            headers=self._proctor_api_headers(),
            timeout=5,
        )
        response.raise_for_status()
        return response.json()

    def _open_proctor_class(self, class_id):
        if not self.proctor_token:
            return {"status": "error", "message": "Chưa có phiên giám thị hợp lệ trên API Server."}
        response = requests.post(
            f"{self.api_url}/api/monitor/proctor/{self.proctor_id}/classes/{class_id}/unlock",
            headers=self._proctor_api_headers(),
            timeout=5,
        )
        response.raise_for_status()
        return response.json()

    def _on_snapshot_request_success(self, payload):
        self.snapshot_request_inflight = False
        self.btn_request_snapshot.setEnabled(bool(self.current_monitor_session_token))
        self.btn_request_snapshot.setText("YÊU CẦU SNAPSHOT")
        if payload.get("status") == "success":
            self._show_dialog("Đã gửi yêu cầu", payload.get("message", "Máy sinh viên sẽ gửi snapshot khi đồng bộ phiên tiếp theo."))
            self.load_selected_monitor_detail()
        else:
            self._show_dialog("Không gửi được yêu cầu", payload.get("message", "Server từ chối yêu cầu snapshot."))

    def _on_snapshot_request_error(self, payload):
        self.snapshot_request_inflight = False
        self.btn_request_snapshot.setEnabled(bool(self.current_monitor_session_token))
        self.btn_request_snapshot.setText("YÊU CẦU SNAPSHOT")
        self._show_dialog("Không gửi được yêu cầu", payload.get("message", "Không thể yêu cầu snapshot từ máy sinh viên."))

    def submit_preview_review(self, action):
        session_token = self.current_monitor_session_token
        if not session_token or self.preview_review_inflight:
            return
        self.preview_review_inflight = True
        self.btn_preview_confirm.setEnabled(False)
        self.btn_preview_reject.setEnabled(False)
        note = self.inp_preview_review_note.text().strip()
        self._run_background_task(
            lambda: self._fetch_proctor_preview_review(session_token, action, note),
            self._on_preview_review_success,
            on_error=self._on_preview_review_error,
        )

    def _on_preview_review_success(self, payload):
        self.preview_review_inflight = False
        self.btn_preview_confirm.setEnabled(bool(self.current_monitor_session_token))
        self.btn_preview_reject.setEnabled(bool(self.current_monitor_session_token))
        if payload.get("status") == "success":
            review_db = (payload.get("data") or {}).get("violation_review_db") or {}
            if review_db.get("updated"):
                detail = f"Đã cập nhật violations.violation_id={review_db.get('violation_id')} trong DB."
            else:
                detail = f"DB chưa cập nhật: {review_db.get('reason', 'unknown')}"
            self._show_dialog("Đã lưu đánh giá", payload.get("message", "Đã cập nhật đánh giá preview."), detail)
            self.load_selected_monitor_detail()
        else:
            self._show_dialog("Không lưu được đánh giá", payload.get("message", "Server từ chối đánh giá preview."))

    def _on_preview_review_error(self, payload):
        self.preview_review_inflight = False
        self.btn_preview_confirm.setEnabled(bool(self.current_monitor_session_token))
        self.btn_preview_reject.setEnabled(bool(self.current_monitor_session_token))
        self._show_dialog("Lỗi đánh giá preview", payload.get("message", "Không thể gửi đánh giá preview tới máy chủ."))

    def _on_close_class_success(self, payload):
        self.close_class_inflight = False
        self.btn_close_class.setText("KHÓA PHÒNG THI")
        self.btn_close_class.setEnabled(bool(self.current_monitor_class_id))
        self.btn_open_class.setText("MỞ LẠI PHÒNG THI")
        self.btn_open_class.setEnabled(bool(self.current_monitor_class_id))
        if payload.get("status") == "success":
            self._show_dialog("Đã khóa phòng thi", payload.get("message", "Lớp thi đã được khóa thủ công."))
            self.load_monitor_overview()
            self.load_selected_monitor_detail()
            self.refresh_proctor_stats()
        else:
            self._show_dialog("Không khóa được phòng thi", payload.get("message", "Hệ thống từ chối thao tác khóa phòng thi."))

    def _on_close_class_error(self, payload):
        self.close_class_inflight = False
        self.btn_close_class.setText("KHÓA PHÒNG THI")
        self.btn_close_class.setEnabled(bool(self.current_monitor_class_id))
        self.btn_open_class.setText("MỞ LẠI PHÒNG THI")
        self.btn_open_class.setEnabled(bool(self.current_monitor_class_id))
        self._show_dialog("Không khóa được phòng thi", payload.get("message", "Không thể khóa phòng thi đã chọn."))

    def _on_open_class_success(self, payload):
        self.close_class_inflight = False
        self.btn_open_class.setText("MỞ LẠI PHÒNG THI")
        self.btn_open_class.setEnabled(bool(self.current_monitor_class_id))
        self.btn_close_class.setText("KHÓA PHÒNG THI")
        self.btn_close_class.setEnabled(bool(self.current_monitor_class_id))
        if payload.get("status") == "success":
            self._show_dialog("Đã mở lại phòng thi", payload.get("message", "Lớp thi đã được mở lại."))
            self.load_monitor_overview()
            self.load_selected_monitor_detail()
            self.refresh_proctor_stats()
        else:
            self._show_dialog("Không mở được phòng thi", payload.get("message", "Hệ thống từ chối thao tác mở lại phòng thi."))

    def _on_open_class_error(self, payload):
        self.close_class_inflight = False
        self.btn_open_class.setText("MỞ LẠI PHÒNG THI")
        self.btn_open_class.setEnabled(bool(self.current_monitor_class_id))
        self.btn_close_class.setText("KHÓA PHÒNG THI")
        self.btn_close_class.setEnabled(bool(self.current_monitor_class_id))
        self._show_dialog("Không mở được phòng thi", payload.get("message", "Không thể mở lại phòng thi đã chọn."))

    def _fetch_proctor_reports(self, keyword=""):
        db = self.get_db()
        try:
            cur = db.cursor(dictionary=True)
            if keyword:
                query = """
                    SELECT v.msv, v.exam_id, v.time_detected, v.error_type, v.evidence_path,
                           COALESCE(v.clip_path, '') AS clip_path, COALESCE(v.duration_seconds, 0) AS duration_seconds,
                           COALESCE(v.review_status, 'pending') AS review_status,
                           COALESCE(v.review_note, '') AS review_note,
                           COALESCE(v.reviewed_by, '') AS reviewed_by,
                           v.reviewed_at
                    FROM violations v JOIN classes c ON v.exam_id = c.class_id
                    JOIN exam_templates et ON c.template_id = et.template_id
                    WHERE et.proctor_id = %s AND (v.msv LIKE %s OR CAST(v.exam_id AS CHAR) LIKE %s OR v.error_type LIKE %s)
                    ORDER BY v.time_detected DESC
                """
                cur.execute(query, (self.proctor_id, f"%{keyword}%", f"%{keyword}%", f"%{keyword}%"))
            else:
                query = """
                    SELECT v.msv, v.exam_id, v.time_detected, v.error_type, v.evidence_path,
                           COALESCE(v.clip_path, '') AS clip_path, COALESCE(v.duration_seconds, 0) AS duration_seconds,
                           COALESCE(v.review_status, 'pending') AS review_status,
                           COALESCE(v.review_note, '') AS review_note,
                           COALESCE(v.reviewed_by, '') AS reviewed_by,
                           v.reviewed_at
                    FROM violations v JOIN classes c ON v.exam_id = c.class_id
                    JOIN exam_templates et ON c.template_id = et.template_id
                    WHERE et.proctor_id = %s ORDER BY v.time_detected DESC
                """
                cur.execute(query, (self.proctor_id,))
            rows = cur.fetchall() or []
            for row in rows:
                row.setdefault("review_status", "pending")
                row.setdefault("review_note", "")
                row.setdefault("reviewed_by", "")
                row.setdefault("reviewed_at", None)
            return rows
        finally:
            db.close()

    def _on_proctor_monitor_overview_loaded(self, rows):
        self._apply_monitor_overview_payload(rows)

    def _on_proctor_monitor_overview_error(self, payload):
        logger.error("Proctor monitor overview failed: %s", payload.get("traceback") or payload.get("message"))

    def _on_proctor_reports_error(self, payload):
        logger.error("Proctor reports load failed: %s", payload.get("traceback") or payload.get("message"))
        self._show_dialog("Không thể tải báo cáo", payload.get("message", "Không tải được báo cáo gian lận."))

    def _selected_template_id(self):
        selected_template = self.cb_templates_for_question.currentText().strip()
        return self.template_dict.get(selected_template)

    def _clear_question_form(self):
        self.current_question_id = None
        self.inp_q.clear()
        self.inp_a.clear()
        self.inp_b.clear()
        self.inp_c.clear()
        self.inp_d.clear()
        self.inp_correct.setCurrentText("A")
        self.inp_points.setValue(1.0)
        self.btn_save_question.setText("THÊM CÂU HỎI VÀO BỘ ĐỀ NÀY")

    def _load_question_into_form(self):
        row = self.tb_template_questions.currentRow()
        if row < 0:
            self._clear_question_form()
            return
        question_id_item = self.tb_template_questions.item(row, 0)
        if question_id_item is None:
            self._clear_question_form()
            return
        self.current_question_id = int(question_id_item.text())
        self.inp_q.setPlainText(self.tb_template_questions.item(row, 1).text() if self.tb_template_questions.item(row, 1) else "")
        self.inp_a.setText(self.tb_template_questions.item(row, 2).text() if self.tb_template_questions.item(row, 2) else "")
        self.inp_b.setText(self.tb_template_questions.item(row, 3).text() if self.tb_template_questions.item(row, 3) else "")
        self.inp_c.setText(self.tb_template_questions.item(row, 4).text() if self.tb_template_questions.item(row, 4) else "")
        self.inp_d.setText(self.tb_template_questions.item(row, 5).text() if self.tb_template_questions.item(row, 5) else "")
        self.inp_correct.setCurrentText(self.tb_template_questions.item(row, 6).text() if self.tb_template_questions.item(row, 6) else "A")
        try:
            self.inp_points.setValue(float(self.tb_template_questions.item(row, 7).text()))
        except Exception:
            self.inp_points.setValue(1.0)
        self.btn_save_question.setText("CẬP NHẬT CÂU HỎI ĐÃ CHỌN")

    def _set_active_template(self, template_id):
        for combo in [self.cb_templates_for_class, self.cb_templates_for_question]:
            for index in range(combo.count()):
                if self.template_dict.get(combo.itemText(index)) == template_id:
                    combo.setCurrentIndex(index)
                    break

    def _load_pixmap_from_bytes(self, image_bytes):
        if not image_bytes:
            return None
        pixmap = QPixmap()
        if pixmap.loadFromData(image_bytes) and not pixmap.isNull():
            return pixmap
        np_buffer = np.frombuffer(image_bytes, np.uint8)
        decoded = cv2.imdecode(np_buffer, cv2.IMREAD_COLOR)
        if decoded is None:
            return None
        rgb = cv2.cvtColor(decoded, cv2.COLOR_BGR2RGB)
        image = QImage(rgb.data, rgb.shape[1], rgb.shape[0], rgb.strides[0], QImage.Format_RGB888)
        return QPixmap.fromImage(image.copy())

    def _load_pixmap_from_base64(self, encoded_image):
        if not encoded_image:
            return None
        try:
            image_bytes = base64.b64decode(encoded_image)
        except Exception:
            return None
        return self._load_pixmap_from_bytes(image_bytes)

    def _resolve_local_evidence_path(self, evidence_filename):
        normalized = str(evidence_filename or "").strip()
        if not normalized:
            return None
        if os.path.isabs(normalized) and os.path.exists(normalized):
            return normalized
        candidates = [
            os.path.join(EVIDENCE_DIR, normalized),
            os.path.join(os.getcwd(), "server_evidence", normalized),
        ]
        for candidate in candidates:
            if os.path.exists(candidate):
                return candidate
        return candidates[0]

    def _read_template_rows_from_csv(self, filepath):
        with open(filepath, "r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            return list(reader)

    def _read_template_rows_from_excel(self, filepath):
        workbook = load_workbook(filepath, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value or "").strip() for value in rows[0]]
        data_rows = []
        for raw_row in rows[1:]:
            if raw_row is None:
                continue
            row_dict = {headers[index]: raw_row[index] for index in range(min(len(headers), len(raw_row)))}
            data_rows.append(row_dict)
        return data_rows

    def _normalize_question_import_rows(self, rows):
        field_aliases = {
            "question_text": {"questiontext", "question", "cauhoi", "cauhoi", "noidung", "noidungcauhoi"},
            "option_a": {"optiona", "dapan a", "dapana", "a"},
            "option_b": {"optionb", "dapanb", "b"},
            "option_c": {"optionc", "dapanc", "c"},
            "option_d": {"optiond", "dapand", "d"},
            "correct_option": {"correctoption", "correctanswer", "dapandung", "dapan", "answer", "correct"},
            "points": {"points", "score", "diem"},
        }
        normalized_rows = []
        for source_row in rows:
            normalized_source = {_normalize_import_key(key): value for key, value in source_row.items()}
            row = {}
            for target_key, aliases in field_aliases.items():
                row[target_key] = ""
                for alias in aliases:
                    if alias in normalized_source and normalized_source[alias] is not None:
                        row[target_key] = str(normalized_source[alias]).strip()
                        break
            if not row["question_text"]:
                continue
            if not row["option_a"] or not row["option_b"]:
                raise ValueError(f"Câu hỏi '{row['question_text'][:60]}' thiếu ít nhất đáp án A hoặc B.")
            correct_option = row["correct_option"].upper() if row["correct_option"] else "A"
            if correct_option not in {"A", "B", "C", "D"}:
                raise ValueError(f"Đáp án đúng không hợp lệ cho câu hỏi '{row['question_text'][:60]}'.")
            try:
                points = float(row["points"] or 1.0)
            except ValueError as exc:
                raise ValueError(f"Điểm số không hợp lệ cho câu hỏi '{row['question_text'][:60]}'.") from exc
            normalized_rows.append(
                {
                    "question_text": row["question_text"],
                    "option_a": row["option_a"],
                    "option_b": row["option_b"],
                    "option_c": row["option_c"] or None,
                    "option_d": row["option_d"] or None,
                    "correct_option": correct_option,
                    "points": points,
                }
            )
        return normalized_rows

    def _load_template_questions(self, template_id):
        db = self.get_db(); cur = db.cursor(dictionary=True)
        try:
            cur.execute(
                "SELECT q_id, question_text, option_a, option_b, option_c, option_d, correct_option, points FROM question_bank WHERE template_id = %s ORDER BY q_id",
                (template_id,),
            )
            return cur.fetchall()
        finally:
            db.close()

    def _populate_template_questions(self, rows, template_label=""):
        self.tb_template_questions.setRowCount(len(rows))
        for index, row in enumerate(rows):
            self.tb_template_questions.setItem(index, 0, QTableWidgetItem(str(row.get("q_id", ""))))
            self.tb_template_questions.setItem(index, 1, QTableWidgetItem(str(row.get("question_text", ""))))
            self.tb_template_questions.setItem(index, 2, QTableWidgetItem(str(row.get("option_a", "") or "")))
            self.tb_template_questions.setItem(index, 3, QTableWidgetItem(str(row.get("option_b", "") or "")))
            self.tb_template_questions.setItem(index, 4, QTableWidgetItem(str(row.get("option_c", "") or "")))
            self.tb_template_questions.setItem(index, 5, QTableWidgetItem(str(row.get("option_d", "") or "")))
            self.tb_template_questions.setItem(index, 6, QTableWidgetItem(str(row.get("correct_option", ""))))
            self.tb_template_questions.setItem(index, 7, QTableWidgetItem(str(row.get("points", ""))))
        question_count = len(rows)
        if template_label:
            self.lbl_template_detail.setText(f"{template_label} | {question_count} câu hỏi")
        else:
            self.lbl_template_detail.setText(f"{question_count} câu hỏi")
        self.template_detail_empty.setVisible(question_count == 0)
        self.tb_template_questions.setVisible(question_count > 0)

    def refresh_template_details(self):
        template_id = self._selected_template_id()
        if template_id is None:
            self.lbl_template_detail.setText("Chưa chọn bộ đề")
            self.template_detail_empty.setVisible(True)
            self.tb_template_questions.setVisible(False)
            self.tb_template_questions.setRowCount(0)
            return
        rows = self._load_template_questions(template_id)
        self._populate_template_questions(rows, self.cb_templates_for_question.currentText())

    def select_template_from_catalog(self):
        row = self.tb_templates.currentRow()
        if row < 0:
            return
        item = self.tb_templates.item(row, 0)
        if item is None:
            return
        template_id = int(item.text())
        self._set_active_template(template_id)
        self.refresh_template_details()
        self._clear_question_form()

    def init_ui(self):
        self.stack = QStackedWidget()
        self.setCentralWidget(self.stack)

        # ================= PAGE 1: LOGIN =================
        self.page_login = QWidget()
        l_layout = QVBoxLayout(self.page_login)
        set_page_margins(self.page_login)
        login_box = QGroupBox("ĐĂNG NHẬP HỆ THỐNG GIÁM THỊ")
        login_box.setFixedSize(400, 250)
        box_layout = QVBoxLayout(login_box)
        
        self.inp_pid = QLineEdit(); self.inp_pid.setPlaceholderText("Mã Giám Thị (Ví dụ: GV01)")
        self.inp_ppw = QLineEdit(); self.inp_ppw.setPlaceholderText("Mật Khẩu"); self.inp_ppw.setEchoMode(QLineEdit.Password)
        self.btn_login = QPushButton("ĐĂNG NHẬP")
        btn_recover = QPushButton("YÊU CẦU CẤP LẠI MẬT KHẨU")
        btn_recover.setStyleSheet("background-color: #2F2F2F;")
        self.btn_login.setStyleSheet("background-color: #111111;")
        self.btn_login.clicked.connect(self.process_login)
        btn_recover.clicked.connect(self.request_proctor_password_reset)
        self.inp_pid.returnPressed.connect(self.process_login)
        self.inp_ppw.returnPressed.connect(self.process_login)
        
        box_layout.addStretch(); box_layout.addWidget(self.inp_pid); box_layout.addWidget(self.inp_ppw)
        box_layout.addWidget(self.btn_login); box_layout.addWidget(btn_recover); box_layout.addStretch()
        login_logo = build_login_logo_label(role="proctor")
        
        center_l = QHBoxLayout(); center_l.addStretch(); center_l.addWidget(login_box); center_l.addStretch()
        l_layout.addStretch(); l_layout.addWidget(login_logo, 0, Qt.AlignCenter); l_layout.addSpacing(6); l_layout.addLayout(center_l); l_layout.addStretch()
        self.stack.addWidget(self.page_login)

        # ================= PAGE 2: DASHBOARD =================
        self.page_dashboard = QWidget()
        dash_layout = QHBoxLayout(self.page_dashboard)
        set_page_margins(self.page_dashboard)
        self.sidebar = SidebarNav(
            "S-MONITOR",
            "Không gian làm việc cho giám thị: tổ chức ca thi, theo dõi vi phạm và chấm điểm.",
            ["Tổng quan", "Lớp thi", "Bộ đề", "Snapshot giám sát", "Báo cáo", "Bảng điểm"],
            role="proctor",
        )
        dash_layout.addWidget(self.sidebar)

        content_layout = QVBoxLayout()
        self.toolbar = AppToolbar(
            "Tổng quan",
            "Theo dõi năng lực coi thi, số lớp, số sinh viên và cảnh báo vi phạm trong phạm vi phụ trách.",
            role="proctor",
        )
        self.lbl_welcome = self.toolbar.title_label
        self.toolbar.badge.setText("Exam ops")
        content_layout.addWidget(self.toolbar)
        
        # Tabs Khởi tạo
        self.tabs = QTabWidget()
        self.tabs.tabBar().hide()
        self.tabs.setStyleSheet("QTabBar::tab { padding: 12px 20px; font-weight: bold; font-size: 14px; }")
        
        self.setup_tab_overview()     # Tab 1: Tổng quan
        self.setup_tab_create_class() # Tab 2: Quản lý Lớp thi
        self.setup_tab_create_quiz()  # Tab 3: Quản lý Bộ đề
        self.setup_tab_live_monitor() # Tab 4: Snapshot giám sát
        self.setup_tab_reports()      # Tab 5: Báo cáo gian lận
        self.setup_tab_scores()       # Tab 6: Quản lý Bảng điểm

        connect_sidebar_to_tabs(
            self.sidebar,
            self.tabs,
            toolbar=self.toolbar,
            section_descriptions=[
                "Theo dõi năng lực coi thi, số lớp, số sinh viên và cảnh báo vi phạm trong phạm vi phụ trách.",
                "Tạo ca thi, gán bộ đề và phân danh sách sinh viên dự thi.",
                "Biên soạn ngân hàng câu hỏi và xuất bộ đề phục vụ vận hành.",
                "Theo dõi snapshot gần nhất, tín hiệu AI và tiến độ làm bài của từng sinh viên đang thi.",
                "Xem bằng chứng vi phạm và rà soát các cảnh báo cần xử lý.",
                "Xem kết quả thi và xuất bảng điểm phục vụ tổng hợp học vụ.",
            ],
        )

        content_layout.addWidget(self.tabs, 1)
        dash_layout.addLayout(content_layout, 1)
        self.stack.addWidget(self.page_dashboard)

    # --- TAB 1: THỐNG KÊ TỔNG QUAN ---
    def setup_tab_overview(self):
        tab = QWidget(); layout = QVBoxLayout(tab)
        lbl_title = QLabel("BÁO CÁO HOẠT ĐỘNG CÁ NHÂN")
        lbl_title.setStyleSheet("font-size: 18px; font-weight: bold; color: #111111; margin-bottom: 10px;")
        layout.addWidget(lbl_title)
        
        stats_layout = QHBoxLayout()
        self.lbl_my_templates = QLabel("Bộ đề của tôi\n0"); self.format_stat_label(self.lbl_my_templates, "#8E44AD")
        self.lbl_my_classes = QLabel("Lớp thi quản lý\n0"); self.format_stat_label(self.lbl_my_classes, "#2980B9")
        self.lbl_my_students = QLabel("Tổng Sinh viên\n0"); self.format_stat_label(self.lbl_my_students, "#27AE60")
        self.lbl_my_violations = QLabel("Cảnh báo Gian lận\n0"); self.format_stat_label(self.lbl_my_violations, "#E74C3C")
        
        stats_layout.addWidget(self.lbl_my_templates); stats_layout.addWidget(self.lbl_my_classes)
        stats_layout.addWidget(self.lbl_my_students); stats_layout.addWidget(self.lbl_my_violations)
        
        btn_refresh = QPushButton("LÀM MỚI SỐ LIỆU")
        btn_refresh.setStyleSheet("background-color: #202020; padding: 15px;")
        btn_refresh.clicked.connect(self.refresh_proctor_stats)
        
        layout.addLayout(stats_layout); layout.addWidget(btn_refresh); layout.addStretch()
        self.tabs.addTab(tab, "📊 Tổng quan")

    def format_stat_label(self, lbl, color):
        style_stat_label(lbl, color)
        lbl.setAlignment(Qt.AlignCenter)

    def refresh_proctor_stats(self):
        if not self.proctor_id: return
        try:
            db = self.get_db(); cur = db.cursor()
            cur.execute("SELECT COUNT(*) FROM exam_templates WHERE proctor_id = %s", (self.proctor_id,))
            self.lbl_my_templates.setText(f"Bộ đề của tôi\n{cur.fetchone()[0]}")
            
            cur.execute("SELECT COUNT(*) FROM classes c JOIN exam_templates et ON c.template_id = et.template_id WHERE et.proctor_id = %s", (self.proctor_id,))
            self.lbl_my_classes.setText(f"Lớp thi quản lý\n{cur.fetchone()[0]}")
            
            cur.execute("SELECT COUNT(DISTINCT cs.msv) FROM class_students cs JOIN classes c ON cs.class_id = c.class_id JOIN exam_templates et ON c.template_id = et.template_id WHERE et.proctor_id = %s", (self.proctor_id,))
            self.lbl_my_students.setText(f"Tổng Sinh viên\n{cur.fetchone()[0]}")
            
            cur.execute("SELECT COUNT(*) FROM violations v JOIN classes c ON v.exam_id = c.class_id JOIN exam_templates et ON c.template_id = et.template_id WHERE et.proctor_id = %s", (self.proctor_id,))
            self.lbl_my_violations.setText(f"Cảnh báo Gian lận\n{cur.fetchone()[0]}")
            db.close()
        except Exception:
            logger.exception("Failed to refresh proctor stats for %s", self.proctor_id)

    # --- TAB 2: TẠO LỚP THI ---
    def setup_tab_create_class(self):
        tab = QWidget(); layout = QHBoxLayout(tab)
        form_gb = QGroupBox("Khởi tạo Ca thi mới")
        form_gb.setFixedWidth(500); fl = QFormLayout(form_gb)
        
        self.inp_c_id = QLineEdit(); self.inp_c_id.setPlaceholderText("Ví dụ: 102")
        self.inp_c_name = QLineEdit(); self.inp_c_name.setPlaceholderText("Ví dụ: Thi Giữa Kỳ AI")
        self.inp_c_pass = QLineEdit(); self.inp_c_pass.setPlaceholderText("Mật khẩu vào phòng")
        self.inp_c_time = QSpinBox(); self.inp_c_time.setRange(15, 180); self.inp_c_time.setValue(60); self.inp_c_time.setSuffix(" Phút")
        
        self.cb_templates_for_class = QComboBox() 
        self.inp_students = QTextEdit()
        self.inp_students.setPlaceholderText("Nhập Mã SV được thi, cách nhau bằng dấu phẩy (Ví dụ: 2200404, 2200508)")
        self.inp_students.setFixedHeight(100)
        
        fl.addRow("ID Lớp thi:", self.inp_c_id)
        fl.addRow("Tên Môn thi:", self.inp_c_name)
        fl.addRow("Mật khẩu Lớp:", self.inp_c_pass)
        fl.addRow("Thời gian (Phút):", self.inp_c_time)
        fl.addRow("Chọn Bộ Đề:", self.cb_templates_for_class)
        fl.addRow("Danh sách MSV:", self.inp_students)
        
        btn_create = QPushButton("TẠO LỚP & CHỐT ĐỀ THI")
        btn_create.setStyleSheet("background-color: #111111; padding: 15px;")
        btn_create.clicked.connect(self.create_class)
        fl.addRow(btn_create)
        
        layout.addWidget(form_gb); layout.addStretch()
        self.tabs.addTab(tab, "🏢 Lớp Thi")

    # --- TAB 3: QUẢN LÝ BỘ ĐỀ & CHẤM ĐIỂM ---
    def setup_tab_create_quiz(self):
        tab = QWidget(); layout = QVBoxLayout(tab)
        workspace_tabs = QTabWidget()
        workspace_tabs.setDocumentMode(True)
        workspace_tabs.setStyleSheet("QTabBar::tab { padding: 10px 16px; }")
        
        # Phần 1: Tạo & Xuất Bộ Đề
        left_layout = QVBoxLayout()
        gb_new_template = QGroupBox("1. Khởi tạo Bộ đề mới")
        fl_template = QVBoxLayout(gb_new_template)
        self.inp_template_name = QLineEdit(); self.inp_template_name.setPlaceholderText("Tên Bộ đề (VD: Đề CV Mã 01)")
        btn_create_template = QPushButton("TẠO BỘ ĐỀ")
        btn_create_template.setStyleSheet("background-color: #111111;")
        btn_create_template.clicked.connect(self.create_exam_template)
        fl_template.addWidget(self.inp_template_name); fl_template.addWidget(btn_create_template); fl_template.addStretch()
        
        gb_export = QGroupBox("Tiện ích Bộ đề")
        fl_export = QVBoxLayout(gb_export)
        btn_export = QPushButton("💾 XUẤT BỘ ĐỀ RA FILE CSV")
        btn_export.setStyleSheet("background-color: #2D2D2D;")
        btn_export.clicked.connect(self.export_template_csv)
        btn_import = QPushButton("📥 NHẬP CÂU HỎI TỪ EXCEL/CSV")
        btn_import.setStyleSheet("background-color: #404040;")
        btn_import.clicked.connect(self.import_template_file)
        fl_export.addWidget(btn_export)
        fl_export.addWidget(btn_import)

        gb_template_list = QGroupBox("Danh sách Bộ đề")
        template_list_layout = QVBoxLayout(gb_template_list)
        self.tb_templates = QTableWidget()
        self.tb_templates.setColumnCount(3)
        self.tb_templates.setHorizontalHeaderLabels(["ID", "Tên bộ đề", "Số câu"])
        self.tb_templates.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.tb_templates.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.tb_templates.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.tb_templates.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tb_templates.setSelectionMode(QAbstractItemView.SingleSelection)
        self.tb_templates.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tb_templates.itemSelectionChanged.connect(self.select_template_from_catalog)
        polish_table(self.tb_templates)
        self.templates_empty = EmptyState("Chưa có bộ đề", "Tạo bộ đề mới hoặc import câu hỏi từ file để bắt đầu.")
        btn_view_template = QPushButton("Xem nội dung bộ đề đã chọn")
        btn_view_template.clicked.connect(self.select_template_from_catalog)
        template_list_layout.addWidget(self.templates_empty)
        template_list_layout.addWidget(self.tb_templates)
        template_list_layout.addWidget(btn_view_template)
        
        left_layout.addWidget(gb_new_template); left_layout.addWidget(gb_export); left_layout.addWidget(gb_template_list, 1)
        
        # Phần 2: Thêm câu hỏi
        gb_questions = QGroupBox("2. Thêm Câu hỏi vào Bộ đề")
        gb_questions.setMinimumWidth(600)
        fl_q = QVBoxLayout(gb_questions)
        
        self.cb_templates_for_question = QComboBox() 
        self.inp_q = QTextEdit(); self.inp_q.setPlaceholderText("Nhập nội dung câu hỏi..."); self.inp_q.setFixedHeight(80)
        self.inp_a = QLineEdit(); self.inp_a.setPlaceholderText("Đáp án A")
        self.inp_b = QLineEdit(); self.inp_b.setPlaceholderText("Đáp án B")
        self.inp_c = QLineEdit(); self.inp_c.setPlaceholderText("Đáp án C")
        self.inp_d = QLineEdit(); self.inp_d.setPlaceholderText("Đáp án D")
        
        config_layout = QHBoxLayout()
        config_layout.addWidget(QLabel("Đáp án đúng:"))
        self.inp_correct = QComboBox()
        self.inp_correct.addItems(['A', 'B', 'C', 'D'])
        self.inp_correct.setStyleSheet("padding: 5px;")
        
        config_layout.addWidget(QLabel("Điểm số:"))
        self.inp_points = QDoubleSpinBox()
        self.inp_points.setRange(0.1, 10.0); self.inp_points.setValue(1.0); self.inp_points.setSingleStep(0.25)
        self.inp_points.setStyleSheet("padding: 5px;")
        config_layout.addWidget(self.inp_correct); config_layout.addWidget(self.inp_points); config_layout.addStretch()
        
        btn_add_q = QPushButton("THÊM CÂU HỎI VÀO BỘ ĐỀ NÀY")
        btn_add_q.clicked.connect(self.add_question)

        self.lbl_template_detail = QLabel("Chọn bộ đề để xem nội dung")
        self.lbl_template_detail.setStyleSheet("font-weight:bold; color:#111111; padding-top:12px;")
        self.tb_template_questions = QTableWidget()
        self.tb_template_questions.setColumnCount(8)
        self.tb_template_questions.setHorizontalHeaderLabels(["ID", "Câu hỏi", "A", "B", "C", "D", "Đúng", "Điểm"])
        self.tb_template_questions.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.tb_template_questions.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        for section in [2, 3, 4, 5]:
            self.tb_template_questions.horizontalHeader().setSectionResizeMode(section, QHeaderView.Stretch)
        self.tb_template_questions.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeToContents)
        self.tb_template_questions.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeToContents)
        self.tb_template_questions.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tb_template_questions.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tb_template_questions.setSelectionMode(QAbstractItemView.SingleSelection)
        self.tb_template_questions.itemSelectionChanged.connect(self._load_question_into_form)
        polish_table(self.tb_template_questions)
        self.template_detail_empty = EmptyState("Chưa có nội dung bộ đề", "Chọn bộ đề ở danh sách bên trái để xem câu hỏi và đáp án.")
        self.btn_save_question = QPushButton("THÊM CÂU HỎI VÀO BỘ ĐỀ NÀY")
        self.btn_save_question.clicked.connect(self.save_question)
        btn_delete_question = QPushButton("XÓA CÂU HỎI ĐÃ CHỌN")
        btn_delete_question.setStyleSheet("background-color: #232323;")
        btn_delete_question.clicked.connect(self.delete_selected_question)
        question_actions = QHBoxLayout()
        question_actions.addWidget(self.btn_save_question)
        question_actions.addWidget(btn_delete_question)
        
        fl_q.addWidget(QLabel("Chọn Bộ Đề đang soạn thảo:")); fl_q.addWidget(self.cb_templates_for_question)
        fl_q.addWidget(self.inp_q); fl_q.addWidget(self.inp_a); fl_q.addWidget(self.inp_b); fl_q.addWidget(self.inp_c); fl_q.addWidget(self.inp_d)
        fl_q.addLayout(config_layout); fl_q.addLayout(question_actions)
        fl_q.addWidget(self.lbl_template_detail)
        fl_q.addWidget(self.template_detail_empty)
        fl_q.addWidget(self.tb_template_questions, 1)
        
        template_page = QWidget()
        template_page_layout = QVBoxLayout(template_page)
        template_page_layout.addWidget(gb_new_template)
        template_page_layout.addWidget(gb_export)
        template_page_layout.addWidget(gb_template_list, 1)

        question_page = QWidget()
        question_page_layout = QVBoxLayout(question_page)
        question_page_layout.addWidget(gb_questions, 1)

        workspace_tabs.addTab(template_page, "Quản lý bộ đề")
        workspace_tabs.addTab(question_page, "Soạn câu hỏi")
        layout.addWidget(workspace_tabs)
        self.tabs.addTab(tab, "📝 Bộ đề thi")

    # --- TAB 4: BÁO CÁO GIAN LẬN & XEM ẢNH ---
    def setup_tab_reports(self):
        tab = QWidget(); layout = QVBoxLayout(tab)
        self.reports_filter_bar = FilterBar("Lọc theo MSV, mã lớp hoặc lỗi vi phạm", "Tìm kiếm", "Tải lại")
        wire_filter_bar(self.reports_filter_bar, self.search_reports, self.load_reports)
        reports_filter_row = QHBoxLayout()
        reports_filter_row.addWidget(self.reports_filter_bar)
        self.cb_reports_class_filter = QComboBox()
        self.cb_reports_class_filter.addItem("Tất cả phòng", "")
        self.cb_reports_class_filter.setMinimumWidth(180)
        self.cb_reports_class_filter.currentIndexChanged.connect(self.on_reports_class_filter_changed)
        reports_filter_row.addWidget(self.cb_reports_class_filter)
        self.cb_reports_status_filter = QComboBox()
        self.cb_reports_status_filter.addItem("Tất cả trạng thái", "")
        self.cb_reports_status_filter.addItem("CONFIRM", "confirm")
        self.cb_reports_status_filter.addItem("REJECT", "reject")
        self.cb_reports_status_filter.addItem("PENDING", "pending")
        self.cb_reports_status_filter.setMinimumWidth(180)
        self.cb_reports_status_filter.currentIndexChanged.connect(self.on_reports_status_filter_changed)
        reports_filter_row.addWidget(self.cb_reports_status_filter)
        
        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(["MSV", "ID Lớp", "Thời gian vi phạm", "Lỗi vi phạm", "Tên file bằng chứng", "Duyệt DB", "Event Clip"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setAlternatingRowColors(True)
        self.table.itemDoubleClicked.connect(self.view_evidence) # Double Click xem ảnh
        polish_table(self.table)
        self.reports_empty = EmptyState("Chưa có báo cáo vi phạm", "Khi AI phát hiện hành vi nghi vấn, bằng chứng sẽ hiển thị tại đây.")
        
        lbl_hint = QLabel("💡 Nhấp đúp chuột vào một dòng để xem ảnh chụp bằng chứng gian lận.")
        lbl_hint.setStyleSheet("color: #5A5A5A; font-style: italic;")
        
        layout.addLayout(reports_filter_row); layout.addWidget(lbl_hint); layout.addWidget(self.reports_empty); layout.addWidget(self.table)
        self.tabs.addTab(tab, "🚨 Báo cáo Gian lận")

    def setup_tab_live_monitor(self):
        tab = QWidget(); layout = QHBoxLayout(tab)
        left_panel = QVBoxLayout()
        self.monitor_filter_bar = FilterBar("Lọc theo MSV, tên sinh viên hoặc lớp thi", "Tìm kiếm", "Tải lại")
        wire_filter_bar(self.monitor_filter_bar, self.apply_monitor_filter, self.load_monitor_overview)
        self.tb_monitor = QTableWidget()
        self.tb_monitor.setColumnCount(9)
        self.tb_monitor.setHorizontalHeaderLabels(["MSV", "Họ tên", "Lớp", "Risk", "Cảnh báo phiên", "Đúng/Sai", "Điểm tạm", "Snapshot", "Ảnh lúc"])
        self.tb_monitor.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.tb_monitor.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.tb_monitor.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        for section in [3, 4, 5, 6, 7, 8]:
            self.tb_monitor.horizontalHeader().setSectionResizeMode(section, QHeaderView.ResizeToContents)
        self.tb_monitor.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tb_monitor.setSelectionMode(QAbstractItemView.SingleSelection)
        self.tb_monitor.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tb_monitor.itemSelectionChanged.connect(self.load_selected_monitor_detail)
        polish_table(self.tb_monitor)
        self.monitor_empty = EmptyState("Chưa có phiên giám sát", "Khi sinh viên vào phòng thi và hoàn tất xác minh, snapshot giám sát sẽ xuất hiện tại đây.")
        left_panel.addWidget(self.monitor_filter_bar)
        left_panel.addWidget(self.monitor_empty)
        left_panel.addWidget(self.tb_monitor, 1)

        detail_box = QGroupBox("Snapshot giám sát sinh viên")
        detail_layout = QVBoxLayout(detail_box)
        self.lbl_monitor_preview = QLabel("Chưa có khung hình gần nhất")
        self.lbl_monitor_preview.setAlignment(Qt.AlignCenter)
        self.lbl_monitor_preview.setMinimumSize(620, 360)
        self.lbl_monitor_preview.setStyleSheet("background:#111111; color:#E8E8E8; border-radius:14px; border:1px solid #2A2A2A;")
        self.btn_request_snapshot = QPushButton("YÊU CẦU SNAPSHOT")
        self.btn_request_snapshot.setEnabled(False)
        self.btn_request_snapshot.clicked.connect(self.request_snapshot_for_selected_session)
        self.inp_preview_review_note = QLineEdit()
        self.inp_preview_review_note.setPlaceholderText("Ghi chú đánh giá preview (tùy chọn)")
        self.btn_preview_confirm = QPushButton("XÁC NHẬN PREVIEW")
        self.btn_preview_confirm.setEnabled(False)
        self.btn_preview_confirm.clicked.connect(lambda: self.submit_preview_review("confirm"))
        self.btn_preview_reject = QPushButton("TỪ CHỐI PREVIEW")
        self.btn_preview_reject.setEnabled(False)
        self.btn_preview_reject.clicked.connect(lambda: self.submit_preview_review("reject"))
        self.btn_close_class = QPushButton("KHÓA PHÒNG THI")
        self.btn_close_class.setEnabled(False)
        self.btn_close_class.setStyleSheet("background-color: #111111; padding: 6px 12px;")
        self.btn_close_class.clicked.connect(self.close_selected_class)
        self.btn_open_class = QPushButton("MỞ LẠI PHÒNG THI")
        self.btn_open_class.setEnabled(False)
        self.btn_open_class.setStyleSheet("background-color: #1E3A5F; color: white; padding: 6px 12px;")
        self.btn_open_class.clicked.connect(self.open_selected_class)
        self.lbl_manual_review_badge = QLabel("Không có case cần review thủ công")
        self.lbl_manual_review_badge.setWordWrap(True)
        self.lbl_manual_review_badge.setStyleSheet("background:#EEF4FF; color:#234A84; border:1px solid #9EB9E5; border-radius:8px; padding:8px 10px; font-size:13px; font-weight:600;")
        self.lbl_manual_review_detail = QLabel("Các vi phạm nghi vấn từ tầng 2 hoặc xác minh danh tính sẽ được đánh dấu tại đây để giám thị duyệt thủ công.")
        self.lbl_manual_review_detail.setWordWrap(True)
        self.lbl_manual_review_detail.setStyleSheet("color:#4B5563; font-size:12px;")
        self.lbl_monitor_student = QLabel("Sinh viên: --")
        self.lbl_monitor_class = QLabel("Lớp thi: --")
        self.lbl_monitor_status = QLabel("Trạng thái AI: --")
        self.lbl_monitor_exam = QLabel("Bài làm: --")
        self.lbl_monitor_score = QLabel("Điểm tạm: --")
        self.lbl_monitor_warnings = QLabel("Cảnh báo: --")
        for info_label in [self.lbl_monitor_student, self.lbl_monitor_class, self.lbl_monitor_status, self.lbl_monitor_exam, self.lbl_monitor_score, self.lbl_monitor_warnings]:
            info_label.setStyleSheet("color:#202020; font-size:13px;")
        self.txt_monitor_predictions = QTextEdit()
        self.txt_monitor_predictions.setReadOnly(True)
        self.txt_monitor_predictions.setFixedHeight(120)
        self.txt_monitor_predictions.setPlaceholderText("Tóm tắt snapshot và sự kiện AI sẽ hiển thị ở đây.")
        detail_layout.addWidget(self.lbl_monitor_preview)
        action_row = QHBoxLayout()
        action_row.addWidget(self.btn_request_snapshot)
        action_row.addWidget(self.btn_preview_confirm)
        action_row.addWidget(self.btn_preview_reject)
        action_row.addWidget(self.btn_close_class)
        action_row.addWidget(self.btn_open_class)
        action_row.addStretch()
        detail_layout.addLayout(action_row)
        detail_layout.addWidget(self.inp_preview_review_note)
        detail_layout.addWidget(self.lbl_manual_review_badge)
        detail_layout.addWidget(self.lbl_manual_review_detail)
        detail_layout.addWidget(self.lbl_monitor_student)
        detail_layout.addWidget(self.lbl_monitor_class)
        detail_layout.addWidget(self.lbl_monitor_status)
        detail_layout.addWidget(self.lbl_monitor_exam)
        detail_layout.addWidget(self.lbl_monitor_score)
        detail_layout.addWidget(self.lbl_monitor_warnings)

        bar_specs = [
            ("Nguy cơ điện thoại", "pb_monitor_phone"),
            ("Nguy cơ nhiều người", "pb_monitor_people"),
            ("Nguy cơ mất tập trung", "pb_monitor_away"),
            ("Nguy cơ người lạ", "pb_monitor_intruder"),
        ]
        for label_text, attr_name in bar_specs:
            detail_layout.addWidget(QLabel(label_text))
            progress_bar = QProgressBar()
            progress_bar.setRange(0, 100)
            progress_bar.setValue(0)
            progress_bar.setFormat("%p%")
            setattr(self, attr_name, progress_bar)
            detail_layout.addWidget(progress_bar)
        detail_layout.addWidget(QLabel("Tóm tắt dự đoán và sự kiện"))
        detail_layout.addWidget(self.txt_monitor_predictions)

        layout.addLayout(left_panel, 3)
        layout.addWidget(detail_box, 4)
        self.tabs.addTab(tab, "🛰 Snapshot giám sát")

    # --- TAB 5: QUẢN LÝ BẢNG ĐIỂM ---
    def setup_tab_scores(self):
        tab = QWidget(); layout = QVBoxLayout(tab)
        
        self.scores_filter_bar = FilterBar("Lọc theo MSV, môn thi hoặc mã lớp", "Tìm kiếm", "Tải lại")
        wire_filter_bar(self.scores_filter_bar, self.search_scores, self.load_scores)
        btn_export = QPushButton("Xuất bảng điểm")
        btn_export.clicked.connect(self.export_scores_csv)
        layout.addWidget(self.scores_filter_bar)
        layout.addWidget(btn_export, 0, Qt.AlignLeft)
        
        self.tb_scores = QTableWidget()
        self.tb_scores.setColumnCount(6)
        self.tb_scores.setHorizontalHeaderLabels(["MSV", "Họ tên Sinh viên", "ID Lớp", "Môn thi", "Điểm số", "Thời gian nộp"])
        self.tb_scores.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tb_scores.setAlternatingRowColors(True)
        self.tb_scores.setStyleSheet("QTableWidget { font-size: 14px; } QTableWidget::item:column(4) { font-weight: bold; color: #111111; }")
        polish_table(self.tb_scores)
        self.scores_empty = EmptyState("Chưa có bảng điểm", "Kết quả nộp bài của sinh viên sẽ được tổng hợp tại đây theo từng lớp thi.")
        
        layout.addWidget(self.scores_empty)
        layout.addWidget(self.tb_scores)
        self.tabs.addTab(tab, "📈 Bảng điểm")

    # ================= LOGIC XỬ LÝ DATABASE & API =================
    def process_login(self):
        pid, pw = self.inp_pid.text().strip(), self.inp_ppw.text().strip()
        if self.proctor_login_inflight:
            return
        if not pid or not pw:
            QMessageBox.warning(self, "Lỗi", "Vui lòng nhập ID và mật khẩu Giám thị.")
            return
        self._set_login_busy(True)
        self._run_background_task(
            lambda: self._fetch_proctor_login_bundle(pid, pw),
            self._on_proctor_login_finished,
            on_error=self._on_proctor_login_error,
            on_finished=lambda: self._set_login_busy(False),
        )

    def _populate_monitor_table(self, rows):
        self.tb_monitor.setRowCount(len(rows))
        self.monitor_empty.setVisible(len(rows) == 0)
        self.tb_monitor.setVisible(len(rows) > 0)
        for row_index, row in enumerate(rows):
            snapshot_status = str(row.get("snapshot_status") or "missing")
            if bool(row.get("manual_review_required", False)):
                snapshot_text = "Can duyet"
            elif snapshot_status == "pending":
                snapshot_text = "Dang cho"
            elif snapshot_status == "available":
                snapshot_text = "Da co"
            else:
                snapshot_text = "Chua co"
            msv_item = QTableWidgetItem(str(row.get("msv", "")))
            msv_item.setData(Qt.UserRole, row.get("session_token"))
            self.tb_monitor.setItem(row_index, 0, msv_item)
            self.tb_monitor.setItem(row_index, 1, QTableWidgetItem(str(row.get("full_name", ""))))
            self.tb_monitor.setItem(row_index, 2, QTableWidgetItem(f"{row.get('class_id', '')} | {row.get('class_name', '')}"))
            self.tb_monitor.setItem(row_index, 3, QTableWidgetItem(f"{row.get('risk_score', 0):.0f}%"))
            self.tb_monitor.setItem(row_index, 4, QTableWidgetItem(f"{row.get('session_warning_count', 0)}/{row.get('max_warnings', 0)}"))
            self.tb_monitor.setItem(row_index, 5, QTableWidgetItem(f"{row.get('correct_count', 0)}/{row.get('wrong_count', 0)}"))
            self.tb_monitor.setItem(row_index, 6, QTableWidgetItem(f"{row.get('current_score', 0.0):.2f}"))
            self.tb_monitor.setItem(row_index, 7, QTableWidgetItem(snapshot_text))
            self.tb_monitor.setItem(row_index, 8, QTableWidgetItem(str(row.get("snapshot_updated_at") or row.get("updated_at", ""))))

    def apply_monitor_filter(self):
        keyword = self.monitor_filter_bar.search_input.text().strip().lower()
        if not keyword:
            return self._populate_monitor_table(self.monitor_entries)
        filtered_rows = [
            row for row in self.monitor_entries
            if keyword in str(row.get("msv", "")).lower()
            or keyword in str(row.get("full_name", "")).lower()
            or keyword in str(row.get("class_name", "")).lower()
            or keyword in str(row.get("class_id", "")).lower()
        ]
        self._populate_monitor_table(filtered_rows)

    def load_monitor_overview(self):
        if not self.proctor_id:
            return
        if self.monitor_overview_inflight:
            return
        self.monitor_overview_inflight = True
        self._run_background_task(
            self._fetch_proctor_monitor_overview,
            self._on_proctor_monitor_overview_loaded,
            on_error=self._on_proctor_monitor_overview_error,
            on_finished=lambda: setattr(self, "monitor_overview_inflight", False),
        )

    def _set_monitor_risk_bar(self, progress_bar, detail, key):
        ratio = float((detail.get("violation_ratios") or {}).get(key, 0.0))
        progress_bar.setValue(int(max(0.0, min(100.0, ratio * 100.0))))

    def load_selected_monitor_detail(self):
        row = self.tb_monitor.currentRow()
        if row < 0:
            self.btn_request_snapshot.setEnabled(False)
            self.btn_preview_confirm.setEnabled(False)
            self.btn_preview_reject.setEnabled(False)
            self.btn_close_class.setEnabled(False)
            self.btn_open_class.setEnabled(False)
            self.current_monitor_class_id = None
            return
        item = self.tb_monitor.item(row, 0)
        if item is None:
            self.btn_request_snapshot.setEnabled(False)
            self.btn_preview_confirm.setEnabled(False)
            self.btn_preview_reject.setEnabled(False)
            self.btn_close_class.setEnabled(False)
            self.btn_open_class.setEnabled(False)
            self.current_monitor_class_id = None
            return
        session_token = item.data(Qt.UserRole)
        if not session_token:
            self.btn_request_snapshot.setEnabled(False)
            self.btn_preview_confirm.setEnabled(False)
            self.btn_preview_reject.setEnabled(False)
            self.btn_close_class.setEnabled(False)
            self.btn_open_class.setEnabled(False)
            self.current_monitor_class_id = None
            return
        self.current_monitor_session_token = session_token
        self.btn_request_snapshot.setEnabled(not self.snapshot_request_inflight)
        self.btn_close_class.setEnabled(False)
        self.btn_open_class.setEnabled(False)
        self._run_background_task(
            lambda: self._fetch_proctor_monitor_detail(session_token),
            lambda payload: self._apply_monitor_detail_payload(payload["session_token"], payload["detail"]),
            on_error=lambda payload: logger.error("Proctor monitor detail failed for %s: %s", session_token, payload.get("traceback") or payload.get("message")),
        )

    def request_snapshot_for_selected_session(self):
        session_token = self.current_monitor_session_token
        if not session_token or self.snapshot_request_inflight:
            return
        self.snapshot_request_inflight = True
        self.btn_request_snapshot.setEnabled(False)
        self.btn_request_snapshot.setText("ĐANG GỬI YÊU CẦU...")
        self._run_background_task(
            lambda: self._fetch_proctor_snapshot_request(session_token),
            self._on_snapshot_request_success,
            on_error=self._on_snapshot_request_error,
        )

    def close_selected_class(self):
        class_id = self.current_monitor_class_id
        if not class_id or self.close_class_inflight:
            return
        reply = QMessageBox.question(
            self,
            "Xác nhận khóa phòng",
            f"Khóa thủ công lớp thi {class_id}? Sinh viên trong lớp sẽ không thể tiếp tục thao tác khi hệ thống đồng bộ trạng thái.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return
        self.close_class_inflight = True
        self.btn_close_class.setEnabled(False)
        self.btn_open_class.setEnabled(False)
        self.btn_close_class.setText("ĐANG KHÓA...")
        self._run_background_task(
            lambda: self._close_proctor_class(class_id),
            self._on_close_class_success,
            on_error=self._on_close_class_error,
        )

    def open_selected_class(self):
        class_id = self.current_monitor_class_id
        if not class_id or self.close_class_inflight:
            return
        reply = QMessageBox.question(
            self,
            "Xác nhận mở lại phòng",
            f"Mở lại lớp thi {class_id}? Sinh viên trong lớp sẽ có thể tiếp tục thao tác sau lần đồng bộ tiếp theo.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return
        self.close_class_inflight = True
        self.btn_open_class.setEnabled(False)
        self.btn_close_class.setEnabled(False)
        self.btn_open_class.setText("ĐANG MỞ...")
        self._run_background_task(
            lambda: self._open_proctor_class(class_id),
            self._on_open_class_success,
            on_error=self._on_open_class_error,
        )

    def load_templates(self):
        try:
            db = self.get_db(); cur = db.cursor(dictionary=True)
            cur.execute(
                """
                SELECT et.template_id, et.template_name, COUNT(q.q_id) AS question_count
                FROM exam_templates et
                LEFT JOIN question_bank q ON q.template_id = et.template_id
                WHERE et.proctor_id = %s
                GROUP BY et.template_id, et.template_name
                ORDER BY et.template_id DESC
                """,
                (self.proctor_id,),
            )
            rows = cur.fetchall(); db.close()
            
            current_template_id = self._selected_template_id()
            self.cb_templates_for_class.clear(); self.cb_templates_for_question.clear()
            self.tb_templates.setRowCount(len(rows))
            self.template_dict.clear()
            
            for row_index, r in enumerate(rows):
                tname = f"[{r['template_id']}] {r['template_name']}"
                self.template_dict[tname] = r['template_id']
                self.cb_templates_for_class.addItem(tname)
                self.cb_templates_for_question.addItem(tname)
                self.tb_templates.setItem(row_index, 0, QTableWidgetItem(str(r['template_id'])))
                self.tb_templates.setItem(row_index, 1, QTableWidgetItem(str(r['template_name'])))
                self.tb_templates.setItem(row_index, 2, QTableWidgetItem(str(r.get('question_count', 0))))
            self.templates_empty.setVisible(len(rows) == 0)
            self.tb_templates.setVisible(len(rows) > 0)
            if current_template_id is not None:
                self._set_active_template(current_template_id)
            if self.cb_templates_for_question.count() > 0:
                if self.cb_templates_for_question.currentIndex() < 0:
                    self.cb_templates_for_question.setCurrentIndex(0)
                self.refresh_template_details()
            else:
                self.refresh_template_details()
        except Exception:
            logger.exception("Failed to load templates for %s", self.proctor_id)
            QMessageBox.critical(self, "Lỗi", "Không tải được danh sách bộ đề.")

    def create_exam_template(self):
        t_name = self.inp_template_name.text().strip()
        if not t_name: return QMessageBox.warning(self, "Lỗi", "Vui lòng nhập tên Bộ đề!")
        try:
            db = self.get_db(); cur = db.cursor()
            try:
                # Newer schemas require exam_name (NOT NULL), keep both names in sync.
                cur.execute(
                    "INSERT INTO exam_templates (proctor_id, template_name, exam_name) VALUES (%s, %s, %s)",
                    (self.proctor_id, t_name, t_name),
                )
            except Exception:
                # Backward compatibility for schemas that only have template_name.
                cur.execute("INSERT INTO exam_templates (proctor_id, template_name) VALUES (%s, %s)", (self.proctor_id, t_name))
            template_id = cur.lastrowid
            db.commit(); db.close()
            self.inp_template_name.clear()
            self.load_templates() 
            self._set_active_template(template_id)
            self.refresh_template_details()
            self.refresh_proctor_stats()
            QMessageBox.information(self, "Thành công", f"Đã tạo bộ đề: {t_name}")
        except Exception as e: QMessageBox.critical(self, "Lỗi", str(e))

    def add_question(self):
        self.save_question()

    def save_question(self):
        selected_template = self.cb_templates_for_question.currentText()
        if not selected_template: return QMessageBox.warning(self, "Lỗi", "Vui lòng chọn Bộ đề!")
        
        template_id = self.template_dict[selected_template]
        q, a, b, c, d = self.inp_q.toPlainText().strip(), self.inp_a.text().strip(), self.inp_b.text().strip(), self.inp_c.text().strip(), self.inp_d.text().strip()
        correct_ans = self.inp_correct.currentText()
        pts = self.inp_points.value()
        
        if not q or not a or not b: return QMessageBox.warning(self, "Lỗi", "Nhập câu hỏi và ít nhất 2 đáp án!")
        try:
            db = self.get_db(); cur = db.cursor()
            if self.current_question_id is None:
                query = "INSERT INTO question_bank (template_id, question_text, option_a, option_b, option_c, option_d, correct_option, points) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"
                cur.execute(query, (template_id, q, a, b, c, d, correct_ans, pts))
                success_message = f"Đã thêm câu hỏi ({pts} điểm) vào Bộ đề!"
            else:
                query = "UPDATE question_bank SET template_id=%s, question_text=%s, option_a=%s, option_b=%s, option_c=%s, option_d=%s, correct_option=%s, points=%s WHERE q_id=%s"
                cur.execute(query, (template_id, q, a, b, c, d, correct_ans, pts, self.current_question_id))
                success_message = "Đã cập nhật câu hỏi đã chọn."
            db.commit(); db.close()
            self.load_templates()
            self._set_active_template(template_id)
            self.refresh_template_details()
            QMessageBox.information(self, "Thành công", success_message)
            self._clear_question_form()
        except Exception as e: QMessageBox.critical(self, "Lỗi", str(e))

    def delete_selected_question(self):
        row = self.tb_template_questions.currentRow()
        if row < 0:
            return QMessageBox.warning(self, "Lỗi", "Hãy chọn câu hỏi cần xóa trong bảng nội dung bộ đề.")
        question_id_item = self.tb_template_questions.item(row, 0)
        if question_id_item is None:
            return QMessageBox.warning(self, "Lỗi", "Không xác định được ID câu hỏi cần xóa.")
        question_id = int(question_id_item.text())
        reply = QMessageBox.question(self, "Xác nhận xóa", f"Bạn có chắc muốn xóa câu hỏi ID {question_id}?", QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return
        try:
            db = self.get_db(); cur = db.cursor()
            cur.execute("DELETE FROM question_bank WHERE q_id = %s", (question_id,))
            db.commit(); db.close()
            self.load_templates()
            self.refresh_template_details()
            self._clear_question_form()
            QMessageBox.information(self, "Thành công", f"Đã xóa câu hỏi ID {question_id}.")
        except Exception as exc:
            logger.exception("Failed to delete question %s", question_id)
            self._show_dialog("Không thể xóa câu hỏi", str(exc))

    def import_template_file(self):
        template_id = self._selected_template_id()
        if template_id is None:
            return QMessageBox.warning(self, "Lỗi", "Vui lòng chọn bộ đề cần nhập câu hỏi.")
        filepath, _ = QFileDialog.getOpenFileName(
            self,
            "Chọn file bộ đề",
            "",
            "Excel Files (*.xlsx *.xlsm);;CSV Files (*.csv)",
        )
        if not filepath:
            return
        try:
            extension = os.path.splitext(filepath)[1].lower()
            if extension == ".csv":
                raw_rows = self._read_template_rows_from_csv(filepath)
            elif extension in {".xlsx", ".xlsm"}:
                raw_rows = self._read_template_rows_from_excel(filepath)
            else:
                raise ValueError("Chỉ hỗ trợ file .xlsx, .xlsm hoặc .csv")

            question_rows = self._normalize_question_import_rows(raw_rows)
            if not question_rows:
                raise ValueError("Không tìm thấy câu hỏi hợp lệ trong file đã chọn.")

            db = self.get_db(); cur = db.cursor()
            cur.executemany(
                "INSERT INTO question_bank (template_id, question_text, option_a, option_b, option_c, option_d, correct_option, points) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
                [
                    (
                        template_id,
                        row["question_text"],
                        row["option_a"],
                        row["option_b"],
                        row["option_c"],
                        row["option_d"],
                        row["correct_option"],
                        row["points"],
                    )
                    for row in question_rows
                ],
            )
            db.commit(); db.close()
            self.load_templates()
            self._set_active_template(template_id)
            self.refresh_template_details()
            QMessageBox.information(self, "Thành công", f"Đã nhập {len(question_rows)} câu hỏi từ file vào bộ đề.")
        except Exception as exc:
            logger.exception("Failed to import template file %s", filepath)
            self._show_dialog("Không thể nhập bộ đề", str(exc), "File cần có các cột như question_text, option_a, option_b, option_c, option_d, correct_option, points.")

    def export_template_csv(self):
        selected_template = self.cb_templates_for_question.currentText()
        if not selected_template: return QMessageBox.warning(self, "Lỗi", "Chọn 1 Bộ đề để xuất!")
        template_id = self.template_dict[selected_template]
        try:
            db = self.get_db(); cur = db.cursor(dictionary=True)
            cur.execute("SELECT question_text, option_a, option_b, option_c, option_d, correct_option, points FROM question_bank WHERE template_id = %s", (template_id,))
            rows = cur.fetchall(); db.close()
            
            if not rows: return QMessageBox.warning(self, "Trống", "Bộ đề chưa có câu hỏi!")
            filepath, _ = QFileDialog.getSaveFileName(self, "Lưu file CSV", f"Bo_De_{template_id}.csv", "CSV Files (*.csv)")
            if filepath:
                with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.DictWriter(f, fieldnames=rows[0].keys())
                    writer.writeheader()
                    writer.writerows(rows)
                QMessageBox.information(self, "Thành công", "Đã xuất câu hỏi ra file Excel thành công!")
        except Exception as e: QMessageBox.critical(self, "Lỗi", str(e))

    def create_class(self):
        cid, cname, cpass = self.inp_c_id.text().strip(), self.inp_c_name.text().strip(), self.inp_c_pass.text().strip()
        ctime = self.inp_c_time.value()
        students_raw = self.inp_students.toPlainText().strip()
        selected_template = self.cb_templates_for_class.currentText()
        
        if not cid or not cname or not cpass or not selected_template:
            return QMessageBox.warning(self, "Lỗi", "Vui lòng nhập đủ thông tin và chọn Bộ đề!")
            
        template_id = self.template_dict[selected_template]
        try:
            db = self.get_db(); cur = db.cursor()
            cur.execute("INSERT INTO classes (class_id, class_name, class_password, duration_minutes, template_id) VALUES (%s, %s, %s, %s, %s)",
                        (cid, cname, cpass, ctime, template_id))
            
            if students_raw:
                student_list = [s.strip() for s in students_raw.split(',') if s.strip()]
                failed_students = []
                for msv in student_list:
                    try:
                        cur.execute("INSERT INTO class_students (class_id, msv) VALUES (%s, %s)", (cid, msv))
                    except Exception:
                        failed_students.append(msv)
                        logger.exception("Failed to enroll student %s into class %s", msv, cid)
            db.commit(); db.close()
            self.refresh_proctor_stats()
            if students_raw and failed_students:
                failed_text = ", ".join(failed_students)
                QMessageBox.warning(self, "Tạo lớp có cảnh báo", f"Đã tạo Lớp {cid} dùng Bộ đề {template_id}, nhưng không thêm được các sinh viên: {failed_text}")
            else:
                QMessageBox.information(self, "Thành công", f"Đã tạo Lớp {cid} dùng Bộ đề {template_id}!")
        except Exception as e: QMessageBox.critical(self, "Lỗi", f"Lỗi tạo lớp!\n{e}")

    def load_reports(self):
        if not self.proctor_id: return
        if self.reports_inflight:
            return
        self.reports_inflight = True
        self._run_background_task(
            lambda: self._fetch_proctor_reports(),
            self._populate_reports_table,
            on_error=self._on_proctor_reports_error,
            on_finished=lambda: setattr(self, "reports_inflight", False),
        )

    def search_reports(self):
        keyword = self.reports_filter_bar.search_input.text().strip()
        if not keyword:
            return self.load_reports()
        if self.reports_inflight:
            return
        self.reports_inflight = True
        self._run_background_task(
            lambda: self._fetch_proctor_reports(keyword),
            self._populate_reports_table,
            on_error=self._on_proctor_reports_error,
            on_finished=lambda: setattr(self, "reports_inflight", False),
        )

    def view_evidence(self, item):
        row = item.row()
        evidence_filename = self.table.item(row, 4).text()
        clip_item = self.table.item(row, 6)
        clip_path = str(clip_item.data(Qt.UserRole) or "").strip() if clip_item else ""
        clip_duration = float(clip_item.data(Qt.UserRole + 1) or 0.0) if clip_item else 0.0
        has_clip = bool(clip_path)

        if not evidence_filename:
            return self._show_dialog("Thiếu dữ liệu", "Dòng báo cáo này không có tên tệp bằng chứng.")
        img_url = f"{self.api_url}/evidence_images/{evidence_filename}"
        local_path = self._resolve_local_evidence_path(evidence_filename)
        api_error = None

        try:
            r = requests.get(img_url, timeout=10)
            if r.status_code == 200:
                pixmap = self._load_pixmap_from_bytes(r.content)
                if pixmap is not None and not pixmap.isNull():
                    self._open_evidence_dialog(
                        pixmap,
                        row,
                        has_clip=has_clip,
                        clip_path=clip_path,
                        clip_duration=clip_duration,
                    )
                    return
                api_error = "API trả dữ liệu nhưng Qt không giải mã được ảnh."
            else:
                api_error = f"API trả về mã {r.status_code}."
        except requests.RequestException as exc:
            logger.warning("Evidence API unavailable for %s: %s", evidence_filename, exc)
            api_error = str(exc)

        if local_path and os.path.exists(local_path):
            with open(local_path, "rb") as handle:
                pixmap = self._load_pixmap_from_bytes(handle.read())
            if pixmap is not None and not pixmap.isNull():
                self._open_evidence_dialog(
                    pixmap,
                    row,
                    has_clip=has_clip,
                    clip_path=clip_path,
                    clip_duration=clip_duration,
                )
                return

        self._show_dialog(
            "Không tải được bằng chứng",
            "Không thể mở ảnh vi phạm từ API hoặc thư mục cục bộ.",
            f"Tệp cần tìm: {evidence_filename}\nĐường dẫn cục bộ: {os.path.abspath(local_path or '')}\nTrạng thái API: {api_error or 'Không có phản hồi.'}",
        )

    def _open_evidence_dialog(self, pixmap, row, has_clip: bool = False, clip_path: str = "", clip_duration: float = 0.0):
        dlg = QDialog(self)
        dlg.setWindowTitle(f"BẰNG CHỨNG VI PHẠM - {self.table.item(row, 0).text()}")
        dlg.resize(860, 680)
        dlg.setStyleSheet("QDialog { background: #FCFCFC; }")
        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(14)
        title = QLabel("Bằng chứng vi phạm")
        title.setStyleSheet("font-size: 20px; font-weight: 700; color: #111111;")
        meta = QLabel(
            f"MSV: {self.table.item(row, 0).text()} | Lớp: {self.table.item(row, 1).text()} | Lỗi: {self.table.item(row, 3).text()}"
        )
        meta.setStyleSheet("color: #5C5C5C; font-size: 12px;")
        img_label = QLabel()
        img_label.setAlignment(Qt.AlignCenter)
        img_label.setStyleSheet("background:#F4F4F4; border:1px solid #E2E2E2; border-radius:16px;")
        img_label.setPixmap(pixmap.scaled(780, 480, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        layout.addWidget(title)
        layout.addWidget(meta)
        layout.addWidget(img_label, 1)
        # Thông tin event clip
        clip_item = self.table.item(row, 6)
        clip_info_text = clip_item.text() if clip_item else ""
        review_item = self.table.item(row, 5)
        review_status = review_item.text() if review_item else "PENDING"
        review_note = str(review_item.data(Qt.UserRole) or "").strip() if review_item else ""
        review_by = str(review_item.data(Qt.UserRole + 1) or "").strip() if review_item else ""
        review_at = str(review_item.data(Qt.UserRole + 2) or "--").strip() if review_item else "--"
        review_lbl = QLabel(
            f"🧾 Duyệt DB: {review_status} | Người duyệt: {review_by or '--'} | Lúc: {review_at or '--'}"
        )
        review_lbl.setStyleSheet("color: #2A5CAA; font-size: 12px; font-weight: 600;")
        review_lbl.setWordWrap(True)
        layout.addWidget(review_lbl)
        if review_note:
            review_note_lbl = QLabel(f"Ghi chú duyệt: {review_note}")
            review_note_lbl.setStyleSheet("color: #4B5563; font-size: 11px;")
            review_note_lbl.setWordWrap(True)
            layout.addWidget(review_note_lbl)
        clip_lbl = QLabel(
            f"🎬 Event clip đã ghi: {clip_info_text}" if has_clip
            else "📷 Chỉ có ảnh đơn — chưa có event clip cho sự kiện này."
        )
        clip_lbl.setStyleSheet(
            "color: #1A7F37; font-size: 12px; font-weight: 600;" if has_clip
            else "color: #888888; font-size: 11px;"
        )
        clip_lbl.setWordWrap(True)
        layout.addWidget(clip_lbl)

        if has_clip:
            btn_open_clip = QPushButton(f"Xem clip vi phạm ({clip_duration:.1f}s)")
            btn_open_clip.setCursor(Qt.PointingHandCursor)
            btn_open_clip.clicked.connect(lambda: self._open_clip_evidence(clip_path))
            layout.addWidget(btn_open_clip)

        dlg.exec_()

    def _open_clip_evidence(self, clip_path: str):
        """Mở clip vi phạm trực tiếp trong giao diện proctor."""
        clip_path = str(clip_path or "").strip()
        if not clip_path:
            self._show_dialog("Thiếu clip", "Bản ghi này chưa có đường dẫn clip vi phạm.")
            return

        local_path = self._resolve_local_evidence_path(clip_path)
        temp_downloaded = False
        # Nếu chưa có file cục bộ, tải về file tạm để phát trong app.
        clip_url = f"{self.api_url}/evidence_images/{clip_path}"
        try:
            if not (local_path and os.path.exists(local_path)):
                resp = requests.get(clip_url, timeout=15)
                if resp.status_code != 200 or not resp.content:
                    self._show_dialog("Không tải được clip", f"API trả mã {resp.status_code} cho clip: {clip_path}")
                    return
                ext = os.path.splitext(clip_path)[1] or ".mp4"
                with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as fh:
                    fh.write(resp.content)
                    local_path = fh.name
                temp_downloaded = True

            self._open_clip_dialog(local_path, clip_path, temp_downloaded)
        except Exception as exc:
            logger.exception("Failed to open clip evidence")
            if temp_downloaded and local_path and os.path.exists(local_path):
                try:
                    os.remove(local_path)
                except OSError:
                    logger.debug("Could not cleanup temporary clip %s", local_path, exc_info=True)
            self._show_dialog("Không mở được clip", "Không thể tải/hiển thị clip vi phạm trong giao diện.", str(exc))

    def _open_clip_dialog(self, local_path: str, clip_label: str, temp_downloaded: bool = False):
        local_path = str(local_path or "").strip()
        if not local_path or not os.path.exists(local_path):
            self._show_dialog("Không mở được clip", "File clip không tồn tại trên máy.")
            return

        cap = cv2.VideoCapture(local_path)
        if not cap.isOpened():
            cap.release()
            self._show_dialog("Không mở được clip", "OpenCV không thể giải mã clip vi phạm này.")
            if temp_downloaded:
                try:
                    os.remove(local_path)
                except OSError:
                    logger.debug("Could not cleanup temporary clip %s", local_path, exc_info=True)
            return

        fps = float(cap.get(cv2.CAP_PROP_FPS) or 0.0)
        if fps <= 0.0 or np.isnan(fps):
            fps = 12.0
        interval_ms = int(max(25, min(120, 1000.0 / fps)))

        dlg = QDialog(self)
        dlg.setWindowTitle(f"CLIP VI PHẠM - {clip_label}")
        dlg.resize(920, 680)
        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(10)

        title = QLabel("Xem clip vi phạm trong giao diện giám thị")
        title.setStyleSheet("font-size: 18px; font-weight: 700; color: #111111;")
        info = QLabel(f"Nguồn clip: {clip_label}")
        info.setStyleSheet("color: #555555; font-size: 12px;")
        video_label = QLabel("Đang tải clip...")
        video_label.setAlignment(Qt.AlignCenter)
        video_label.setMinimumSize(860, 520)
        video_label.setStyleSheet("background:#0B0B0B; color:#F0F0F0; border:1px solid #2A2A2A; border-radius:12px;")

        controls = QHBoxLayout()
        btn_toggle = QPushButton("TẠM DỪNG")
        btn_restart = QPushButton("PHÁT LẠI")
        btn_close = QPushButton("ĐÓNG")
        controls.addWidget(btn_toggle)
        controls.addWidget(btn_restart)
        controls.addStretch()
        controls.addWidget(btn_close)

        layout.addWidget(title)
        layout.addWidget(info)
        layout.addWidget(video_label, 1)
        layout.addLayout(controls)

        timer = QTimer(dlg)
        state = {"playing": True}

        def render_next_frame():
            if not state["playing"]:
                return
            ok, frame = cap.read()
            if not ok:
                state["playing"] = False
                btn_toggle.setText("TIẾP TỤC")
                video_label.setText("Đã phát hết clip. Bấm PHÁT LẠI để xem lại.")
                return
            rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            image = QImage(rgb.data, rgb.shape[1], rgb.shape[0], rgb.strides[0], QImage.Format_RGB888)
            pixmap = QPixmap.fromImage(image).scaled(video_label.width(), video_label.height(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
            video_label.setPixmap(pixmap)

        def toggle_playback():
            state["playing"] = not state["playing"]
            btn_toggle.setText("TẠM DỪNG" if state["playing"] else "TIẾP TỤC")

        def restart_playback():
            cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
            state["playing"] = True
            btn_toggle.setText("TẠM DỪNG")
            video_label.setText("Đang phát lại clip...")

        cleaned = {"done": False}

        def cleanup_resources():
            if cleaned["done"]:
                return
            cleaned["done"] = True
            timer.stop()
            try:
                cap.release()
            except Exception:
                logger.debug("Failed releasing clip capture", exc_info=True)
            if temp_downloaded:
                try:
                    if os.path.exists(local_path):
                        os.remove(local_path)
                except OSError:
                    logger.debug("Could not cleanup temporary clip %s", local_path, exc_info=True)

        btn_toggle.clicked.connect(toggle_playback)
        btn_restart.clicked.connect(restart_playback)
        btn_close.clicked.connect(dlg.close)
        timer.timeout.connect(render_next_frame)
        dlg.finished.connect(lambda _=0: cleanup_resources())

        render_next_frame()
        timer.start(interval_ms)
        dlg.exec_()

    def load_scores(self):
        if not self.proctor_id: return
        try:
            db = self.get_db(); cur = db.cursor(dictionary=True)
            query = """SELECT r.msv, s.full_name, c.class_id, c.class_name, r.score, r.submission_time
                       FROM exam_results r JOIN students s ON r.msv = s.msv
                       JOIN classes c ON r.exam_id = c.class_id JOIN exam_templates et ON c.template_id = et.template_id
                       WHERE et.proctor_id = %s ORDER BY r.submission_time DESC"""
            cur.execute(query, (self.proctor_id,))
            rows = cur.fetchall(); db.close()
            
            self.tb_scores.setRowCount(len(rows))
            self.scores_empty.setVisible(len(rows) == 0)
            self.tb_scores.setVisible(len(rows) > 0)
            for i, r in enumerate(rows):
                self.tb_scores.setItem(i, 0, QTableWidgetItem(str(r['msv'])))
                self.tb_scores.setItem(i, 1, QTableWidgetItem(str(r['full_name'])))
                self.tb_scores.setItem(i, 2, QTableWidgetItem(str(r['class_id'])))
                self.tb_scores.setItem(i, 3, QTableWidgetItem(str(r['class_name'])))
                
                score_item = QTableWidgetItem(f"{r['score']} Điểm")
                score_item.setTextAlignment(Qt.AlignCenter)
                self.tb_scores.setItem(i, 4, score_item)
                self.tb_scores.setItem(i, 5, QTableWidgetItem(str(r['submission_time'])))
        except Exception:
            logger.exception("Failed to load scores for %s", self.proctor_id)
            QMessageBox.critical(self, "Lỗi", "Không tải được bảng điểm.")

    def search_scores(self):
        keyword = self.scores_filter_bar.search_input.text().strip()
        if not keyword:
            return self.load_scores()
        try:
            db = self.get_db(); cur = db.cursor(dictionary=True)
            query = """SELECT r.msv, s.full_name, c.class_id, c.class_name, r.score, r.submission_time
                       FROM exam_results r JOIN students s ON r.msv = s.msv
                       JOIN classes c ON r.exam_id = c.class_id JOIN exam_templates et ON c.template_id = et.template_id
                       WHERE et.proctor_id = %s AND (r.msv LIKE %s OR c.class_name LIKE %s OR CAST(c.class_id AS CHAR) LIKE %s)
                       ORDER BY r.submission_time DESC"""
            cur.execute(query, (self.proctor_id, f"%{keyword}%", f"%{keyword}%", f"%{keyword}%"))
            rows = cur.fetchall(); db.close()
            self.tb_scores.setRowCount(len(rows))
            self.scores_empty.setVisible(len(rows) == 0)
            self.tb_scores.setVisible(len(rows) > 0)
            for i, r in enumerate(rows):
                self.tb_scores.setItem(i, 0, QTableWidgetItem(str(r['msv'])))
                self.tb_scores.setItem(i, 1, QTableWidgetItem(str(r['full_name'])))
                self.tb_scores.setItem(i, 2, QTableWidgetItem(str(r['class_id'])))
                self.tb_scores.setItem(i, 3, QTableWidgetItem(str(r['class_name'])))
                score_item = QTableWidgetItem(f"{r['score']} Điểm")
                score_item.setTextAlignment(Qt.AlignCenter)
                self.tb_scores.setItem(i, 4, score_item)
                self.tb_scores.setItem(i, 5, QTableWidgetItem(str(r['submission_time'])))
        except Exception as exc:
            self._show_dialog("Không thể tìm bảng điểm", str(exc))

    def export_scores_csv(self):
        if not self.proctor_id: return
        try:
            db = self.get_db(); cur = db.cursor(dictionary=True)
            query = """SELECT r.msv AS 'Mã Sinh Viên', s.full_name AS 'Họ và Tên', c.class_id AS 'Mã Lớp', 
                       c.class_name AS 'Môn Thi', r.score AS 'Điểm Số', r.submission_time AS 'Thời gian nộp bài'
                       FROM exam_results r JOIN students s ON r.msv = s.msv JOIN classes c ON r.exam_id = c.class_id
                       JOIN exam_templates et ON c.template_id = et.template_id
                       WHERE et.proctor_id = %s ORDER BY c.class_id, r.score DESC"""
            cur.execute(query, (self.proctor_id,))
            rows = cur.fetchall(); db.close()
            
            if not rows: return QMessageBox.warning(self, "Trống", "Chưa có sinh viên nào nộp bài!")
            filepath, _ = QFileDialog.getSaveFileName(self, "Lưu Bảng Điểm", "Bang_Diem_Tong_Hop.csv", "CSV Files (*.csv)")
            if filepath:
                with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.DictWriter(f, fieldnames=rows[0].keys())
                    writer.writeheader()
                    writer.writerows(rows)
                QMessageBox.information(self, "Thành công", f"Đã xuất bảng điểm của {len(rows)} lượt thi ra file Excel!")
        except Exception as e: QMessageBox.critical(self, "Lỗi", str(e))

if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = ProctorDashboard(); win.show()
    sys.exit(app.exec_())