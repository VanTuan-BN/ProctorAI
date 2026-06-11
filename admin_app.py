import sys, os, shutil, csv, logging, requests, unicodedata, base64
from datetime import datetime
import cv2
import mysql.connector
import numpy as np
from openpyxl import load_workbook
from PyQt5.QtCore import Qt, QTimer, QThreadPool
from PyQt5.QtGui import QImage, QPixmap
from PyQt5.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QComboBox,
    QDialog,
    QDoubleSpinBox,
    QFileDialog,
    QFormLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QInputDialog,
    QPushButton,
    QProgressBar,
    QSpinBox,
    QStackedWidget,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)
from ui_branding import build_login_logo_label
from ui_components import AppToolbar, EmptyState, EnterpriseDialog, FilterBar, SidebarNav, connect_sidebar_to_tabs, wire_filter_bar
from qt_async import BackgroundTask
from ui_theme import apply_theme, polish_table, set_page_margins, style_stat_label, style_status_pill
from auth_security import hash_password, verify_and_upgrade_password
from password_recovery_dialog import PasswordRecoveryDialog

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s: %(message)s")
logger = logging.getLogger(__name__)

DEFAULT_FACE_THRESHOLD = 0.75
MIN_FACE_THRESHOLD = 0.70
MAX_FACE_THRESHOLD = 0.99
ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
EVIDENCE_DIR = os.path.join(ROOT_DIR, "server_evidence")
DB_HOST = os.getenv("S_MONITOR_DB_HOST", "127.0.0.1")
DB_USER = os.getenv("S_MONITOR_DB_USER", "root")
DB_PASSWORD = os.getenv("S_MONITOR_DB_PASSWORD", "12345")
DB_NAME = os.getenv("S_MONITOR_DB_NAME", "exam_monitor_db")


def _normalize_import_key(value):
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    return "".join(ch for ch in text.lower() if ch.isalnum())


def _password_role_mapping(role):
    mapping = {
        "student": ("students", "msv"),
        "proctor": ("proctors", "proctor_id"),
        "admin": ("admins", "admin_id"),
    }
    return mapping[str(role or "").strip().lower()]

class AdminDashboard(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("S-MONITOR: HỆ THỐNG QUẢN TRỊ CẤP CAO (SUPER ADMIN)")
        self.setMinimumSize(1280, 800)
        apply_theme(self, role="admin")
        
        self.admin_name = ""
        self.admin_token = None
        self.api_url = "http://127.0.0.1:8000"
        self.template_dict = {}
        self.admin_current_question_id = None
        self.monitor_entries = []
        self.current_monitor_session_token = None
        self.selected_student_face_ref_id = None
        self.selected_student_face_image = ""
        self.selected_student_face_targets = []
        self.thread_pool = QThreadPool.globalInstance()
        self.admin_login_inflight = False
        self.monitor_overview_inflight = False
        self.reports_inflight = False
        self.init_ui()

        # Timer cho Tính năng 3: System Health Check (Ping server 3 giây/lần)
        self.health_timer = QTimer()
        self.health_timer.timeout.connect(self.check_server_health)
        self.monitor_timer = QTimer()
        self.monitor_timer.timeout.connect(self.load_monitor_overview)

    def get_db(self):
        return mysql.connector.connect(host=DB_HOST, user=DB_USER, password=DB_PASSWORD, database=DB_NAME)

    def _admin_api_headers(self, token=None):
        active_token = token or self.admin_token
        return {"X-Admin-Token": active_token} if active_token else {}

    def _clear_proctor_form(self):
        self.inp_p_id.clear()
        self.inp_p_name.clear()
        self.inp_p_pass.clear()

    def log_action(self, action_text):
        """Hàm ghi log tự động cho Tính năng 2: Audit Logs"""
        try:
            db = self.get_db(); cur = db.cursor()
            cur.execute("INSERT INTO audit_logs (actor, action) VALUES (%s, %s)", (self.admin_name, action_text))
            db.commit(); db.close()
        except Exception:
            logger.exception("Failed to write audit log for %s", self.admin_name)

    def _show_dialog(self, title, message, detail=""):
        EnterpriseDialog(title, message, detail=detail, role="admin", parent=self).exec_()

    def request_admin_password_reset(self):
        dialog = PasswordRecoveryDialog("Quản trị viên", "ID Quản Trị", account_value=self.inp_aid.text().strip(), parent=self)
        if dialog.exec_() != QDialog.Accepted:
            return
        payload = dialog.get_payload()
        if not payload["account_id"] or not payload["full_name"]:
            return self._show_dialog("Thiếu thông tin", "Vui lòng nhập ID Quản Trị và Họ tên để gửi yêu cầu.")
        try:
            response = requests.post(
                f"{self.api_url}/api/password-recovery/request",
                data={
                    "role": "admin",
                    "account_id": payload["account_id"],
                    "full_name": payload["full_name"],
                    "note": payload["note"],
                },
                timeout=10,
            )
            response.raise_for_status()
            data = response.json()
            if data.get("status") == "success":
                self._show_dialog("Đã gửi yêu cầu", data.get("message", "Một quản trị viên khác sẽ cấp lại mật khẩu tạm sau khi kiểm tra."))
            else:
                self._show_dialog("Không gửi được yêu cầu", data.get("message", "Hệ thống từ chối yêu cầu cấp lại mật khẩu."))
        except requests.RequestException as exc:
            logger.exception("Admin password recovery request failed")
            self._show_dialog("Kết nối thất bại", str(exc))
        except ValueError:
            logger.exception("Admin password recovery returned invalid JSON")
            self._show_dialog("Phản hồi không hợp lệ", "Server trả về dữ liệu không đọc được.")

    def _run_background_task(self, fn, on_success, on_error=None, on_finished=None):
        task = BackgroundTask(fn)
        task.signals.result.connect(on_success)
        task.signals.error.connect(on_error or self._log_background_error)
        if on_finished is not None:
            task.signals.finished.connect(on_finished)
        self.thread_pool.start(task)

    def _log_background_error(self, payload):
        logger.error("Background task failed: %s", payload.get("traceback") or payload.get("message"))

    def _set_login_busy(self, busy, button_text=None):
        self.admin_login_inflight = busy
        self.inp_aid.setEnabled(not busy)
        self.inp_apw.setEnabled(not busy)
        self.btn_login.setEnabled(not busy)
        self.btn_login.setText(button_text or ("ĐĂNG NHẬP HỆ THỐNG" if not busy else "ĐANG ĐĂNG NHẬP..."))

    def _apply_health_status(self, online, warning=False):
        if online:
            self.lbl_server_status.setText("🟢 API Server: ONLINE")
            style_status_pill(self.lbl_server_status, "success")
            return
        if warning:
            self.lbl_server_status.setText("🟠 API Server: LỖI KIỂM TRA")
            style_status_pill(self.lbl_server_status, "warning")
            return
        self.lbl_server_status.setText("🔴 API Server: OFFLINE")
        style_status_pill(self.lbl_server_status, "error")

    def _apply_admin_stats(self, stats):
        self.lbl_total_sv.setText(f"Sinh viên\n{stats.get('students', 0)}")
        self.lbl_total_gv.setText(f"Giám thị\n{stats.get('proctors', 0)}")
        self.lbl_total_class.setText(f"Lớp đang mở\n{stats.get('active_classes', 0)}")
        self.lbl_total_err.setText(f"Lượt cảnh báo\n{stats.get('violations', 0)}")

    def _populate_proctors_table(self, rows):
        self.tb_proctors.setRowCount(len(rows))
        self.proctors_empty.setVisible(len(rows) == 0)
        self.tb_proctors.setVisible(len(rows) > 0)
        for i, row in enumerate(rows):
            self.tb_proctors.setItem(i, 0, QTableWidgetItem(str(row['proctor_id'])))
            self.tb_proctors.setItem(i, 1, QTableWidgetItem(str(row['full_name'])))
            self.tb_proctors.setItem(i, 2, QTableWidgetItem(""))

    def _populate_template_proctor_choices(self, rows):
        choices = [(f"{row['proctor_id']} - {row['full_name']}", row['proctor_id']) for row in rows]
        for combo in [getattr(self, 'cb_template_proctor', None), getattr(self, 'cb_manage_template_proctor', None)]:
            if combo is None:
                continue
            current_value = combo.currentData()
            combo.blockSignals(True)
            combo.clear()
            for label, value in choices:
                combo.addItem(label, value)
            if current_value is not None:
                for index in range(combo.count()):
                    if combo.itemData(index) == current_value:
                        combo.setCurrentIndex(index)
                        break
            combo.blockSignals(False)

    def _populate_classes_table(self, rows):
        self.tb_classes.setRowCount(len(rows))
        self.classes_empty.setVisible(len(rows) == 0)
        self.tb_classes.setVisible(len(rows) > 0)
        for i, row in enumerate(rows):
            self.tb_classes.setItem(i, 0, QTableWidgetItem(str(row['class_id'])))
            self.tb_classes.setItem(i, 1, QTableWidgetItem(str(row['class_name'])))
            self.tb_classes.setItem(i, 2, QTableWidgetItem(str(row['status'])))
            self.tb_classes.setCellWidget(i, 3, self._build_class_action_widget(row))

    def _build_class_action_widget(self, row):
        class_id = row['class_id']
        status = str(row.get('status') or 'active').lower()
        container = QWidget()
        layout = QHBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)

        btn_close = QPushButton("KHÓA PHÒNG THI")
        btn_close.setStyleSheet("background-color: #111111; padding: 5px;")
        btn_close.clicked.connect(lambda checked=False, cid=class_id: self.force_close_class(cid))
        btn_close.setEnabled(status != 'closed')

        btn_open = QPushButton("MỞ LẠI PHÒNG THI")
        btn_open.setStyleSheet("background-color: #1E3A5F; color: white; padding: 5px;")
        btn_open.clicked.connect(lambda checked=False, cid=class_id: self.reopen_class(cid))
        btn_open.setEnabled(status == 'closed')

        layout.addWidget(btn_close)
        layout.addWidget(btn_open)
        return container

    def _populate_reports_table(self, rows):
        self.tb_reports.setRowCount(len(rows))
        self.reports_empty.setVisible(len(rows) == 0)
        self.tb_reports.setVisible(len(rows) > 0)
        for i, row in enumerate(rows):
            self.tb_reports.setItem(i, 0, QTableWidgetItem(str(row['msv'])))
            self.tb_reports.setItem(i, 1, QTableWidgetItem(str(row['exam_id'])))
            self.tb_reports.setItem(i, 2, QTableWidgetItem(str(row['time_detected'])))
            self.tb_reports.setItem(i, 3, QTableWidgetItem(str(row['error_type'])))
            self.tb_reports.setItem(i, 4, QTableWidgetItem(str(row['evidence_path'])))

    def _populate_audit_logs_table(self, rows):
        self.tb_logs.setRowCount(len(rows))
        self.logs_empty.setVisible(len(rows) == 0)
        self.tb_logs.setVisible(len(rows) > 0)
        for i, row in enumerate(rows):
            self.tb_logs.setItem(i, 0, QTableWidgetItem(str(row['log_id'])))
            self.tb_logs.setItem(i, 1, QTableWidgetItem(str(row['timestamp'])))
            self.tb_logs.setItem(i, 2, QTableWidgetItem(str(row['actor'])))
            self.tb_logs.setItem(i, 3, QTableWidgetItem(str(row['action'])))

    def _apply_configs_rows(self, rows):
        for row in rows:
            if row['setting_key'] == 'ai_face_threshold':
                self.inp_threshold.setValue(self._normalize_face_threshold_value(row['setting_value']))
            if row['setting_key'] == 'max_warnings':
                self.inp_warnings.setValue(int(row['setting_value']))

    def _apply_monitor_overview_payload(self, rows):
        self.monitor_entries = rows
        self.apply_monitor_filter()
        if self.current_monitor_session_token:
            for row_index in range(self.tb_monitor.rowCount()):
                item = self.tb_monitor.item(row_index, 0)
                if item and item.data(Qt.UserRole) == self.current_monitor_session_token:
                    self.tb_monitor.selectRow(row_index)
                    break
        elif self.tb_monitor.rowCount() > 0:
            self.tb_monitor.selectRow(0)

    def _apply_monitor_detail_payload(self, session_token, detail):
        if session_token != self.current_monitor_session_token:
            return
        self.lbl_monitor_meta.setText(
            f"Phiên thi: {detail.get('msv', '--')} | {detail.get('full_name', '')} | Lớp {detail.get('class_id', '--')} | Giám thị {detail.get('proctor_id', '--')}"
        )
        self.lbl_monitor_risk.setText(
            f"Mức rủi ro: {detail.get('risk_level', '--')} | Phiên {detail.get('session_warning_count', 0)}/{detail.get('max_warnings', 0)} | Lịch sử {detail.get('historical_warning_count', 0)} | Tổng {detail.get('total_warning_count', 0)} | Khóa phiên sinh viên: {'Có' if detail.get('warning_locked') else 'Không'} | Khóa phòng thi: {'Có' if detail.get('class_locked') else 'Không'}"
        )
        self.lbl_monitor_state.setText(
            f"Trạng thái AI: {detail.get('identity_status', '--')} | People {detail.get('people_count', 0)} | Active {'Có' if detail.get('active') else 'Không'}"
        )
        self.lbl_monitor_update.setText(f"Cập nhật cuối: {detail.get('updated_at', '--')} | {detail.get('seconds_since_update', 0)}s trước")
        self.pb_monitor_total_risk.setValue(int(max(0.0, min(100.0, float(detail.get('risk_score', 0.0))))))
        summary_lines = [
            f"Sự kiện gần nhất: {detail.get('last_event', 'Không có')}",
            f"Broadcast trạng thái: {detail.get('status_text', 'Không có')}",
            f"Snapshot giám sát: {detail.get('snapshot_status', 'missing')} | Cập nhật {detail.get('snapshot_updated_at', '--')}",
            f"Phone detected: {detail.get('phone_detected')}, head-pose audit: {detail.get('head_pose_status', 'unavailable')}",
            f"Pitch/Yaw: {detail.get('pitch', 0.0)} / {detail.get('yaw', 0.0)} | reliable={detail.get('pose_reliable')}",
            "Admin chỉ xem metadata audit, bằng chứng vi phạm và khả năng khóa phòng; không hiển thị ảnh snapshot trực tiếp.",
        ]
        self.txt_monitor_summary.setPlainText("\n".join(summary_lines))

    def _fetch_admin_login_bundle(self, aid, pw):
        db = self.get_db()
        try:
            cur = db.cursor(dictionary=True)
            cur.execute("SELECT admin_id, full_name, password FROM admins WHERE admin_id = %s", (aid,))
            res = cur.fetchone()
            if not res:
                return {"status": "error", "message": "Sai ID hoặc Mật khẩu Quản trị!"}

            verified, upgraded = verify_and_upgrade_password(cur, "admins", "admin_id", aid, pw, res.get("password"))
            if not verified:
                return {"status": "error", "message": "Sai ID hoặc Mật khẩu Quản trị!"}
            if upgraded:
                db.commit()

            cur.execute("SELECT COUNT(*) AS value FROM students")
            students_count = int(cur.fetchone()["value"])
            cur.execute("SELECT COUNT(*) AS value FROM proctors")
            proctors_count = int(cur.fetchone()["value"])
            cur.execute("SELECT COUNT(*) AS value FROM classes WHERE status='active'")
            active_classes = int(cur.fetchone()["value"])
            cur.execute("SELECT COUNT(*) AS value FROM violations")
            violations_count = int(cur.fetchone()["value"])
            cur.execute("SELECT * FROM students")
            students = cur.fetchall()
            cur.execute("SELECT * FROM proctors")
            proctors = cur.fetchall()
            cur.execute("SELECT * FROM classes ORDER BY created_at DESC")
            classes = cur.fetchall()
            cur.execute("SELECT * FROM configs")
            configs = cur.fetchall()
            cur.execute("SELECT * FROM audit_logs ORDER BY timestamp DESC LIMIT 100")
            audit_logs = cur.fetchall()
        finally:
            db.close()

        try:
            requests.get(self.api_url, timeout=1)
            health = {"online": True, "warning": False}
        except requests.RequestException:
            health = {"online": False, "warning": False}
        except Exception:
            health = {"online": False, "warning": True}

        admin_token = None
        try:
            login_response = requests.post(
                f"{self.api_url}/api/admin/login",
                data={"admin_id": aid, "password": pw},
                timeout=10,
            )
            login_response.raise_for_status()
            login_payload = login_response.json()
            if login_payload.get("status") == "success":
                admin_token = login_payload.get("token")
                response = requests.get(
                    f"{self.api_url}/api/monitor/admin",
                    headers=self._admin_api_headers(admin_token),
                    timeout=5,
                )
                response.raise_for_status()
                monitor_payload = response.json()
                monitor_entries = monitor_payload.get("data", []) if monitor_payload.get("status") == "success" else []
            else:
                monitor_entries = []
        except Exception:
            logger.exception("Failed to preload admin monitor overview")
            monitor_entries = []

        return {
            "status": "success",
            "admin_name": res["full_name"],
            "admin_token": admin_token,
            "health": health,
            "stats": {
                "students": students_count,
                "proctors": proctors_count,
                "active_classes": active_classes,
                "violations": violations_count,
            },
            "students": students,
            "proctors": proctors,
            "classes": classes,
            "configs": configs,
            "audit_logs": audit_logs,
            "monitor_entries": monitor_entries,
        }

    def _on_admin_login_finished(self, payload):
        if payload.get("status") != "success":
            QMessageBox.warning(self, "Lỗi", payload.get("message", "Đăng nhập thất bại"))
            return
        self.admin_name = payload["admin_name"]
        self.admin_token = payload.get("admin_token")
        self.lbl_welcome.setText(f"QUẢN TRỊ VIÊN: {self.admin_name.upper()}")
        self.stack.setCurrentIndex(1)
        self._apply_health_status(payload["health"]["online"], payload["health"].get("warning", False))
        self._apply_admin_stats(payload["stats"])
        self.populate_student_table(payload["students"])
        self._populate_proctors_table(payload["proctors"])
        self._populate_template_proctor_choices(payload["proctors"])
        self._populate_classes_table(payload["classes"])
        self._apply_monitor_overview_payload(payload["monitor_entries"])
        self._apply_configs_rows(payload["configs"])
        self.load_password_reset_requests()
        self._populate_audit_logs_table(payload["audit_logs"])
        self.health_timer.start(5000)
        self.monitor_timer.start(5000)
        self._run_background_task(lambda: self.log_action("Đăng nhập vào hệ thống Admin"), lambda _result: None)

    def _on_admin_login_error(self, payload):
        logger.error("Admin login background task failed: %s", payload.get("traceback") or payload.get("message"))
        QMessageBox.critical(self, "Lỗi DB", payload.get("message", "Không thể đăng nhập."))

    def _fetch_admin_monitor_overview(self):
        if not self.admin_token:
            return []
        response = requests.get(f"{self.api_url}/api/monitor/admin", headers=self._admin_api_headers(), timeout=5)
        response.raise_for_status()
        payload = response.json()
        return payload.get("data", []) if payload.get("status") == "success" else []

    def _fetch_admin_monitor_detail(self, session_token):
        if not self.admin_token:
            return {"session_token": session_token, "detail": {}}
        response = requests.get(
            f"{self.api_url}/api/monitor/admin/{session_token}",
            headers=self._admin_api_headers(),
            timeout=5,
        )
        response.raise_for_status()
        payload = response.json()
        if payload.get("status") != "success":
            return {"session_token": session_token, "detail": {}}
        return {"session_token": session_token, "detail": payload.get("data", {})}

    def _fetch_admin_reports(self, keyword=""):
        db = self.get_db()
        try:
            cur = db.cursor(dictionary=True)
            if keyword:
                cur.execute(
                    "SELECT msv, exam_id, time_detected, error_type, evidence_path FROM violations WHERE msv LIKE %s OR CAST(exam_id AS CHAR) LIKE %s OR error_type LIKE %s ORDER BY time_detected DESC",
                    (f"%{keyword}%", f"%{keyword}%", f"%{keyword}%"),
                )
            else:
                cur.execute("SELECT msv, exam_id, time_detected, error_type, evidence_path FROM violations ORDER BY time_detected DESC")
            return cur.fetchall()
        finally:
            db.close()

    def _on_admin_monitor_overview_loaded(self, rows):
        self._apply_monitor_overview_payload(rows)

    def _on_admin_monitor_overview_error(self, payload):
        logger.error("Admin monitor overview failed: %s", payload.get("traceback") or payload.get("message"))

    def _on_admin_report_load_error(self, payload):
        logger.error("Admin reports load failed: %s", payload.get("traceback") or payload.get("message"))
        self._show_dialog("Không thể tải báo cáo", payload.get("message", "Không tải được danh sách vi phạm toàn hệ thống."))

    def _normalize_face_threshold_value(self, raw_value):
        try:
            threshold = float(raw_value)
        except (TypeError, ValueError):
            return DEFAULT_FACE_THRESHOLD
        if threshold < MIN_FACE_THRESHOLD or threshold > MAX_FACE_THRESHOLD:
            return DEFAULT_FACE_THRESHOLD
        return threshold

    def _load_pixmap_from_bytes(self, image_bytes):
        if not image_bytes:
            return None
        pixmap = QPixmap()
        if pixmap.loadFromData(image_bytes) and not pixmap.isNull():
            return pixmap
        np_buffer = np.frombuffer(image_bytes, np.uint8)
        decoded = cv2.imdecode(np_buffer, cv2.IMREAD_COLOR)
        if decoded is None:
            return None
        rgb = cv2.cvtColor(decoded, cv2.COLOR_BGR2RGB)
        image = QImage(rgb.data, rgb.shape[1], rgb.shape[0], rgb.strides[0], QImage.Format_RGB888)
        return QPixmap.fromImage(image.copy())

    def _load_pixmap_from_base64(self, encoded_image):
        if not encoded_image:
            return None
        try:
            image_bytes = base64.b64decode(encoded_image)
        except Exception:
            return None
        return self._load_pixmap_from_bytes(image_bytes)

    def _resolve_local_evidence_path(self, evidence_filename):
        normalized = str(evidence_filename or "").strip()
        if not normalized:
            return None
        if os.path.isabs(normalized) and os.path.exists(normalized):
            return normalized
        candidates = [
            os.path.join(EVIDENCE_DIR, normalized),
            os.path.join(os.getcwd(), "server_evidence", normalized),
        ]
        for candidate in candidates:
            if os.path.exists(candidate):
                return candidate
        return candidates[0]

    def _read_template_rows_from_csv(self, filepath):
        with open(filepath, "r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            return list(reader)

    def _read_template_rows_from_excel(self, filepath):
        workbook = load_workbook(filepath, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value or "").strip() for value in rows[0]]
        data_rows = []
        for raw_row in rows[1:]:
            if raw_row is None:
                continue
            row_dict = {headers[index]: raw_row[index] for index in range(min(len(headers), len(raw_row)))}
            data_rows.append(row_dict)
        return data_rows

    def _normalize_question_import_rows(self, rows):
        field_aliases = {
            "question_text": {"questiontext", "question", "cauhoi", "noidung", "noidungcauhoi"},
            "option_a": {"optiona", "dapana", "a"},
            "option_b": {"optionb", "dapanb", "b"},
            "option_c": {"optionc", "dapanc", "c"},
            "option_d": {"optiond", "dapand", "d"},
            "correct_option": {"correctoption", "correctanswer", "dapandung", "dapan", "answer", "correct"},
            "points": {"points", "score", "diem"},
        }
        normalized_rows = []
        for source_row in rows:
            normalized_source = {_normalize_import_key(key): value for key, value in source_row.items()}
            row = {}
            for target_key, aliases in field_aliases.items():
                row[target_key] = ""
                for alias in aliases:
                    if alias in normalized_source and normalized_source[alias] is not None:
                        row[target_key] = str(normalized_source[alias]).strip()
                        break
            if not row["question_text"]:
                continue
            if not row["option_a"] or not row["option_b"]:
                raise ValueError(f"Câu hỏi '{row['question_text'][:60]}' thiếu ít nhất đáp án A hoặc B.")
            correct_option = row["correct_option"].upper() if row["correct_option"] else "A"
            if correct_option not in {"A", "B", "C", "D"}:
                raise ValueError(f"Đáp án đúng không hợp lệ cho câu hỏi '{row['question_text'][:60]}'.")
            try:
                points = float(row["points"] or 1.0)
            except ValueError as exc:
                raise ValueError(f"Điểm số không hợp lệ cho câu hỏi '{row['question_text'][:60]}'.") from exc
            normalized_rows.append(
                {
                    "question_text": row["question_text"],
                    "option_a": row["option_a"],
                    "option_b": row["option_b"],
                    "option_c": row["option_c"] or None,
                    "option_d": row["option_d"] or None,
                    "correct_option": correct_option,
                    "points": points,
                }
            )
        return normalized_rows

    def init_ui(self):
        self.stack = QStackedWidget()
        self.setCentralWidget(self.stack)

        # ================= PAGE 1: LOGIN =================
        self.page_login = QWidget()
        l_layout = QVBoxLayout(self.page_login)
        set_page_margins(self.page_login)
        login_box = QGroupBox("XÁC THỰC QUẢN TRỊ VIÊN")
        login_box.setFixedSize(400, 250)
        box_layout = QVBoxLayout(login_box)
        
        self.inp_aid = QLineEdit(); self.inp_aid.setPlaceholderText("ID Quản Trị (Ví dụ: ADMIN_THDO)")
        self.inp_apw = QLineEdit(); self.inp_apw.setPlaceholderText("Mật Khẩu"); self.inp_apw.setEchoMode(QLineEdit.Password)
        self.btn_login = QPushButton("ĐĂNG NHẬP HỆ THỐNG")
        btn_recover = QPushButton("YÊU CẦU CẤP LẠI MẬT KHẨU")
        btn_recover.setStyleSheet("background-color: #2F2F2F;")
        self.btn_login.setStyleSheet("background-color: #111111;")
        self.btn_login.clicked.connect(self.process_login)
        btn_recover.clicked.connect(self.request_admin_password_reset)
        self.inp_aid.returnPressed.connect(self.process_login)
        self.inp_apw.returnPressed.connect(self.process_login)
        
        box_layout.addStretch(); box_layout.addWidget(self.inp_aid); box_layout.addWidget(self.inp_apw)
        box_layout.addWidget(self.btn_login); box_layout.addWidget(btn_recover); box_layout.addStretch()
        login_logo = build_login_logo_label(role="admin")
        
        center_l = QHBoxLayout(); center_l.addStretch(); center_l.addWidget(login_box); center_l.addStretch()
        l_layout.addStretch(); l_layout.addWidget(login_logo, 0, Qt.AlignCenter); l_layout.addSpacing(6); l_layout.addLayout(center_l); l_layout.addStretch()
        self.stack.addWidget(self.page_login)

        # ================= PAGE 2: MAIN DASHBOARD =================
        self.page_dashboard = QWidget()
        dash_layout = QHBoxLayout(self.page_dashboard)
        set_page_margins(self.page_dashboard)
        self.sidebar = SidebarNav(
            "S-MONITOR",
            "Super admin workspace cho vận hành, điều phối và kiểm soát toàn hệ thống.",
            ["Tổng quan", "Sinh viên", "Giám thị", "Lớp thi", "Audit snapshot", "Bộ đề", "Báo cáo", "Bảng điểm", "Cấu hình AI", "Khôi phục MK", "Nhật ký", "Sao lưu"],
            role="admin",
        )
        dash_layout.addWidget(self.sidebar)

        content_layout = QVBoxLayout()
        self.toolbar = AppToolbar(
            "Tổng quan",
            "Theo dõi tình trạng nền tảng, nhân sự và các snapshot audit phục vụ vận hành.",
            role="admin",
        )
        self.lbl_welcome = self.toolbar.title_label
        self.lbl_server_status = self.toolbar.badge
        self.lbl_server_status.setAlignment(Qt.AlignCenter)
        style_status_pill(self.lbl_server_status, "info")
        content_layout.addWidget(self.toolbar)
        
        self.tabs = QTabWidget()
        self.tabs.tabBar().hide()
        self.tabs.setStyleSheet("QTabBar::tab { padding: 12px 20px; font-weight: bold; font-size: 14px; }")
        
        self.setup_tab_overview()     # Thống kê tổng quan
        self.setup_tab_students()     # Quản lý SV + Search + Import CSV
        self.setup_tab_proctors()     # Quản lý Giám thị
        self.setup_tab_master()       # Master Control
        self.setup_tab_monitor()      # Audit snapshot toàn hệ thống
        self.setup_tab_templates()    # Quản lý bộ đề toàn hệ thống
        self.setup_tab_reports()      # Báo cáo vi phạm toàn hệ thống
        self.setup_tab_scores()       # Bảng điểm toàn hệ thống
        self.setup_tab_configs()      # Cấu hình AI
        self.setup_tab_password_resets()  # Yêu cầu khôi phục mật khẩu
        self.setup_tab_audit_logs()   # NHẬT KÝ HỆ THỐNG (Mới)
        self.setup_tab_backup()       # Sao lưu dữ liệu

        connect_sidebar_to_tabs(
            self.sidebar,
            self.tabs,
            toolbar=self.toolbar,
            section_descriptions=[
                "Theo dõi tình trạng nền tảng, nhân sự và sự kiện vận hành theo thời gian gần thực.",
                "Quản lý hồ sơ sinh viên, ảnh tham chiếu và import dữ liệu hàng loạt.",
                "Điều phối tài khoản giám thị và phân quyền vận hành ca thi.",
                "Can thiệp lớp thi và khóa phòng trong các tình huống vận hành khẩn cấp.",
                "Theo dõi snapshot AI và trạng thái phiên thi toàn hệ thống phục vụ audit vận hành.",
                "Toàn quyền tạo, nhập, rà soát và chỉnh sửa bộ đề trên toàn hệ thống.",
                "Xem bằng chứng vi phạm của tất cả lớp thi và mở ảnh đối chứng khi cần.",
                "Theo dõi điểm thi toàn hệ thống và xuất dữ liệu phục vụ đối soát.",
                "Điều chỉnh ngưỡng AI và mức cảnh báo theo chính sách giám sát hiện hành.",
                "Tiếp nhận và xử lý yêu cầu cấp lại mật khẩu bằng mật khẩu tạm cho từng vai trò.",
                "Truy vết hoạt động quản trị và lịch sử thao tác quan trọng.",
                "Xuất dữ liệu phục vụ kiểm toán, báo cáo hoặc lưu trữ định kỳ.",
            ],
        )

        content_layout.addWidget(self.tabs, 1)
        dash_layout.addLayout(content_layout, 1)
        self.stack.addWidget(self.page_dashboard)

    def check_server_health(self):
        """Tính năng 3: Cập nhật trạng thái máy chủ API"""
        try:
            r = requests.get(self.api_url, timeout=1)
            # Nếu request gửi đi không bị sập mạng (200 hoặc 404 đều là server đang sống)
            self.lbl_server_status.setText("🟢 API Server: ONLINE")
            style_status_pill(self.lbl_server_status, "success")
        except requests.RequestException:
            self.lbl_server_status.setText("🔴 API Server: OFFLINE")
            style_status_pill(self.lbl_server_status, "error")
        except Exception:
            logger.exception("Unexpected error during server health check")
            self.lbl_server_status.setText("🟠 API Server: LỖI KIỂM TRA")
            style_status_pill(self.lbl_server_status, "warning")

    # ---------------- TAB 1: THỐNG KÊ TỔNG QUAN ----------------
    def setup_tab_overview(self):
        tab = QWidget(); layout = QVBoxLayout(tab)
        stats_layout = QHBoxLayout()
        self.lbl_total_sv = QLabel("Sinh viên\n0"); self.format_stat_label(self.lbl_total_sv, "#3498DB")
        self.lbl_total_gv = QLabel("Giám thị\n0"); self.format_stat_label(self.lbl_total_gv, "#9B59B6")
        self.lbl_total_class = QLabel("Lớp thi Active\n0"); self.format_stat_label(self.lbl_total_class, "#2ECC71")
        self.lbl_total_err = QLabel("Lượt gian lận\n0"); self.format_stat_label(self.lbl_total_err, "#E74C3C")
        
        stats_layout.addWidget(self.lbl_total_sv); stats_layout.addWidget(self.lbl_total_gv)
        stats_layout.addWidget(self.lbl_total_class); stats_layout.addWidget(self.lbl_total_err)
        btn_refresh = QPushButton("LÀM MỚI SỐ LIỆU"); btn_refresh.clicked.connect(self.refresh_stats)
        layout.addLayout(stats_layout); layout.addWidget(btn_refresh); layout.addStretch()
        self.tabs.addTab(tab, "📊 Tổng quan")

    def format_stat_label(self, lbl, color):
        style_stat_label(lbl, color)
        lbl.setAlignment(Qt.AlignCenter)

    def refresh_stats(self):
        try:
            db = self.get_db(); cur = db.cursor()
            cur.execute("SELECT COUNT(*) FROM students"); self.lbl_total_sv.setText(f"Sinh viên\n{cur.fetchone()[0]}")
            cur.execute("SELECT COUNT(*) FROM proctors"); self.lbl_total_gv.setText(f"Giám thị\n{cur.fetchone()[0]}")
            cur.execute("SELECT COUNT(*) FROM classes WHERE status='active'"); self.lbl_total_class.setText(f"Lớp đang mở\n{cur.fetchone()[0]}")
            cur.execute("SELECT COUNT(*) FROM violations"); self.lbl_total_err.setText(f"Lượt cảnh báo\n{cur.fetchone()[0]}")
            db.close()
        except Exception:
            logger.exception("Failed to refresh admin statistics")
            QMessageBox.warning(self, "Lỗi", "Không tải được số liệu tổng quan.")

    # ---------------- TAB 2: QUẢN LÝ SINH VIÊN (TÌM KIẾM + IMPORT CSV) ----------------
    def setup_tab_students(self):
        tab = QWidget(); layout = QHBoxLayout(tab)
        
        # Form nhập liệu bên trái
        form_gb = QGroupBox("Thông tin Sinh viên")
        form_gb.setFixedWidth(350); fl = QVBoxLayout(form_gb)
        self.inp_s_msv = QLineEdit(); self.inp_s_msv.setPlaceholderText("Mã SV (Ví dụ: 2200404)")
        self.inp_s_name = QLineEdit(); self.inp_s_name.setPlaceholderText("Họ tên")
        self.inp_s_class = QLineEdit(); self.inp_s_class.setPlaceholderText("Lớp hành chính (VD: K22_CNTT)")
        self.inp_s_pass = QLineEdit(); self.inp_s_pass.setPlaceholderText("Mật khẩu")
        
        btn_add = QPushButton("THÊM / CẬP NHẬT"); btn_add.clicked.connect(self.save_student)
        btn_upload = QPushButton("📸 TẢI LÊN ẢNH THAM CHIẾU"); btn_upload.setStyleSheet("background-color: #2A2A2A;")
        btn_upload.clicked.connect(self.upload_face)
        self.btn_delete_face = QPushButton("🗑️ XÓA ẢNH ĐÃ CHỌN")
        self.btn_delete_face.setStyleSheet("background-color: #2A2A2A;")
        self.btn_delete_face.setEnabled(False)
        self.btn_delete_face.clicked.connect(self.delete_selected_face)
        btn_del = QPushButton("XÓA SINH VIÊN"); btn_del.setStyleSheet("background-color: #1A1A1A;"); btn_del.clicked.connect(self.del_student)
        
        # Tính năng 1: Nhập hàng loạt (Bulk Import CSV)
        btn_import_csv = QPushButton("📂 NHẬP HÀNG LOẠT (CSV)"); btn_import_csv.setStyleSheet("background-color: #3B3B3B;")
        btn_import_csv.clicked.connect(self.import_students_csv)
        
        fl.addWidget(self.inp_s_msv); fl.addWidget(self.inp_s_name); fl.addWidget(self.inp_s_class); fl.addWidget(self.inp_s_pass)
        fl.addWidget(btn_add); fl.addWidget(btn_upload); fl.addWidget(self.btn_delete_face); fl.addWidget(btn_del); fl.addSpacing(20); fl.addWidget(btn_import_csv); fl.addStretch()
        
        # Bảng hiển thị bên phải (Có Tìm kiếm)
        right_layout = QVBoxLayout()
        self.student_filter_bar = FilterBar("Tìm theo mã sinh viên hoặc họ tên", "Tìm kiếm", "Tải lại")
        self.inp_search_sv = self.student_filter_bar.search_input
        wire_filter_bar(self.student_filter_bar, self.search_students, self.load_students)
        
        self.tb_students = QTableWidget(); self.tb_students.setColumnCount(5)
        self.tb_students.setHorizontalHeaderLabels(["MSV", "Họ tên", "Lớp", "Mật khẩu", "Ảnh tham chiếu"])
        self.tb_students.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tb_students.setColumnHidden(3, True)
        self.tb_students.itemClicked.connect(self.select_student)
        polish_table(self.tb_students)
        self.students_empty = EmptyState("Chưa có sinh viên", "Thêm thủ công hoặc import CSV để khởi tạo danh sách sinh viên.")

        self.tb_student_faces = QTableWidget(); self.tb_student_faces.setColumnCount(3)
        self.tb_student_faces.setHorizontalHeaderLabels(["Ref ID", "Tên file ảnh", "Ảnh chính"])
        self.tb_student_faces.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.tb_student_faces.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.tb_student_faces.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.tb_student_faces.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tb_student_faces.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.tb_student_faces.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tb_student_faces.itemSelectionChanged.connect(self.select_student_face_reference)
        polish_table(self.tb_student_faces)
        self.student_faces_empty = EmptyState("Chưa có ảnh tham chiếu", "Chọn sinh viên để xem danh sách từng ảnh tham chiếu; có thể chọn nhiều ảnh để xóa cùng lúc.")
        
        right_layout.addWidget(self.student_filter_bar); right_layout.addWidget(self.students_empty); right_layout.addWidget(self.tb_students)
        right_layout.addWidget(QLabel("Danh sách ảnh tham chiếu của sinh viên đang chọn (giữ Ctrl/Shift để chọn nhiều ảnh)"))
        right_layout.addWidget(self.student_faces_empty)
        right_layout.addWidget(self.tb_student_faces)
        layout.addWidget(form_gb); layout.addLayout(right_layout)
        self.tabs.addTab(tab, "🎓 Sinh viên")

    def _clear_student_face_selection(self):
        self.selected_student_face_ref_id = None
        self.selected_student_face_image = ""
        self.selected_student_face_targets = []
        if hasattr(self, "btn_delete_face"):
            self.btn_delete_face.setEnabled(False)

    def _collect_selected_student_faces(self):
        targets = []
        selected_rows = sorted({idx.row() for idx in self.tb_student_faces.selectionModel().selectedRows()})
        for row in selected_rows:
            ref_item = self.tb_student_faces.item(row, 0)
            if ref_item is None:
                continue
            image_name = str(ref_item.data(Qt.UserRole + 1) or "").strip()
            if not image_name:
                continue
            targets.append({
                "ref_id": ref_item.data(Qt.UserRole),
                "face_image": image_name,
            })
        return targets

    def _fetch_student_face_rows(self, msv):
        rows_out = []
        if not msv:
            return rows_out
        db = self.get_db(); cur = db.cursor(dictionary=True)
        try:
            seen = set()
            try:
                cur.execute(
                    """
                    SELECT ref_id, face_image, is_primary
                    FROM student_face_images
                    WHERE msv = %s AND face_image IS NOT NULL AND TRIM(face_image) <> ''
                    ORDER BY is_primary DESC, ref_id ASC
                    """,
                    (msv,),
                )
                for row in cur.fetchall() or []:
                    face_image = str(row.get("face_image") or "").strip()
                    if not face_image or face_image in seen:
                        continue
                    seen.add(face_image)
                    rows_out.append({
                        "ref_id": int(row.get("ref_id")) if row.get("ref_id") is not None else None,
                        "face_image": face_image,
                        "is_primary": bool(row.get("is_primary")),
                    })
            except mysql.connector.Error:
                logger.debug("student_face_images unavailable while loading references", exc_info=True)

            cur.execute("SELECT face_image FROM students WHERE msv = %s", (msv,))
            legacy_row = cur.fetchone() or {}
            legacy_face = str(legacy_row.get("face_image") or "").strip()
            if legacy_face and legacy_face not in seen:
                rows_out.append({"ref_id": None, "face_image": legacy_face, "is_primary": True})
            return rows_out
        finally:
            db.close()

    def populate_student_face_table(self, rows):
        self._clear_student_face_selection()
        self.tb_student_faces.clearSelection()
        self.tb_student_faces.setRowCount(len(rows))
        self.student_faces_empty.setVisible(len(rows) == 0)
        self.tb_student_faces.setVisible(len(rows) > 0)
        for i, row in enumerate(rows):
            ref_id_text = str(row.get("ref_id")) if row.get("ref_id") is not None else "--"
            ref_item = QTableWidgetItem(ref_id_text)
            ref_item.setData(Qt.UserRole, row.get("ref_id"))
            ref_item.setData(Qt.UserRole + 1, str(row.get("face_image") or ""))
            self.tb_student_faces.setItem(i, 0, ref_item)
            self.tb_student_faces.setItem(i, 1, QTableWidgetItem(str(row.get("face_image") or "")))
            primary_item = QTableWidgetItem("Có" if row.get("is_primary") else "Không")
            primary_item.setTextAlignment(Qt.AlignCenter)
            self.tb_student_faces.setItem(i, 2, primary_item)

    def load_student_faces(self, msv):
        try:
            rows = self._fetch_student_face_rows(msv)
            self.populate_student_face_table(rows)
        except Exception:
            logger.exception("Failed to load face references for %s", msv)
            self.populate_student_face_table([])
            QMessageBox.warning(self, "Lỗi", f"Không tải được ảnh tham chiếu của sinh viên {msv}.")

    def _fetch_student_rows(self, keyword=None):
        db = self.get_db(); cur = db.cursor(dictionary=True)
        try:
            try:
                base_query = """
                    SELECT s.*, COALESCE(refs.reference_count, 0) AS reference_count
                    FROM students s
                    LEFT JOIN (
                        SELECT msv, COUNT(*) AS reference_count
                        FROM student_face_images
                        GROUP BY msv
                    ) refs ON refs.msv = s.msv
                """
                params = ()
                if keyword:
                    base_query += " WHERE s.msv LIKE %s OR s.full_name LIKE %s"
                    params = (f"%{keyword}%", f"%{keyword}%")
                base_query += " ORDER BY s.msv"
                cur.execute(base_query, params)
            except mysql.connector.Error:
                if keyword:
                    cur.execute("SELECT * FROM students WHERE msv LIKE %s OR full_name LIKE %s ORDER BY msv", (f"%{keyword}%", f"%{keyword}%"))
                else:
                    cur.execute("SELECT * FROM students ORDER BY msv")
            return cur.fetchall()
        finally:
            db.close()

    def load_students(self):
        self.inp_search_sv.clear()
        try:
            rows = self._fetch_student_rows()
            self.populate_student_table(rows)
            self.populate_student_face_table([])
        except Exception:
            logger.exception("Failed to load students")
            QMessageBox.warning(self, "Lỗi", "Không tải được danh sách sinh viên.")

    def search_students(self):
        keyword = self.inp_search_sv.text().strip()
        if not keyword: return self.load_students()
        try:
            rows = self._fetch_student_rows(keyword)
            self.populate_student_table(rows)
        except Exception as e: QMessageBox.critical(self, "Lỗi", str(e))

    def populate_student_table(self, rows):
        self.tb_students.setRowCount(len(rows))
        self.students_empty.setVisible(len(rows) == 0)
        self.tb_students.setVisible(len(rows) > 0)
        for i, r in enumerate(rows):
            self.tb_students.setItem(i, 0, QTableWidgetItem(str(r['msv'])))
            self.tb_students.setItem(i, 1, QTableWidgetItem(str(r['full_name'])))
            self.tb_students.setItem(i, 2, QTableWidgetItem(str(r['class_name'])))
            self.tb_students.setItem(i, 3, QTableWidgetItem(""))
            reference_count = int(r.get('reference_count', 0) or 0)
            primary_face = str(r.get('face_image') or '')
            if reference_count > 0:
                summary = f"{reference_count} ảnh | chính: {primary_face or '--'}"
            else:
                summary = primary_face or "Chưa có"
            self.tb_students.setItem(i, 4, QTableWidgetItem(summary))

    def import_students_csv(self):
        """Tính năng 1: Hàm nhập dữ liệu sinh viên hàng loạt từ Excel/CSV"""
        filepath, _ = QFileDialog.getOpenFileName(self, "Chọn file CSV danh sách Sinh viên", "", "CSV Files (*.csv)")
        if not filepath: return
        
        db = None
        try:
            db = self.get_db()
            cur = db.cursor()
            count = 0
            failed_rows = []
            
            with open(filepath, 'r', encoding='utf-8') as file:
                reader = csv.DictReader(file)
                
                # Fix #16: Validate CSV structure before processing
                if not reader.fieldnames:
                    raise ValueError("File CSV không có tiêu đề cột (header)")
                
                required_cols = ['msv', 'full_name']
                missing_cols = [col for col in required_cols if col not in reader.fieldnames]
                if missing_cols:
                    raise ValueError(f"File CSV thiếu cột bắt buộc: {', '.join(missing_cols)}")
                
                for row_num, row in enumerate(reader, 1):
                    try:
                        # Fix #16: Validate data in each row
                        msv = str(row.get('msv', '') or '').strip()
                        name = str(row.get('full_name', '') or '').strip()
                        cls = str(row.get('class_name', '') or '').strip()
                        pw = str(row.get('password', '123456') or '123456').strip()
                        
                        # Validate required fields
                        if not msv or not name:
                            failed_rows.append(f"Dòng {row_num}: msv hoặc full_name bị trống")
                            continue
                        
                        # Validate field lengths
                        if len(msv) > 20:
                            failed_rows.append(f"Dòng {row_num}: msv quá dài (max 20 ký tự)")
                            continue
                        if len(name) > 255:
                            failed_rows.append(f"Dòng {row_num}: full_name quá dài (max 255 ký tự)")
                            continue
                        
                        cur.execute(
                            """
                            INSERT INTO students (msv, full_name, class_name, password)
                            VALUES (%s, %s, %s, %s)
                            ON DUPLICATE KEY UPDATE
                                full_name = VALUES(full_name),
                                class_name = VALUES(class_name),
                                password = VALUES(password)
                            """,
                            (msv, name, cls, hash_password(pw)),
                        )
                        count += 1
                    except Exception as e:
                        failed_rows.append(f"Dòng {row_num}: {str(e)}")
                        continue
            
            db.commit()
            self.log_action(f"Import sinh viên: {count} thành công, {len(failed_rows)} lỗi")
            self.load_students()
            self.refresh_stats()
            
            if failed_rows:
                error_details = "\\n".join(failed_rows[:10])  # Show first 10 errors
                if len(failed_rows) > 10:
                    error_details += f"\\n... và {len(failed_rows) - 10} lỗi khác"
                QMessageBox.warning(self, "Cảnh báo", f"Đã nhập {count} sinh viên thành công.\\n\\nLỗi:\\n{error_details}")
            else:
                QMessageBox.information(self, "Thành công", f"Đã nhập thành công {count} sinh viên vào hệ thống!")
        except Exception as e:
            logger.exception("CSV import failed")
            QMessageBox.critical(self, "Lỗi đọc file", f"Không thể nhập file CSV.\\n\\nChi tiết: {str(e)}")
        finally:
            # Fix #14: Ensure database connection is always closed
            if db:
                try:
                    db.close()
                except Exception as e:
                    logger.debug("Error closing database: %s", e)

    def save_student(self):
        msv = self.inp_s_msv.text().strip()
        name = self.inp_s_name.text().strip()
        cls = self.inp_s_class.text().strip()
        pw = self.inp_s_pass.text().strip()
        if not msv or not name or not cls:
            return QMessageBox.warning(self, "Lỗi", "Vui lòng nhập đủ Mã SV, Họ tên và Lớp.")
        try:
            db = self.get_db()
            cur = db.cursor(dictionary=True)
            cur.execute("SELECT password FROM students WHERE msv = %s", (msv,))
            existing_row = cur.fetchone()
            if existing_row and not pw:
                password_to_store = existing_row["password"]
            elif pw:
                password_to_store = hash_password(pw)
            else:
                db.close()
                return QMessageBox.warning(self, "Lỗi", "Vui lòng nhập mật khẩu cho sinh viên mới.")

            cur.execute(
                """
                INSERT INTO students (msv, full_name, class_name, password)
                VALUES (%s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    full_name = VALUES(full_name),
                    class_name = VALUES(class_name),
                    password = VALUES(password)
                """,
                (msv, name, cls, password_to_store),
            )
            db.commit()
            db.close()
            self.load_students()
            self.refresh_stats()
            self.log_action(f"Thêm/Cập nhật sinh viên MSV: {msv}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def del_student(self):
        msv = self.inp_s_msv.text()
        try:
            face_rows = self._fetch_student_face_rows(msv)
            db = self.get_db(); cur = db.cursor();
            try:
                cur.execute("DELETE FROM student_face_images WHERE msv=%s", (msv,))
            except mysql.connector.Error:
                logger.debug("student_face_images unavailable while deleting student", exc_info=True)
            cur.execute("DELETE FROM students WHERE msv=%s", (msv,))
            db.commit(); db.close();
            for row in face_rows:
                face_path = os.path.join("server_database", str(row.get("face_image") or ""))
                if os.path.exists(face_path):
                    try:
                        os.remove(face_path)
                    except OSError:
                        logger.warning("Could not remove face file %s", face_path, exc_info=True)
            self.load_students(); self.inp_s_msv.clear(); self.inp_s_name.clear(); self.inp_s_class.clear(); self.inp_s_pass.clear(); self.refresh_stats()
            self.log_action(f"Đã xóa sinh viên MSV: {msv}")
        except Exception:
            logger.exception("Failed to delete student %s", msv)
            QMessageBox.critical(self, "Lỗi", f"Không xóa được sinh viên {msv}.")

    def select_student(self, item):
        r = item.row()
        self.inp_s_msv.setText(self.tb_students.item(r, 0).text() if self.tb_students.item(r, 0) else "")
        self.inp_s_name.setText(self.tb_students.item(r, 1).text() if self.tb_students.item(r, 1) else "")
        self.inp_s_class.setText(self.tb_students.item(r, 2).text() if self.tb_students.item(r, 2) else "")
        self.inp_s_pass.clear()
        self.load_student_faces(self.inp_s_msv.text().strip())

    def select_student_face_reference(self):
        targets = self._collect_selected_student_faces()
        self.selected_student_face_targets = targets
        if targets:
            self.selected_student_face_ref_id = targets[0].get("ref_id")
            self.selected_student_face_image = str(targets[0].get("face_image") or "").strip()
        else:
            self.selected_student_face_ref_id = None
            self.selected_student_face_image = ""
        self.btn_delete_face.setEnabled(bool(self.inp_s_msv.text().strip()) and len(targets) > 0)

    def delete_selected_face(self):
        msv = self.inp_s_msv.text().strip()
        if not msv:
            return QMessageBox.warning(self, "Lỗi", "Hãy chọn sinh viên trước khi xóa ảnh tham chiếu.")
        targets = list(self.selected_student_face_targets or self._collect_selected_student_faces())
        if not targets:
            return QMessageBox.warning(self, "Lỗi", "Hãy chọn ít nhất một ảnh tham chiếu trong danh sách để xóa.")

        selected_names = [str(target.get("face_image") or "").strip() for target in targets if str(target.get("face_image") or "").strip()]
        if not selected_names:
            return QMessageBox.warning(self, "Lỗi", "Không đọc được danh sách ảnh đã chọn để xóa.")

        if QMessageBox.question(self, "Xác nhận xóa ảnh", f"Bạn có chắc muốn xóa {len(selected_names)} ảnh tham chiếu của sinh viên {msv}?", QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            return

        db = None
        try:
            current_rows = self._fetch_student_face_rows(msv)
            target_keys = {
                (target.get("ref_id"), str(target.get("face_image") or "").strip())
                for target in targets
            }
            existing_targets = [
                row for row in current_rows
                if (row.get("ref_id"), str(row.get("face_image") or "").strip()) in target_keys
                or (None, str(row.get("face_image") or "").strip()) in target_keys
            ]
            if not existing_targets:
                return QMessageBox.warning(self, "Không tìm thấy ảnh", "Các ảnh tham chiếu đã chọn không còn tồn tại trong hệ thống.")

            deleting_images = {str(row.get("face_image") or "").strip() for row in existing_targets}
            remaining_rows = [row for row in current_rows if str(row.get("face_image") or "").strip() not in deleting_images]
            next_primary = next((row for row in remaining_rows if row.get("is_primary")), None)
            if next_primary is None and remaining_rows:
                next_primary = remaining_rows[0]

            db = self.get_db(); cur = db.cursor(dictionary=True)
            try:
                cur.execute("UPDATE student_face_images SET is_primary = 0 WHERE msv = %s", (msv,))
                for row in existing_targets:
                    row_ref_id = row.get("ref_id")
                    row_image = str(row.get("face_image") or "").strip()
                    if row_ref_id is not None:
                        cur.execute("DELETE FROM student_face_images WHERE ref_id = %s", (row_ref_id,))
                    else:
                        cur.execute("DELETE FROM student_face_images WHERE msv = %s AND face_image = %s", (msv, row_image))
                if next_primary and next_primary.get("ref_id") is not None:
                    cur.execute("UPDATE student_face_images SET is_primary = 1 WHERE ref_id = %s", (next_primary.get("ref_id"),))
            except mysql.connector.Error:
                logger.debug("student_face_images unavailable while deleting face reference", exc_info=True)

            next_primary_image = str(next_primary.get("face_image") or "").strip() if next_primary else None
            cur.execute("UPDATE students SET face_image = %s WHERE msv = %s", (next_primary_image, msv))
            db.commit()
            db.close(); db = None

            missing_files = 0
            failed_remove_files = 0
            for image_name in deleting_images:
                face_path = os.path.join("server_database", image_name)
                if os.path.exists(face_path):
                    try:
                        os.remove(face_path)
                    except OSError:
                        failed_remove_files += 1
                        logger.warning("Could not remove face file %s", face_path, exc_info=True)
                else:
                    missing_files += 1

            self.load_students()
            self.load_student_faces(msv)
            self.log_action(f"Xóa {len(deleting_images)} ảnh tham chiếu của sinh viên {msv}")
            suffix_parts = []
            if missing_files > 0:
                suffix_parts.append(f"{missing_files} file vật lý đã không còn tồn tại")
            if failed_remove_files > 0:
                suffix_parts.append(f"{failed_remove_files} file vật lý không thể xóa (đang mở hoặc thiếu quyền)")
            suffix = "" if not suffix_parts else "\n(" + "; ".join(suffix_parts) + ")"
            QMessageBox.information(self, "Thành công", f"Đã xóa {len(deleting_images)} ảnh tham chiếu của sinh viên {msv}.{suffix}")
        except FileNotFoundError:
            self.load_students()
            self.load_student_faces(msv)
            self.log_action(f"Xóa metadata ảnh tham chiếu của sinh viên {msv}; một số file vật lý không còn tồn tại")
            QMessageBox.information(self, "Thành công", "Đã xóa bản ghi ảnh tham chiếu. Một số file vật lý không còn tồn tại trên đĩa.")
        except Exception:
            if db is not None:
                try:
                    db.rollback()
                    db.close()
                except Exception:
                    logger.debug("Rollback delete face failed", exc_info=True)
            logger.exception("Failed to delete selected face references for %s", msv)
            QMessageBox.critical(self, "Lỗi", f"Không xóa được các ảnh tham chiếu đã chọn của sinh viên {msv}.")

    def upload_face(self):
        msv = self.inp_s_msv.text()
        if not msv: return QMessageBox.warning(self, "Lỗi", "Hãy chọn MSV trước!")
        filepaths, _ = QFileDialog.getOpenFileNames(self, "Chọn ảnh tham chiếu", "", "Images (*.png *.jpg *.jpeg)")
        if not filepaths:
            return
        os.makedirs("server_database", exist_ok=True)
        try:
            db = self.get_db(); cur = db.cursor(dictionary=True)
            cur.execute("SELECT face_image FROM students WHERE msv = %s", (msv,))
            student_row = cur.fetchone()
            primary_face = str(student_row.get("face_image") or "") if student_row else ""

            inserted_count = 0
            first_saved_name = None
            for index, filepath in enumerate(filepaths, start=1):
                ext = os.path.splitext(filepath)[1].lower() or ".jpg"
                timestamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
                new_name = f"{msv}_ref_{timestamp}_{index}{ext}"
                shutil.copy(filepath, os.path.join("server_database", new_name))
                if first_saved_name is None:
                    first_saved_name = new_name
                try:
                    cur.execute(
                        "INSERT INTO student_face_images (msv, face_image, is_primary) VALUES (%s, %s, %s)",
                        (msv, new_name, 1 if not primary_face and inserted_count == 0 else 0),
                    )
                    inserted_count += 1
                except mysql.connector.Error:
                    logger.debug("student_face_images unavailable; skipping multi-reference insert", exc_info=True)

            if not primary_face and first_saved_name:
                cur.execute("UPDATE students SET face_image=%s WHERE msv=%s", (first_saved_name, msv))
            db.commit(); db.close(); self.load_students()
            self.log_action(f"Cập nhật {inserted_count or len(filepaths)} ảnh tham chiếu cho MSV: {msv}")
            QMessageBox.information(self, "OK", f"Đã nạp {len(filepaths)} ảnh tham chiếu cho sinh viên {msv}.")
        except Exception:
            logger.exception("Failed to upload face image for %s", msv)
            QMessageBox.critical(self, "Lỗi", f"Không cập nhật được ảnh khuôn mặt cho {msv}.")

    # ---------------- TAB 3: QUẢN LÝ GIÁM THỊ ----------------
    def setup_tab_proctors(self):
        tab = QWidget(); layout = QHBoxLayout(tab)
        form_gb = QGroupBox("Thông tin Giám thị")
        form_gb.setFixedWidth(350); fl = QVBoxLayout(form_gb)
        self.inp_p_id = QLineEdit(); self.inp_p_id.setPlaceholderText("ID (VD: GV02)")
        self.inp_p_name = QLineEdit(); self.inp_p_name.setPlaceholderText("Họ tên")
        self.inp_p_pass = QLineEdit(); self.inp_p_pass.setPlaceholderText("Mật khẩu")
        btn_add = QPushButton("THÊM / CẬP NHẬT"); btn_add.clicked.connect(self.save_proctor)
        btn_delete = QPushButton("XÓA GIÁM THỊ")
        btn_delete.setStyleSheet("background-color: #1A1A1A;")
        btn_delete.clicked.connect(self.delete_proctor)
        fl.addWidget(self.inp_p_id); fl.addWidget(self.inp_p_name); fl.addWidget(self.inp_p_pass); fl.addWidget(btn_add); fl.addWidget(btn_delete); fl.addStretch()
        
        right_layout = QVBoxLayout()
        self.proctor_filter_bar = FilterBar("Tìm theo mã hoặc tên giám thị", "Tìm kiếm", "Tải lại")
        wire_filter_bar(self.proctor_filter_bar, self.search_proctors, self.load_proctors)
        self.tb_proctors = QTableWidget(); self.tb_proctors.setColumnCount(3)
        self.tb_proctors.setHorizontalHeaderLabels(["ID Giám thị", "Họ tên", "Mật khẩu"])
        self.tb_proctors.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tb_proctors.setColumnHidden(2, True)
        self.tb_proctors.itemClicked.connect(self.select_proctor)
        polish_table(self.tb_proctors)
        self.proctors_empty = EmptyState("Chưa có giám thị", "Tạo tài khoản giám thị để phân công coi thi và quản lý bộ đề.")
        right_layout.addWidget(self.proctor_filter_bar)
        right_layout.addWidget(self.proctors_empty)
        right_layout.addWidget(self.tb_proctors)
        layout.addWidget(form_gb); layout.addLayout(right_layout)
        self.tabs.addTab(tab, "👨‍🏫 Giám thị")

    def load_proctors(self):
        try:
            db = self.get_db(); cur = db.cursor(dictionary=True)
            cur.execute("SELECT * FROM proctors"); rows = cur.fetchall(); db.close()
            self._populate_proctors_table(rows)
            self._populate_template_proctor_choices(rows)
        except Exception:
            logger.exception("Failed to load proctors")
            QMessageBox.warning(self, "Lỗi", "Không tải được danh sách giám thị.")

    def search_proctors(self):
        keyword = self.proctor_filter_bar.search_input.text().strip()
        if not keyword:
            return self.load_proctors()
        try:
            db = self.get_db(); cur = db.cursor(dictionary=True)
            cur.execute("SELECT * FROM proctors WHERE proctor_id LIKE %s OR full_name LIKE %s", (f"%{keyword}%", f"%{keyword}%"))
            rows = cur.fetchall(); db.close()
            self._populate_proctors_table(rows)
        except Exception as exc:
            self._show_dialog("Không thể tìm giám thị", str(exc))

    def select_proctor(self, item):
        row = item.row()
        self.inp_p_id.setText(self.tb_proctors.item(row, 0).text() if self.tb_proctors.item(row, 0) else "")
        self.inp_p_name.setText(self.tb_proctors.item(row, 1).text() if self.tb_proctors.item(row, 1) else "")
        self.inp_p_pass.clear()

    def save_proctor(self):
        pid = self.inp_p_id.text().strip()
        name = self.inp_p_name.text().strip()
        pw = self.inp_p_pass.text().strip()
        if not pid or not name:
            return QMessageBox.warning(self, "Lỗi", "Vui lòng nhập đủ ID và họ tên giám thị.")
        try:
            db = self.get_db()
            cur = db.cursor(dictionary=True)
            cur.execute("SELECT password FROM proctors WHERE proctor_id = %s", (pid,))
            existing_row = cur.fetchone()
            if existing_row and not pw:
                password_to_store = existing_row["password"]
            elif pw:
                password_to_store = hash_password(pw)
            else:
                db.close()
                return QMessageBox.warning(self, "Lỗi", "Vui lòng nhập mật khẩu cho giám thị mới.")

            cur.execute(
                """
                INSERT INTO proctors (proctor_id, full_name, password)
                VALUES (%s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    full_name = VALUES(full_name),
                    password = VALUES(password)
                """,
                (pid, name, password_to_store),
            )
            db.commit()
            db.close()
            self.load_proctors()
            self.refresh_stats()
            self._clear_proctor_form()
            self.log_action(f"Cập nhật tài khoản Giám thị: {pid}")
        except Exception:
            logger.exception("Failed to save proctor %s", pid)
            QMessageBox.critical(self, "Lỗi", f"Không lưu được giám thị {pid}.")

    def delete_proctor(self):
        pid = self.inp_p_id.text().strip()
        if not pid:
            return QMessageBox.warning(self, "Lỗi", "Hãy chọn giám thị cần xóa trong bảng hoặc nhập ID giám thị.")
        if QMessageBox.question(self, "Xác nhận xóa", f"Bạn có chắc muốn xóa giám thị {pid}?", QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            return
        try:
            db = self.get_db(); cur = db.cursor(dictionary=True)
            cur.execute("SELECT COUNT(*) AS total FROM exam_templates WHERE proctor_id = %s", (pid,))
            template_count = int((cur.fetchone() or {}).get("total", 0))
            if template_count > 0:
                db.close()
                return QMessageBox.warning(self, "Không thể xóa", f"Giám thị {pid} đang phụ trách {template_count} bộ đề. Hãy chuyển bộ đề sang giám thị khác trước.")
            cur = db.cursor()
            cur.execute("DELETE FROM proctors WHERE proctor_id = %s", (pid,))
            db.commit(); db.close(); self.load_proctors(); self.refresh_stats(); self._clear_proctor_form()
            self.log_action(f"Xóa tài khoản Giám thị: {pid}")
            QMessageBox.information(self, "Thành công", f"Đã xóa giám thị {pid}.")
        except Exception:
            logger.exception("Failed to delete proctor %s", pid)
            QMessageBox.critical(self, "Lỗi", f"Không xóa được giám thị {pid}.")

    # ---------------- TAB 4: MASTER CONTROL ----------------
    def setup_tab_master(self):
        tab = QWidget(); layout = QVBoxLayout(tab)
        self.class_filter_bar = FilterBar("Lọc theo mã lớp, tên môn hoặc trạng thái", "Tìm kiếm", "Tải lại")
        wire_filter_bar(self.class_filter_bar, self.search_classes, self.load_classes)
        self.tb_classes = QTableWidget(); self.tb_classes.setColumnCount(4)
        self.tb_classes.setHorizontalHeaderLabels(["ID Lớp", "Tên Môn", "Trạng thái", "Hành động"])
        self.tb_classes.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        polish_table(self.tb_classes)
        self.classes_empty = EmptyState("Không có lớp thi phù hợp", "Tạo lớp mới hoặc bỏ bộ lọc để xem toàn bộ danh sách.")
        layout.addWidget(self.class_filter_bar); layout.addWidget(self.classes_empty); layout.addWidget(self.tb_classes)
        self.tabs.addTab(tab, "🏢 Lớp thi (Master)")

    def load_classes(self):
        try:
            db = self.get_db(); cur = db.cursor(dictionary=True)
            cur.execute("SELECT * FROM classes ORDER BY created_at DESC"); rows = cur.fetchall(); db.close()
            self._populate_classes_table(rows)
        except Exception:
            logger.exception("Failed to load classes")
            QMessageBox.warning(self, "Lỗi", "Không tải được danh sách lớp thi.")

    def search_classes(self):
        keyword = self.class_filter_bar.search_input.text().strip()
        if not keyword:
            return self.load_classes()
        try:
            db = self.get_db(); cur = db.cursor(dictionary=True)
            cur.execute(
                "SELECT * FROM classes WHERE CAST(class_id AS CHAR) LIKE %s OR class_name LIKE %s OR status LIKE %s ORDER BY created_at DESC",
                (f"%{keyword}%", f"%{keyword}%", f"%{keyword}%"),
            )
            rows = cur.fetchall(); db.close()
            self._populate_classes_table(rows)
        except Exception as exc:
            self._show_dialog("Không thể tìm lớp", str(exc))

    def force_close_class(self, class_id):
        try:
            if not self.admin_token:
                return self._show_dialog("Thiếu phiên quản trị", "Bạn cần đăng nhập quản trị qua API để khóa phòng thi.")
            response = requests.post(
                f"{self.api_url}/api/monitor/admin/classes/{class_id}/lock",
                headers=self._admin_api_headers(),
                timeout=5,
            )
            response.raise_for_status()
            payload = response.json()
            if payload.get("status") == "success":
                self.load_classes(); self.refresh_stats()
                QMessageBox.information(self, "OK", payload.get("message", f"Đã khóa phòng thi của lớp {class_id}"))
            else:
                self._show_dialog("Không khóa được phòng thi", payload.get("message", f"Hệ thống từ chối khóa phòng thi của lớp {class_id}."))
        except Exception:
            logger.exception("Failed to close class %s", class_id)
            QMessageBox.critical(self, "Lỗi", f"Không khóa được phòng thi của lớp {class_id}.")

    def reopen_class(self, class_id):
        try:
            if not self.admin_token:
                return self._show_dialog("Thiếu phiên quản trị", "Bạn cần đăng nhập quản trị qua API để mở lại phòng thi.")
            response = requests.post(
                f"{self.api_url}/api/monitor/admin/classes/{class_id}/unlock",
                headers=self._admin_api_headers(),
                timeout=5,
            )
            response.raise_for_status()
            payload = response.json()
            if payload.get("status") == "success":
                self.load_classes(); self.refresh_stats()
                QMessageBox.information(self, "OK", payload.get("message", f"Đã mở lại phòng thi của lớp {class_id}"))
            else:
                self._show_dialog("Không mở được phòng thi", payload.get("message", f"Hệ thống từ chối mở lại phòng thi của lớp {class_id}."))
        except Exception:
            logger.exception("Failed to reopen class %s", class_id)
            QMessageBox.critical(self, "Lỗi", f"Không mở lại được phòng thi của lớp {class_id}.")

    def setup_tab_monitor(self):
        tab = QWidget(); layout = QHBoxLayout(tab)
        left_layout = QVBoxLayout()
        self.monitor_filter_bar = FilterBar("Lọc theo MSV, lớp hoặc giám thị", "Tìm kiếm", "Tải lại")
        wire_filter_bar(self.monitor_filter_bar, self.apply_monitor_filter, self.load_monitor_overview)
        self.tb_monitor = QTableWidget(); self.tb_monitor.setColumnCount(7)
        self.tb_monitor.setHorizontalHeaderLabels(["MSV", "Sinh viên", "Lớp", "Giám thị", "Risk", "Cảnh báo phiên", "Cập nhật"])
        self.tb_monitor.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        for section in [1, 2, 3]:
            self.tb_monitor.horizontalHeader().setSectionResizeMode(section, QHeaderView.Stretch)
        for section in [4, 5, 6]:
            self.tb_monitor.horizontalHeader().setSectionResizeMode(section, QHeaderView.ResizeToContents)
        self.tb_monitor.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tb_monitor.setSelectionMode(QAbstractItemView.SingleSelection)
        self.tb_monitor.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tb_monitor.itemSelectionChanged.connect(self.load_selected_monitor_detail)
        polish_table(self.tb_monitor)
        self.monitor_empty = EmptyState("Chưa có phiên AI hoạt động", "Metadata của các snapshot audit sẽ xuất hiện tại đây khi sinh viên vào phòng thi.")
        left_layout.addWidget(self.monitor_filter_bar)
        left_layout.addWidget(self.monitor_empty)
        left_layout.addWidget(self.tb_monitor, 1)

        detail_box = QGroupBox("Audit snapshot")
        detail_layout = QVBoxLayout(detail_box)
        self.lbl_monitor_meta = QLabel("Phiên thi: --")
        self.lbl_monitor_risk = QLabel("Mức rủi ro: --")
        self.lbl_monitor_state = QLabel("Trạng thái AI: --")
        self.lbl_monitor_update = QLabel("Cập nhật cuối: --")
        for info_label in [self.lbl_monitor_meta, self.lbl_monitor_risk, self.lbl_monitor_state, self.lbl_monitor_update]:
            info_label.setStyleSheet("color:#202020; font-size:13px;")
        self.pb_monitor_total_risk = QProgressBar(); self.pb_monitor_total_risk.setRange(0, 100); self.pb_monitor_total_risk.setValue(0); self.pb_monitor_total_risk.setFormat("%p%")
        self.txt_monitor_summary = QTextEdit(); self.txt_monitor_summary.setReadOnly(True); self.txt_monitor_summary.setFixedHeight(140)
        detail_layout.addWidget(self.lbl_monitor_meta)
        detail_layout.addWidget(self.lbl_monitor_risk)
        detail_layout.addWidget(self.lbl_monitor_state)
        detail_layout.addWidget(self.lbl_monitor_update)
        detail_layout.addWidget(QLabel("Tổng mức rủi ro"))
        detail_layout.addWidget(self.pb_monitor_total_risk)
        detail_layout.addWidget(QLabel("Tóm tắt audit"))
        detail_layout.addWidget(self.txt_monitor_summary)

        layout.addLayout(left_layout, 3)
        layout.addWidget(detail_box, 4)
        self.tabs.addTab(tab, "🛰 Audit snapshot")

    def _populate_monitor_table(self, rows):
        self.tb_monitor.setRowCount(len(rows))
        self.monitor_empty.setVisible(len(rows) == 0)
        self.tb_monitor.setVisible(len(rows) > 0)
        for row_index, row in enumerate(rows):
            msv_item = QTableWidgetItem(str(row.get("msv", "")))
            msv_item.setData(Qt.UserRole, row.get("session_token"))
            self.tb_monitor.setItem(row_index, 0, msv_item)
            self.tb_monitor.setItem(row_index, 1, QTableWidgetItem(str(row.get("full_name", ""))))
            self.tb_monitor.setItem(row_index, 2, QTableWidgetItem(f"{row.get('class_id', '')} | {row.get('class_name', '')}"))
            self.tb_monitor.setItem(row_index, 3, QTableWidgetItem(str(row.get("proctor_id", ""))))
            self.tb_monitor.setItem(row_index, 4, QTableWidgetItem(f"{row.get('risk_score', 0):.0f}%"))
            self.tb_monitor.setItem(row_index, 5, QTableWidgetItem(f"{row.get('session_warning_count', 0)}/{row.get('max_warnings', 0)}"))
            self.tb_monitor.setItem(row_index, 6, QTableWidgetItem(str(row.get("updated_at", ""))))

    def apply_monitor_filter(self):
        keyword = self.monitor_filter_bar.search_input.text().strip().lower()
        if not keyword:
            return self._populate_monitor_table(self.monitor_entries)
        filtered_rows = [
            row for row in self.monitor_entries
            if keyword in str(row.get("msv", "")).lower()
            or keyword in str(row.get("full_name", "")).lower()
            or keyword in str(row.get("class_name", "")).lower()
            or keyword in str(row.get("class_id", "")).lower()
            or keyword in str(row.get("proctor_id", "")).lower()
        ]
        self._populate_monitor_table(filtered_rows)

    def load_monitor_overview(self):
        if self.monitor_overview_inflight:
            return
        self.monitor_overview_inflight = True
        self._run_background_task(
            self._fetch_admin_monitor_overview,
            self._on_admin_monitor_overview_loaded,
            on_error=self._on_admin_monitor_overview_error,
            on_finished=lambda: setattr(self, "monitor_overview_inflight", False),
        )

    def load_selected_monitor_detail(self):
        row = self.tb_monitor.currentRow()
        if row < 0:
            return
        item = self.tb_monitor.item(row, 0)
        if item is None:
            return
        session_token = item.data(Qt.UserRole)
        if not session_token:
            return
        self.current_monitor_session_token = session_token
        self._run_background_task(
            lambda: self._fetch_admin_monitor_detail(session_token),
            lambda payload: self._apply_monitor_detail_payload(payload["session_token"], payload["detail"]),
            on_error=lambda payload: logger.error("Admin monitor detail failed for %s: %s", session_token, payload.get("traceback") or payload.get("message")),
        )

    # ---------------- TAB 5: BỘ ĐỀ TOÀN HỆ THỐNG ----------------
    def setup_tab_templates(self):
        tab = QWidget(); layout = QVBoxLayout(tab)
        workspace_tabs = QTabWidget()
        workspace_tabs.setDocumentMode(True)
        workspace_tabs.setStyleSheet("QTabBar::tab { padding: 10px 16px; }")

        left_layout = QVBoxLayout()
        gb_new_template = QGroupBox("Khởi tạo Bộ đề mới")
        fl_template = QFormLayout(gb_new_template)
        self.cb_template_proctor = QComboBox()
        self.inp_template_name = QLineEdit(); self.inp_template_name.setPlaceholderText("Tên Bộ đề")
        btn_create_template = QPushButton("TẠO BỘ ĐỀ")
        btn_create_template.clicked.connect(self.create_exam_template)
        fl_template.addRow("Gán cho giám thị:", self.cb_template_proctor)
        fl_template.addRow("Tên bộ đề:", self.inp_template_name)
        fl_template.addRow(btn_create_template)

        gb_manage_template = QGroupBox("Quản lý bộ đề đã có")
        manage_layout = QFormLayout(gb_manage_template)
        self.lbl_manage_template_id = QLabel("Chưa chọn bộ đề")
        self.inp_manage_template_name = QLineEdit(); self.inp_manage_template_name.setPlaceholderText("Tên bộ đề")
        self.cb_manage_template_proctor = QComboBox()
        btn_update_template = QPushButton("LƯU THÔNG TIN BỘ ĐỀ")
        btn_update_template.clicked.connect(self.update_selected_template)
        btn_delete_template = QPushButton("XÓA BỘ ĐỀ ĐÃ CHỌN")
        btn_delete_template.setStyleSheet("background-color: #232323;")
        btn_delete_template.clicked.connect(self.delete_selected_template)
        manage_actions = QHBoxLayout()
        manage_actions.addWidget(btn_update_template)
        manage_actions.addWidget(btn_delete_template)
        manage_layout.addRow("Bộ đề đang chọn:", self.lbl_manage_template_id)
        manage_layout.addRow("Tên bộ đề:", self.inp_manage_template_name)
        manage_layout.addRow("Giám thị phụ trách:", self.cb_manage_template_proctor)
        manage_layout.addRow(manage_actions)

        gb_import = QGroupBox("Nhập / Xuất Bộ đề")
        import_layout = QVBoxLayout(gb_import)
        btn_import_template = QPushButton("📥 NHẬP TỪ EXCEL/CSV")
        btn_import_template.clicked.connect(self.import_template_file)
        btn_export_template = QPushButton("💾 XUẤT BỘ ĐỀ RA CSV")
        btn_export_template.setStyleSheet("background-color: #2A2A2A;")
        btn_export_template.clicked.connect(self.export_template_csv)
        import_layout.addWidget(btn_import_template)
        import_layout.addWidget(btn_export_template)

        gb_template_list = QGroupBox("Danh sách Bộ đề")
        template_list_layout = QVBoxLayout(gb_template_list)
        self.template_filter_bar = FilterBar("Lọc theo tên bộ đề hoặc giám thị", "Tìm kiếm", "Tải lại")
        wire_filter_bar(self.template_filter_bar, self.search_templates, self.load_templates)
        self.tb_templates = QTableWidget(); self.tb_templates.setColumnCount(4)
        self.tb_templates.setHorizontalHeaderLabels(["ID", "Tên bộ đề", "Giám thị", "Số câu"])
        self.tb_templates.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.tb_templates.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.tb_templates.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.tb_templates.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.tb_templates.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tb_templates.setSelectionMode(QAbstractItemView.SingleSelection)
        self.tb_templates.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tb_templates.itemSelectionChanged.connect(self.select_template_from_catalog)
        polish_table(self.tb_templates)
        self.templates_empty = EmptyState("Chưa có bộ đề", "Admin có thể tạo mới hoặc nhập câu hỏi từ file Excel/CSV.")
        template_list_layout.addWidget(self.template_filter_bar)
        template_list_layout.addWidget(self.templates_empty)
        template_list_layout.addWidget(self.tb_templates)

        left_layout.addWidget(gb_new_template)
        left_layout.addWidget(gb_manage_template)
        left_layout.addWidget(gb_import)
        left_layout.addWidget(gb_template_list, 1)

        gb_questions = QGroupBox("Chi tiết Bộ đề")
        question_layout = QVBoxLayout(gb_questions)
        self.cb_templates_for_question = QComboBox()
        self.cb_templates_for_question.currentIndexChanged.connect(self.refresh_template_details)
        self.inp_q = QTextEdit(); self.inp_q.setPlaceholderText("Nhập nội dung câu hỏi..."); self.inp_q.setFixedHeight(80)
        self.inp_a = QLineEdit(); self.inp_a.setPlaceholderText("Đáp án A")
        self.inp_b = QLineEdit(); self.inp_b.setPlaceholderText("Đáp án B")
        self.inp_c = QLineEdit(); self.inp_c.setPlaceholderText("Đáp án C")
        self.inp_d = QLineEdit(); self.inp_d.setPlaceholderText("Đáp án D")
        answer_layout = QHBoxLayout()
        answer_layout.addWidget(QLabel("Đáp án đúng:"))
        self.inp_correct = QComboBox(); self.inp_correct.addItems(["A", "B", "C", "D"])
        answer_layout.addWidget(self.inp_correct)
        answer_layout.addWidget(QLabel("Điểm số:"))
        self.inp_points = QDoubleSpinBox(); self.inp_points.setRange(0.1, 10.0); self.inp_points.setSingleStep(0.25); self.inp_points.setValue(1.0)
        answer_layout.addWidget(self.inp_points)
        answer_layout.addStretch()
        self.btn_save_question = QPushButton("THÊM CÂU HỎI VÀO BỘ ĐỀ NÀY")
        self.btn_save_question.clicked.connect(self.save_question)
        btn_delete_question = QPushButton("XÓA CÂU HỎI ĐÃ CHỌN")
        btn_delete_question.setStyleSheet("background-color: #232323;")
        btn_delete_question.clicked.connect(self.delete_selected_question)
        question_actions = QHBoxLayout()
        question_actions.addWidget(self.btn_save_question)
        question_actions.addWidget(btn_delete_question)
        self.lbl_template_detail = QLabel("Chọn bộ đề để xem nội dung")
        self.lbl_template_detail.setStyleSheet("font-weight:bold; color:#111111; padding-top:12px;")
        self.tb_template_questions = QTableWidget(); self.tb_template_questions.setColumnCount(8)
        self.tb_template_questions.setHorizontalHeaderLabels(["ID", "Câu hỏi", "A", "B", "C", "D", "Đúng", "Điểm"])
        self.tb_template_questions.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.tb_template_questions.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        for section in [2, 3, 4, 5]:
            self.tb_template_questions.horizontalHeader().setSectionResizeMode(section, QHeaderView.Stretch)
        self.tb_template_questions.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeToContents)
        self.tb_template_questions.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeToContents)
        self.tb_template_questions.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tb_template_questions.setSelectionMode(QAbstractItemView.SingleSelection)
        self.tb_template_questions.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tb_template_questions.itemSelectionChanged.connect(self._load_question_into_form)
        polish_table(self.tb_template_questions)
        self.template_detail_empty = EmptyState("Chưa có nội dung bộ đề", "Chọn bộ đề ở danh sách bên trái để xem và chỉnh sửa nội dung.")
        question_layout.addWidget(QLabel("Chọn Bộ Đề đang soạn thảo:"))
        question_layout.addWidget(self.cb_templates_for_question)
        question_layout.addWidget(self.inp_q)
        question_layout.addWidget(self.inp_a)
        question_layout.addWidget(self.inp_b)
        question_layout.addWidget(self.inp_c)
        question_layout.addWidget(self.inp_d)
        question_layout.addLayout(answer_layout)
        question_layout.addLayout(question_actions)
        question_layout.addWidget(self.lbl_template_detail)
        question_layout.addWidget(self.template_detail_empty)
        question_layout.addWidget(self.tb_template_questions, 1)

        template_page = QWidget()
        template_page_layout = QVBoxLayout(template_page)
        template_page_layout.addWidget(gb_new_template)
        template_page_layout.addWidget(gb_import)
        template_page_layout.addWidget(gb_template_list, 1)

        question_page = QWidget()
        question_page_layout = QVBoxLayout(question_page)
        question_page_layout.addWidget(gb_questions, 1)

        workspace_tabs.addTab(template_page, "Danh mục bộ đề")
        workspace_tabs.addTab(question_page, "Soạn câu hỏi")
        layout.addWidget(workspace_tabs)
        self.tabs.addTab(tab, "📝 Bộ đề")

    def _clear_question_form(self):
        self.admin_current_question_id = None
        self.inp_q.clear(); self.inp_a.clear(); self.inp_b.clear(); self.inp_c.clear(); self.inp_d.clear()
        self.inp_correct.setCurrentText("A")
        self.inp_points.setValue(1.0)
        self.btn_save_question.setText("THÊM CÂU HỎI VÀO BỘ ĐỀ NÀY")

    def _clear_template_management_form(self):
        self.lbl_manage_template_id.setText("Chưa chọn bộ đề")
        self.inp_manage_template_name.clear()
        if self.cb_manage_template_proctor.count() > 0:
            self.cb_manage_template_proctor.setCurrentIndex(0)

    def _set_combo_data(self, combo, value):
        for index in range(combo.count()):
            if combo.itemData(index) == value:
                combo.setCurrentIndex(index)
                return

    def _load_template_metadata(self, template_id):
        db = self.get_db(); cur = db.cursor(dictionary=True)
        try:
            cur.execute("SELECT template_id, template_name, proctor_id FROM exam_templates WHERE template_id = %s LIMIT 1", (template_id,))
            return cur.fetchone()
        finally:
            db.close()

    def _apply_template_management_selection(self, template_row):
        if not template_row:
            self._clear_template_management_form()
            return
        self.lbl_manage_template_id.setText(f"ID {template_row['template_id']}")
        self.inp_manage_template_name.setText(str(template_row.get('template_name') or ''))
        self._set_combo_data(self.cb_manage_template_proctor, template_row.get('proctor_id'))

    def _selected_template_id(self):
        selected_template = self.cb_templates_for_question.currentText().strip()
        return self.template_dict.get(selected_template)

    def _set_active_template(self, template_id):
        for combo in [self.cb_templates_for_question]:
            for index in range(combo.count()):
                if self.template_dict.get(combo.itemText(index)) == template_id:
                    combo.setCurrentIndex(index)
                    break

    def _load_question_into_form(self):
        row = self.tb_template_questions.currentRow()
        if row < 0:
            self._clear_question_form()
            return
        item = self.tb_template_questions.item(row, 0)
        if item is None:
            self._clear_question_form()
            return
        self.admin_current_question_id = int(item.text())
        self.inp_q.setPlainText(self.tb_template_questions.item(row, 1).text() if self.tb_template_questions.item(row, 1) else "")
        self.inp_a.setText(self.tb_template_questions.item(row, 2).text() if self.tb_template_questions.item(row, 2) else "")
        self.inp_b.setText(self.tb_template_questions.item(row, 3).text() if self.tb_template_questions.item(row, 3) else "")
        self.inp_c.setText(self.tb_template_questions.item(row, 4).text() if self.tb_template_questions.item(row, 4) else "")
        self.inp_d.setText(self.tb_template_questions.item(row, 5).text() if self.tb_template_questions.item(row, 5) else "")
        self.inp_correct.setCurrentText(self.tb_template_questions.item(row, 6).text() if self.tb_template_questions.item(row, 6) else "A")
        try:
            self.inp_points.setValue(float(self.tb_template_questions.item(row, 7).text()))
        except Exception:
            self.inp_points.setValue(1.0)
        self.btn_save_question.setText("CẬP NHẬT CÂU HỎI ĐÃ CHỌN")

    def load_template_proctors(self):
        try:
            db = self.get_db(); cur = db.cursor(dictionary=True)
            cur.execute("SELECT proctor_id, full_name FROM proctors ORDER BY proctor_id")
            rows = cur.fetchall(); db.close()
            self._populate_template_proctor_choices(rows)
        except Exception:
            logger.exception("Failed to load proctors for template assignment")

    def load_templates(self):
        try:
            db = self.get_db(); cur = db.cursor(dictionary=True)
            cur.execute(
                """
                SELECT et.template_id, et.template_name, COALESCE(p.full_name, et.proctor_id, '') AS proctor_name, COUNT(q.q_id) AS question_count
                FROM exam_templates et
                LEFT JOIN proctors p ON p.proctor_id = et.proctor_id
                LEFT JOIN question_bank q ON q.template_id = et.template_id
                GROUP BY et.template_id, et.template_name, p.full_name, et.proctor_id
                ORDER BY et.template_id DESC
                """
            )
            rows = cur.fetchall(); db.close()
            current_template_id = self._selected_template_id()
            self.cb_templates_for_question.clear()
            self.tb_templates.setRowCount(len(rows))
            self.template_dict.clear()
            for row_index, row in enumerate(rows):
                label = f"[{row['template_id']}] {row['template_name']}"
                self.template_dict[label] = row['template_id']
                self.cb_templates_for_question.addItem(label)
                self.tb_templates.setItem(row_index, 0, QTableWidgetItem(str(row['template_id'])))
                self.tb_templates.setItem(row_index, 1, QTableWidgetItem(str(row['template_name'])))
                self.tb_templates.setItem(row_index, 2, QTableWidgetItem(str(row.get('proctor_name') or '')))
                self.tb_templates.setItem(row_index, 3, QTableWidgetItem(str(row.get('question_count', 0))))
            self.templates_empty.setVisible(len(rows) == 0)
            self.tb_templates.setVisible(len(rows) > 0)
            if current_template_id is not None:
                self._set_active_template(current_template_id)
            if self.cb_templates_for_question.count() > 0:
                if self.cb_templates_for_question.currentIndex() < 0:
                    self.cb_templates_for_question.setCurrentIndex(0)
            self.refresh_template_details()
            if current_template_id is not None:
                self._apply_template_management_selection(self._load_template_metadata(current_template_id))
            elif rows:
                self._apply_template_management_selection(self._load_template_metadata(rows[0]['template_id']))
            else:
                self._clear_template_management_form()
        except Exception:
            logger.exception("Failed to load templates for admin")
            self._show_dialog("Không thể tải bộ đề", "Không tải được danh sách bộ đề toàn hệ thống.")

    def search_templates(self):
        keyword = self.template_filter_bar.search_input.text().strip()
        if not keyword:
            return self.load_templates()
        try:
            db = self.get_db(); cur = db.cursor(dictionary=True)
            cur.execute(
                """
                SELECT et.template_id, et.template_name, COALESCE(p.full_name, et.proctor_id, '') AS proctor_name, COUNT(q.q_id) AS question_count
                FROM exam_templates et
                LEFT JOIN proctors p ON p.proctor_id = et.proctor_id
                LEFT JOIN question_bank q ON q.template_id = et.template_id
                WHERE et.template_name LIKE %s OR COALESCE(p.full_name, et.proctor_id, '') LIKE %s
                GROUP BY et.template_id, et.template_name, p.full_name, et.proctor_id
                ORDER BY et.template_id DESC
                """,
                (f"%{keyword}%", f"%{keyword}%"),
            )
            rows = cur.fetchall(); db.close()
            self.tb_templates.setRowCount(len(rows))
            for row_index, row in enumerate(rows):
                self.tb_templates.setItem(row_index, 0, QTableWidgetItem(str(row['template_id'])))
                self.tb_templates.setItem(row_index, 1, QTableWidgetItem(str(row['template_name'])))
                self.tb_templates.setItem(row_index, 2, QTableWidgetItem(str(row.get('proctor_name') or '')))
                self.tb_templates.setItem(row_index, 3, QTableWidgetItem(str(row.get('question_count', 0))))
            self.templates_empty.setVisible(len(rows) == 0)
            self.tb_templates.setVisible(len(rows) > 0)
        except Exception as exc:
            self._show_dialog("Không thể tìm bộ đề", str(exc))

    def select_template_from_catalog(self):
        row = self.tb_templates.currentRow()
        if row < 0:
            return
        item = self.tb_templates.item(row, 0)
        if item is None:
            return
        template_id = int(item.text())
        self._set_active_template(template_id)
        self._apply_template_management_selection(self._load_template_metadata(template_id))
        self.refresh_template_details()
        self._clear_question_form()

    def _load_template_questions(self, template_id):
        db = self.get_db(); cur = db.cursor(dictionary=True)
        try:
            cur.execute("SELECT q_id, question_text, option_a, option_b, option_c, option_d, correct_option, points FROM question_bank WHERE template_id = %s ORDER BY q_id", (template_id,))
            return cur.fetchall()
        finally:
            db.close()

    def _populate_template_questions(self, rows, template_label=""):
        self.tb_template_questions.setRowCount(len(rows))
        for index, row in enumerate(rows):
            self.tb_template_questions.setItem(index, 0, QTableWidgetItem(str(row.get('q_id', ''))))
            self.tb_template_questions.setItem(index, 1, QTableWidgetItem(str(row.get('question_text', ''))))
            self.tb_template_questions.setItem(index, 2, QTableWidgetItem(str(row.get('option_a', '') or '')))
            self.tb_template_questions.setItem(index, 3, QTableWidgetItem(str(row.get('option_b', '') or '')))
            self.tb_template_questions.setItem(index, 4, QTableWidgetItem(str(row.get('option_c', '') or '')))
            self.tb_template_questions.setItem(index, 5, QTableWidgetItem(str(row.get('option_d', '') or '')))
            self.tb_template_questions.setItem(index, 6, QTableWidgetItem(str(row.get('correct_option', ''))))
            self.tb_template_questions.setItem(index, 7, QTableWidgetItem(str(row.get('points', ''))))
        count = len(rows)
        self.lbl_template_detail.setText(f"{template_label} | {count} câu hỏi" if template_label else f"{count} câu hỏi")
        self.template_detail_empty.setVisible(count == 0)
        self.tb_template_questions.setVisible(count > 0)

    def refresh_template_details(self):
        template_id = self._selected_template_id()
        if template_id is None:
            self.lbl_template_detail.setText("Chưa chọn bộ đề")
            self.template_detail_empty.setVisible(True)
            self.tb_template_questions.setVisible(False)
            self.tb_template_questions.setRowCount(0)
            return
        rows = self._load_template_questions(template_id)
        self._populate_template_questions(rows, self.cb_templates_for_question.currentText())

    def create_exam_template(self):
        template_name = self.inp_template_name.text().strip()
        proctor_id = self.cb_template_proctor.currentData()
        if not template_name:
            return QMessageBox.warning(self, "Lỗi", "Vui lòng nhập tên Bộ đề.")
        try:
            db = self.get_db(); cur = db.cursor()
            try:
                # Newer schemas require exam_name (NOT NULL), keep both names in sync.
                cur.execute(
                    "INSERT INTO exam_templates (proctor_id, template_name, exam_name) VALUES (%s, %s, %s)",
                    (proctor_id, template_name, template_name),
                )
            except Exception:
                # Backward compatibility for schemas that only have template_name.
                cur.execute("INSERT INTO exam_templates (proctor_id, template_name) VALUES (%s, %s)", (proctor_id, template_name))
            template_id = cur.lastrowid
            db.commit(); db.close()
            self.inp_template_name.clear()
            self.load_templates(); self._set_active_template(template_id); self.refresh_template_details(); self.refresh_stats()
            self.log_action(f"Tạo bộ đề mới ID {template_id}")
            QMessageBox.information(self, "Thành công", f"Đã tạo bộ đề: {template_name}")
        except Exception as exc:
            self._show_dialog("Không thể tạo bộ đề", str(exc))

    def update_selected_template(self):
        template_id = self._selected_template_id()
        if template_id is None:
            return QMessageBox.warning(self, "Lỗi", "Vui lòng chọn bộ đề cần cập nhật.")
        template_name = self.inp_manage_template_name.text().strip()
        proctor_id = self.cb_manage_template_proctor.currentData()
        if not template_name:
            return QMessageBox.warning(self, "Lỗi", "Vui lòng nhập tên bộ đề.")
        try:
            db = self.get_db(); cur = db.cursor()
            try:
                cur.execute(
                    "UPDATE exam_templates SET template_name = %s, exam_name = %s, proctor_id = %s WHERE template_id = %s",
                    (template_name, template_name, proctor_id, template_id),
                )
            except Exception:
                cur.execute(
                    "UPDATE exam_templates SET template_name = %s, proctor_id = %s WHERE template_id = %s",
                    (template_name, proctor_id, template_id),
                )
            db.commit(); db.close()
            self.load_templates(); self._set_active_template(template_id); self.refresh_template_details()
            self.log_action(f"Cập nhật bộ đề ID {template_id} cho giám thị {proctor_id}")
            QMessageBox.information(self, "Thành công", f"Đã cập nhật bộ đề ID {template_id}.")
        except Exception as exc:
            self._show_dialog("Không thể cập nhật bộ đề", str(exc))

    def delete_selected_template(self):
        template_id = self._selected_template_id()
        if template_id is None:
            return QMessageBox.warning(self, "Lỗi", "Vui lòng chọn bộ đề cần xóa.")
        template_name = self.inp_manage_template_name.text().strip() or f"ID {template_id}"
        if QMessageBox.question(self, "Xác nhận xóa", f"Bạn có chắc muốn xóa bộ đề {template_name}?", QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            return
        try:
            db = self.get_db(); cur = db.cursor(dictionary=True)
            cur.execute("SELECT COUNT(*) AS total FROM classes WHERE template_id = %s", (template_id,))
            class_count = int((cur.fetchone() or {}).get('total', 0))
            if class_count > 0:
                db.close()
                return QMessageBox.warning(self, "Không thể xóa", f"Bộ đề này đang được gán cho {class_count} lớp thi. Hãy chuyển hoặc xóa các lớp đó trước.")
            cur = db.cursor()
            cur.execute("DELETE FROM template_questions WHERE template_id = %s", (template_id,))
            cur.execute("DELETE FROM question_bank WHERE template_id = %s", (template_id,))
            cur.execute("DELETE FROM exam_templates WHERE template_id = %s", (template_id,))
            db.commit(); db.close()
            self._clear_question_form()
            self.load_templates()
            self.log_action(f"Xóa bộ đề ID {template_id}")
            QMessageBox.information(self, "Thành công", f"Đã xóa bộ đề {template_name}.")
        except Exception as exc:
            self._show_dialog("Không thể xóa bộ đề", str(exc))

    def save_question(self):
        template_id = self._selected_template_id()
        if template_id is None:
            return QMessageBox.warning(self, "Lỗi", "Vui lòng chọn Bộ đề.")
        q, a, b, c, d = self.inp_q.toPlainText().strip(), self.inp_a.text().strip(), self.inp_b.text().strip(), self.inp_c.text().strip(), self.inp_d.text().strip()
        correct_ans = self.inp_correct.currentText()
        pts = self.inp_points.value()
        if not q or not a or not b:
            return QMessageBox.warning(self, "Lỗi", "Nhập câu hỏi và ít nhất 2 đáp án.")
        try:
            db = self.get_db(); cur = db.cursor()
            if self.admin_current_question_id is None:
                cur.execute("INSERT INTO question_bank (template_id, question_text, option_a, option_b, option_c, option_d, correct_option, points) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)", (template_id, q, a, b, c, d, correct_ans, pts))
                success_message = f"Đã thêm câu hỏi ({pts} điểm) vào Bộ đề."
            else:
                cur.execute("UPDATE question_bank SET template_id=%s, question_text=%s, option_a=%s, option_b=%s, option_c=%s, option_d=%s, correct_option=%s, points=%s WHERE q_id=%s", (template_id, q, a, b, c, d, correct_ans, pts, self.admin_current_question_id))
                success_message = "Đã cập nhật câu hỏi đã chọn."
            db.commit(); db.close()
            self.load_templates(); self._set_active_template(template_id); self.refresh_template_details(); self._clear_question_form()
            self.log_action(f"Cập nhật nội dung bộ đề ID {template_id}")
            QMessageBox.information(self, "Thành công", success_message)
        except Exception as exc:
            self._show_dialog("Không thể lưu câu hỏi", str(exc))

    def delete_selected_question(self):
        row = self.tb_template_questions.currentRow()
        if row < 0:
            return QMessageBox.warning(self, "Lỗi", "Hãy chọn câu hỏi cần xóa.")
        question_id = int(self.tb_template_questions.item(row, 0).text())
        if QMessageBox.question(self, "Xác nhận xóa", f"Bạn có chắc muốn xóa câu hỏi ID {question_id}?", QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            return
        try:
            db = self.get_db(); cur = db.cursor()
            cur.execute("DELETE FROM question_bank WHERE q_id = %s", (question_id,))
            db.commit(); db.close()
            self.load_templates(); self.refresh_template_details(); self._clear_question_form()
            self.log_action(f"Xóa câu hỏi ID {question_id}")
            QMessageBox.information(self, "Thành công", f"Đã xóa câu hỏi ID {question_id}.")
        except Exception as exc:
            self._show_dialog("Không thể xóa câu hỏi", str(exc))

    def import_template_file(self):
        template_id = self._selected_template_id()
        if template_id is None:
            return QMessageBox.warning(self, "Lỗi", "Vui lòng chọn bộ đề cần nhập câu hỏi.")
        filepath, _ = QFileDialog.getOpenFileName(self, "Chọn file bộ đề", "", "Excel Files (*.xlsx *.xlsm);;CSV Files (*.csv)")
        if not filepath:
            return
        try:
            extension = os.path.splitext(filepath)[1].lower()
            raw_rows = self._read_template_rows_from_csv(filepath) if extension == ".csv" else self._read_template_rows_from_excel(filepath)
            question_rows = self._normalize_question_import_rows(raw_rows)
            if not question_rows:
                raise ValueError("Không tìm thấy câu hỏi hợp lệ trong file đã chọn.")
            db = self.get_db(); cur = db.cursor()
            cur.executemany(
                "INSERT INTO question_bank (template_id, question_text, option_a, option_b, option_c, option_d, correct_option, points) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
                [(template_id, row['question_text'], row['option_a'], row['option_b'], row['option_c'], row['option_d'], row['correct_option'], row['points']) for row in question_rows],
            )
            db.commit(); db.close()
            self.load_templates(); self._set_active_template(template_id); self.refresh_template_details()
            self.log_action(f"Nhập {len(question_rows)} câu hỏi vào bộ đề ID {template_id}")
            QMessageBox.information(self, "Thành công", f"Đã nhập {len(question_rows)} câu hỏi từ file vào bộ đề.")
        except Exception as exc:
            logger.exception("Failed to import template file %s", filepath)
            self._show_dialog("Không thể nhập bộ đề", str(exc), "File cần có các cột như question_text, option_a, option_b, option_c, option_d, correct_option, points.")

    def export_template_csv(self):
        template_id = self._selected_template_id()
        if template_id is None:
            return QMessageBox.warning(self, "Lỗi", "Chọn 1 Bộ đề để xuất.")
        try:
            rows = self._load_template_questions(template_id)
            if not rows:
                return QMessageBox.warning(self, "Trống", "Bộ đề chưa có câu hỏi.")
            filepath, _ = QFileDialog.getSaveFileName(self, "Lưu file CSV", f"Bo_De_{template_id}.csv", "CSV Files (*.csv)")
            if filepath:
                with open(filepath, 'w', newline='', encoding='utf-8-sig') as handle:
                    writer = csv.DictWriter(handle, fieldnames=['q_id', 'question_text', 'option_a', 'option_b', 'option_c', 'option_d', 'correct_option', 'points'])
                    writer.writeheader(); writer.writerows(rows)
                QMessageBox.information(self, "Thành công", "Đã xuất câu hỏi ra file CSV thành công.")
        except Exception as exc:
            self._show_dialog("Không thể xuất bộ đề", str(exc))

    # ---------------- TAB 6: BÁO CÁO TOÀN HỆ THỐNG ----------------
    def setup_tab_reports(self):
        tab = QWidget(); layout = QVBoxLayout(tab)
        self.reports_filter_bar = FilterBar("Lọc theo MSV, mã lớp hoặc lỗi vi phạm", "Tìm kiếm", "Tải lại")
        wire_filter_bar(self.reports_filter_bar, self.search_reports, self.load_reports)
        self.tb_reports = QTableWidget(); self.tb_reports.setColumnCount(5)
        self.tb_reports.setHorizontalHeaderLabels(["MSV", "ID Lớp", "Thời gian vi phạm", "Lỗi vi phạm", "Tên file bằng chứng"])
        self.tb_reports.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tb_reports.setAlternatingRowColors(True)
        self.tb_reports.itemDoubleClicked.connect(self.view_evidence)
        polish_table(self.tb_reports)
        self.reports_empty = EmptyState("Chưa có báo cáo vi phạm", "Các bằng chứng vi phạm toàn hệ thống sẽ xuất hiện tại đây.")
        layout.addWidget(self.reports_filter_bar); layout.addWidget(self.reports_empty); layout.addWidget(self.tb_reports)
        self.tabs.addTab(tab, "🚨 Báo cáo")

    def load_reports(self):
        if self.reports_inflight:
            return
        self.reports_inflight = True
        self._run_background_task(
            lambda: self._fetch_admin_reports(),
            self._populate_reports_table,
            on_error=self._on_admin_report_load_error,
            on_finished=lambda: setattr(self, "reports_inflight", False),
        )

    def search_reports(self):
        keyword = self.reports_filter_bar.search_input.text().strip()
        if not keyword:
            return self.load_reports()
        if self.reports_inflight:
            return
        self.reports_inflight = True
        self._run_background_task(
            lambda: self._fetch_admin_reports(keyword),
            self._populate_reports_table,
            on_error=self._on_admin_report_load_error,
            on_finished=lambda: setattr(self, "reports_inflight", False),
        )

    def view_evidence(self, item):
        row = item.row()
        evidence_filename = self.tb_reports.item(row, 4).text() if self.tb_reports.item(row, 4) else ""
        if not evidence_filename:
            return self._show_dialog("Thiếu dữ liệu", "Dòng báo cáo này không có tên tệp bằng chứng.")
        img_url = f"{self.api_url}/evidence_images/{evidence_filename}"
        local_path = self._resolve_local_evidence_path(evidence_filename)
        api_error = None
        try:
            response = requests.get(img_url, timeout=10)
            if response.status_code == 200:
                pixmap = self._load_pixmap_from_bytes(response.content)
                if pixmap is not None and not pixmap.isNull():
                    return self._open_evidence_dialog(pixmap, row)
                api_error = "API trả dữ liệu nhưng không giải mã được ảnh."
            else:
                api_error = f"API trả về mã {response.status_code}."
        except requests.RequestException as exc:
            api_error = str(exc)
        if local_path and os.path.exists(local_path):
            with open(local_path, "rb") as handle:
                pixmap = self._load_pixmap_from_bytes(handle.read())
            if pixmap is not None and not pixmap.isNull():
                return self._open_evidence_dialog(pixmap, row)
        self._show_dialog("Không tải được bằng chứng", "Không thể mở ảnh vi phạm từ API hoặc thư mục cục bộ.", f"Tệp cần tìm: {evidence_filename}\nĐường dẫn cục bộ: {os.path.abspath(local_path or '')}\nTrạng thái API: {api_error or 'Không có phản hồi.'}")

    def _open_evidence_dialog(self, pixmap, row):
        dlg = QDialog(self)
        dlg.setWindowTitle(f"BẰNG CHỨNG VI PHẠM - {self.tb_reports.item(row, 0).text()}")
        dlg.resize(860, 620)
        dlg.setStyleSheet("QDialog { background: #FCFCFC; }")
        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(14)
        title = QLabel("Bằng chứng vi phạm")
        title.setStyleSheet("font-size: 20px; font-weight: 700; color: #111111;")
        meta = QLabel(f"MSV: {self.tb_reports.item(row, 0).text()} | Lớp: {self.tb_reports.item(row, 1).text()} | Lỗi: {self.tb_reports.item(row, 3).text()}")
        meta.setStyleSheet("color: #5C5C5C; font-size: 12px;")
        img_label = QLabel(); img_label.setAlignment(Qt.AlignCenter)
        img_label.setStyleSheet("background:#F4F4F4; border:1px solid #E2E2E2; border-radius:16px;")
        img_label.setPixmap(pixmap.scaled(780, 520, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        layout.addWidget(title); layout.addWidget(meta); layout.addWidget(img_label, 1)
        dlg.exec_()

    # ---------------- TAB 7: BẢNG ĐIỂM TOÀN HỆ THỐNG ----------------
    def setup_tab_scores(self):
        tab = QWidget(); layout = QVBoxLayout(tab)
        self.scores_filter_bar = FilterBar("Lọc theo MSV, môn thi hoặc mã lớp", "Tìm kiếm", "Tải lại")
        wire_filter_bar(self.scores_filter_bar, self.search_scores, self.load_scores)
        btn_export = QPushButton("Xuất bảng điểm")
        btn_export.clicked.connect(self.export_scores_csv)
        self.tb_scores = QTableWidget(); self.tb_scores.setColumnCount(6)
        self.tb_scores.setHorizontalHeaderLabels(["MSV", "Họ tên Sinh viên", "ID Lớp", "Môn thi", "Điểm số", "Thời gian nộp"])
        self.tb_scores.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        polish_table(self.tb_scores)
        self.scores_empty = EmptyState("Chưa có bảng điểm", "Kết quả thi toàn hệ thống sẽ được tổng hợp tại đây.")
        layout.addWidget(self.scores_filter_bar); layout.addWidget(btn_export, 0, Qt.AlignLeft); layout.addWidget(self.scores_empty); layout.addWidget(self.tb_scores)
        self.tabs.addTab(tab, "📈 Bảng điểm")

    def load_scores(self):
        try:
            db = self.get_db(); cur = db.cursor(dictionary=True)
            cur.execute("SELECT r.msv, s.full_name, c.class_id, c.class_name, r.score, r.submission_time FROM exam_results r JOIN students s ON r.msv = s.msv JOIN classes c ON r.exam_id = c.class_id ORDER BY r.submission_time DESC")
            rows = cur.fetchall(); db.close()
            self.tb_scores.setRowCount(len(rows))
            self.scores_empty.setVisible(len(rows) == 0)
            self.tb_scores.setVisible(len(rows) > 0)
            for i, row in enumerate(rows):
                self.tb_scores.setItem(i, 0, QTableWidgetItem(str(row['msv'])))
                self.tb_scores.setItem(i, 1, QTableWidgetItem(str(row['full_name'])))
                self.tb_scores.setItem(i, 2, QTableWidgetItem(str(row['class_id'])))
                self.tb_scores.setItem(i, 3, QTableWidgetItem(str(row['class_name'])))
                self.tb_scores.setItem(i, 4, QTableWidgetItem(str(row['score'])))
                self.tb_scores.setItem(i, 5, QTableWidgetItem(str(row['submission_time'])))
        except Exception:
            logger.exception("Failed to load admin scores")
            self._show_dialog("Không thể tải bảng điểm", "Không tải được bảng điểm toàn hệ thống.")

    def search_scores(self):
        keyword = self.scores_filter_bar.search_input.text().strip()
        if not keyword:
            return self.load_scores()
        try:
            db = self.get_db(); cur = db.cursor(dictionary=True)
            cur.execute("SELECT r.msv, s.full_name, c.class_id, c.class_name, r.score, r.submission_time FROM exam_results r JOIN students s ON r.msv = s.msv JOIN classes c ON r.exam_id = c.class_id WHERE r.msv LIKE %s OR c.class_name LIKE %s OR CAST(c.class_id AS CHAR) LIKE %s ORDER BY r.submission_time DESC", (f"%{keyword}%", f"%{keyword}%", f"%{keyword}%"))
            rows = cur.fetchall(); db.close()
            self.tb_scores.setRowCount(len(rows))
            self.scores_empty.setVisible(len(rows) == 0)
            self.tb_scores.setVisible(len(rows) > 0)
            for i, row in enumerate(rows):
                self.tb_scores.setItem(i, 0, QTableWidgetItem(str(row['msv'])))
                self.tb_scores.setItem(i, 1, QTableWidgetItem(str(row['full_name'])))
                self.tb_scores.setItem(i, 2, QTableWidgetItem(str(row['class_id'])))
                self.tb_scores.setItem(i, 3, QTableWidgetItem(str(row['class_name'])))
                self.tb_scores.setItem(i, 4, QTableWidgetItem(str(row['score'])))
                self.tb_scores.setItem(i, 5, QTableWidgetItem(str(row['submission_time'])))
        except Exception as exc:
            self._show_dialog("Không thể tìm bảng điểm", str(exc))

    def export_scores_csv(self):
        try:
            db = self.get_db(); cur = db.cursor(dictionary=True)
            cur.execute("SELECT r.msv AS 'Mã Sinh Viên', s.full_name AS 'Họ và Tên', c.class_id AS 'Mã Lớp', c.class_name AS 'Môn Thi', r.score AS 'Điểm Số', r.submission_time AS 'Thời gian nộp bài' FROM exam_results r JOIN students s ON r.msv = s.msv JOIN classes c ON r.exam_id = c.class_id ORDER BY c.class_id, r.score DESC")
            rows = cur.fetchall(); db.close()
            if not rows:
                return QMessageBox.warning(self, "Trống", "Chưa có sinh viên nào nộp bài!")
            filepath, _ = QFileDialog.getSaveFileName(self, "Lưu Bảng Điểm", "Bang_Diem_Tong_Hop.csv", "CSV Files (*.csv)")
            if filepath:
                with open(filepath, 'w', newline='', encoding='utf-8-sig') as handle:
                    writer = csv.DictWriter(handle, fieldnames=rows[0].keys())
                    writer.writeheader(); writer.writerows(rows)
                QMessageBox.information(self, "Thành công", f"Đã xuất bảng điểm của {len(rows)} lượt thi ra file CSV!")
        except Exception as exc:
            self._show_dialog("Không thể xuất bảng điểm", str(exc))

    # ---------------- TAB 8: CẤU HÌNH AI ----------------
    def setup_tab_configs(self):
        tab = QWidget(); layout = QVBoxLayout(tab)
        form_gb = QGroupBox("Tham số Trí Tuệ Nhân Tạo (AI)")
        fl = QFormLayout(form_gb)
        self.inp_threshold = QDoubleSpinBox(); self.inp_threshold.setRange(MIN_FACE_THRESHOLD, MAX_FACE_THRESHOLD); self.inp_threshold.setSingleStep(0.01)
        self.inp_warnings = QSpinBox(); self.inp_warnings.setRange(1, 20)
        btn_save_cfg = QPushButton("LƯU CẤU HÌNH HỆ THỐNG"); btn_save_cfg.clicked.connect(self.save_configs)
        
        fl.addRow("Ngưỡng tương đồng DeepFace:", self.inp_threshold)
        fl.addRow("Số lần cảnh báo tối đa trước khi khóa:", self.inp_warnings)
        layout.addWidget(form_gb); layout.addWidget(btn_save_cfg); layout.addStretch()
        self.tabs.addTab(tab, "⚙️ Cấu hình AI")

    def load_configs(self):
        try:
            db = self.get_db(); cur = db.cursor(dictionary=True)
            cur.execute("SELECT * FROM configs"); rows = cur.fetchall(); db.close()
            for r in rows:
                if r['setting_key'] == 'ai_face_threshold': self.inp_threshold.setValue(self._normalize_face_threshold_value(r['setting_value']))
                if r['setting_key'] == 'max_warnings': self.inp_warnings.setValue(int(r['setting_value']))
        except Exception:
            logger.exception("Failed to load configs")
            QMessageBox.warning(self, "Lỗi", "Không tải được cấu hình AI.")

    def save_configs(self):
        try:
            db = self.get_db(); cur = db.cursor()
            cur.execute("REPLACE INTO configs (setting_key, setting_value) VALUES ('ai_face_threshold', %s)", (str(self._normalize_face_threshold_value(self.inp_threshold.value())),))
            cur.execute("REPLACE INTO configs (setting_key, setting_value) VALUES ('max_warnings', %s)", (str(self.inp_warnings.value()),))
            db.commit(); db.close()
            self.log_action(f"Thay đổi cấu hình AI hệ thống")
            QMessageBox.information(self, "OK", "Đã cập nhật cấu hình AI!")
        except Exception:
            logger.exception("Failed to save configs")
            QMessageBox.critical(self, "Lỗi", "Không lưu được cấu hình AI.")

    # ---------------- TAB 9: KHÔI PHỤC MẬT KHẨU ----------------
    def setup_tab_password_resets(self):
        tab = QWidget(); layout = QVBoxLayout(tab)
        self.reset_filter_bar = FilterBar("Lọc theo vai trò, mã tài khoản hoặc trạng thái", "Tìm kiếm", "Tải lại")
        wire_filter_bar(self.reset_filter_bar, self.search_password_reset_requests, self.load_password_reset_requests)

        action_row = QHBoxLayout()
        btn_approve = QPushButton("CẤP MẬT KHẨU TẠM")
        btn_approve.clicked.connect(self.approve_selected_password_reset)
        btn_reject = QPushButton("TỪ CHỐI YÊU CẦU")
        btn_reject.setStyleSheet("background-color: #2A2A2A;")
        btn_reject.clicked.connect(self.reject_selected_password_reset)
        action_row.addWidget(btn_approve)
        action_row.addWidget(btn_reject)
        action_row.addStretch()

        self.tb_password_resets = QTableWidget(); self.tb_password_resets.setColumnCount(7)
        self.tb_password_resets.setHorizontalHeaderLabels(["ID", "Vai trò", "Mã tài khoản", "Họ tên", "Ghi chú", "Trạng thái", "Tạo lúc"])
        self.tb_password_resets.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tb_password_resets.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.tb_password_resets.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeToContents)
        polish_table(self.tb_password_resets)
        self.password_resets_empty = EmptyState("Chưa có yêu cầu", "Yêu cầu cấp lại mật khẩu từ Student, Proctor, Admin sẽ xuất hiện tại đây.")

        layout.addWidget(self.reset_filter_bar)
        layout.addLayout(action_row)
        layout.addWidget(self.password_resets_empty)
        layout.addWidget(self.tb_password_resets)
        self.tabs.addTab(tab, "🔐 Khôi phục mật khẩu")

    def _populate_password_reset_requests_table(self, rows):
        self.tb_password_resets.setRowCount(len(rows))
        self.password_resets_empty.setVisible(len(rows) == 0)
        self.tb_password_resets.setVisible(len(rows) > 0)
        for i, row in enumerate(rows):
            request_id_item = QTableWidgetItem(str(row["request_id"]))
            request_id_item.setData(Qt.UserRole, row["request_id"])
            self.tb_password_resets.setItem(i, 0, request_id_item)
            self.tb_password_resets.setItem(i, 1, QTableWidgetItem(str(row["account_role"])))
            self.tb_password_resets.setItem(i, 2, QTableWidgetItem(str(row["account_id"])))
            self.tb_password_resets.setItem(i, 3, QTableWidgetItem(str(row["full_name"])))
            self.tb_password_resets.setItem(i, 4, QTableWidgetItem(str(row.get("request_note") or "")))
            self.tb_password_resets.setItem(i, 5, QTableWidgetItem(str(row["status"])))
            self.tb_password_resets.setItem(i, 6, QTableWidgetItem(str(row["created_at"])))

    def load_password_reset_requests(self):
        try:
            if not self.admin_token:
                return self._show_dialog("Thiếu phiên quản trị", "Vui lòng đăng nhập lại để tải danh sách yêu cầu cấp lại mật khẩu.")
            response = requests.get(
                f"{self.api_url}/api/admin/password-recovery/requests",
                headers=self._admin_api_headers(),
                timeout=10,
            )
            response.raise_for_status()
            payload = response.json()
            rows = payload.get("data", []) if payload.get("status") == "success" else []
            self._populate_password_reset_requests_table(rows)
        except Exception:
            logger.exception("Failed to load password reset requests")
            self._show_dialog("Không thể tải yêu cầu", "Không tải được danh sách yêu cầu cấp lại mật khẩu.")

    def search_password_reset_requests(self):
        keyword = self.reset_filter_bar.search_input.text().strip()
        try:
            if not self.admin_token:
                return self._show_dialog("Thiếu phiên quản trị", "Vui lòng đăng nhập lại để tải danh sách yêu cầu cấp lại mật khẩu.")
            response = requests.get(
                f"{self.api_url}/api/admin/password-recovery/requests",
                headers=self._admin_api_headers(),
                params={"keyword": keyword},
                timeout=10,
            )
            response.raise_for_status()
            payload = response.json()
            rows = payload.get("data", []) if payload.get("status") == "success" else []
            self._populate_password_reset_requests_table(rows)
        except Exception as exc:
            self._show_dialog("Không thể tìm yêu cầu", str(exc))

    def _selected_password_reset_request(self):
        row = self.tb_password_resets.currentRow()
        if row < 0:
            return None
        request_id_item = self.tb_password_resets.item(row, 0)
        role_item = self.tb_password_resets.item(row, 1)
        account_item = self.tb_password_resets.item(row, 2)
        status_item = self.tb_password_resets.item(row, 5)
        if not request_id_item or not role_item or not account_item or not status_item:
            return None
        return {
            "request_id": request_id_item.data(Qt.UserRole),
            "account_role": role_item.text().strip(),
            "account_id": account_item.text().strip(),
            "status": status_item.text().strip(),
        }

    def approve_selected_password_reset(self):
        request_row = self._selected_password_reset_request()
        if not request_row:
            return self._show_dialog("Chưa chọn yêu cầu", "Hãy chọn một yêu cầu cấp lại mật khẩu trong bảng.")
        if request_row["status"] != "pending":
            return self._show_dialog("Không thể cấp lại", "Chỉ những yêu cầu đang ở trạng thái pending mới được xử lý.")
        if not self.admin_token:
            return self._show_dialog("Thiếu phiên quản trị", "Vui lòng đăng nhập lại để thực hiện duyệt yêu cầu.")

        temp_password, ok = QInputDialog.getText(self, "Mật khẩu tạm", "Nhập mật khẩu tạm mới cho tài khoản này:", QLineEdit.Password)
        temp_password = temp_password.strip()
        if not ok:
            return
        if len(temp_password) < 6:
            return self._show_dialog("Mật khẩu quá ngắn", "Mật khẩu tạm cần có ít nhất 6 ký tự.")

        try:
            response = requests.post(
                f"{self.api_url}/api/admin/password-recovery/requests/{request_row['request_id']}/approve",
                headers=self._admin_api_headers(),
                data={
                    "temp_password": temp_password,
                    "note": "Đã cấp mật khẩu tạm mới",
                },
                timeout=10,
            )
            response.raise_for_status()
            payload = response.json()
            if payload.get("status") != "success":
                return self._show_dialog("Không thể cấp lại", payload.get("message", "API từ chối cấp lại mật khẩu."))
            self.load_password_reset_requests()
            QMessageBox.information(self, "Đã cấp lại", f"Mật khẩu tạm mới cho tài khoản {request_row['account_id']} là: {temp_password}")
        except Exception as exc:
            logger.exception("Failed to approve password reset request %s", request_row["request_id"])
            self._show_dialog("Không thể cấp lại", str(exc))

    def reject_selected_password_reset(self):
        request_row = self._selected_password_reset_request()
        if not request_row:
            return self._show_dialog("Chưa chọn yêu cầu", "Hãy chọn một yêu cầu cấp lại mật khẩu trong bảng.")
        if request_row["status"] != "pending":
            return self._show_dialog("Không thể từ chối", "Chỉ những yêu cầu đang ở trạng thái pending mới được xử lý.")
        if not self.admin_token:
            return self._show_dialog("Thiếu phiên quản trị", "Vui lòng đăng nhập lại để thực hiện từ chối yêu cầu.")

        reason, ok = QInputDialog.getText(self, "Từ chối yêu cầu", "Nhập lý do từ chối (có thể để trống):")
        if not ok:
            return
        try:
            response = requests.post(
                f"{self.api_url}/api/admin/password-recovery/requests/{request_row['request_id']}/reject",
                headers=self._admin_api_headers(),
                data={"note": reason.strip() or "Từ chối yêu cầu cấp lại mật khẩu"},
                timeout=10,
            )
            response.raise_for_status()
            payload = response.json()
            if payload.get("status") != "success":
                return self._show_dialog("Không thể từ chối", payload.get("message", "API từ chối thao tác."))
            self.load_password_reset_requests()
            QMessageBox.information(self, "Đã từ chối", "Yêu cầu cấp lại mật khẩu đã được từ chối.")
        except Exception as exc:
            logger.exception("Failed to reject password reset request %s", request_row["request_id"])
            self._show_dialog("Không thể từ chối", str(exc))

    # ---------------- TAB 6: NHẬT KÝ HỆ THỐNG (AUDIT LOGS) ----------------
    def setup_tab_audit_logs(self):
        tab = QWidget(); layout = QVBoxLayout(tab)
        self.logs_filter_bar = FilterBar("Lọc theo tác nhân hoặc nội dung hành động", "Tìm kiếm", "Tải lại")
        wire_filter_bar(self.logs_filter_bar, self.search_audit_logs, self.load_audit_logs)
        
        self.tb_logs = QTableWidget(); self.tb_logs.setColumnCount(4)
        self.tb_logs.setHorizontalHeaderLabels(["ID", "Thời gian", "Tài khoản thực hiện", "Hành động chi tiết"])
        self.tb_logs.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tb_logs.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents) # Cột ID co lại cho gọn
        self.tb_logs.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        polish_table(self.tb_logs)
        self.logs_empty = EmptyState("Chưa có nhật ký", "Các thao tác quản trị sẽ xuất hiện tại đây khi hệ thống bắt đầu ghi nhận sự kiện.")
        
        layout.addWidget(self.logs_filter_bar); layout.addWidget(self.logs_empty); layout.addWidget(self.tb_logs)
        self.tabs.addTab(tab, "🛡️ Nhật ký Hệ thống")

    def load_audit_logs(self):
        """Tính năng 2: Load dữ liệu nhật ký bảo mật"""
        try:
            db = self.get_db(); cur = db.cursor(dictionary=True)
            cur.execute("SELECT * FROM audit_logs ORDER BY timestamp DESC LIMIT 100") # Lấy 100 dòng mới nhất
            rows = cur.fetchall(); db.close()
            
            self.tb_logs.setRowCount(len(rows))
            self.logs_empty.setVisible(len(rows) == 0)
            self.tb_logs.setVisible(len(rows) > 0)
            for i, r in enumerate(rows):
                self.tb_logs.setItem(i, 0, QTableWidgetItem(str(r['log_id'])))
                self.tb_logs.setItem(i, 1, QTableWidgetItem(str(r['timestamp'])))
                self.tb_logs.setItem(i, 2, QTableWidgetItem(str(r['actor'])))
                self.tb_logs.setItem(i, 3, QTableWidgetItem(str(r['action'])))
        except Exception:
            logger.exception("Failed to load audit logs")
            QMessageBox.warning(self, "Lỗi", "Không tải được nhật ký hệ thống.")

    def search_audit_logs(self):
        keyword = self.logs_filter_bar.search_input.text().strip()
        if not keyword:
            return self.load_audit_logs()
        try:
            db = self.get_db(); cur = db.cursor(dictionary=True)
            cur.execute(
                "SELECT * FROM audit_logs WHERE actor LIKE %s OR action LIKE %s ORDER BY timestamp DESC LIMIT 100",
                (f"%{keyword}%", f"%{keyword}%"),
            )
            rows = cur.fetchall(); db.close()
            self.tb_logs.setRowCount(len(rows))
            self.logs_empty.setVisible(len(rows) == 0)
            self.tb_logs.setVisible(len(rows) > 0)
            for i, r in enumerate(rows):
                self.tb_logs.setItem(i, 0, QTableWidgetItem(str(r['log_id'])))
                self.tb_logs.setItem(i, 1, QTableWidgetItem(str(r['timestamp'])))
                self.tb_logs.setItem(i, 2, QTableWidgetItem(str(r['actor'])))
                self.tb_logs.setItem(i, 3, QTableWidgetItem(str(r['action'])))
        except Exception as exc:
            self._show_dialog("Không thể tìm nhật ký", str(exc))

    # ---------------- TAB 7: SAO LƯU DỮ LIỆU (BACKUP) ----------------
    def setup_tab_backup(self):
        tab = QWidget(); layout = QVBoxLayout(tab)
        lbl = QLabel("Sao lưu toàn bộ Dữ liệu Hệ thống ra file CSV để báo cáo.")
        lbl.setStyleSheet("font-size: 16px; margin-bottom: 20px;")
        btn_backup = QPushButton("💾 XUẤT DỮ LIỆU BÁO CÁO (CSV)")
        btn_backup.setStyleSheet("background-color: #111111; font-size: 16px; padding: 20px;")
        btn_backup.clicked.connect(self.backup_data)
        layout.addWidget(lbl); layout.addWidget(btn_backup); layout.addStretch()
        self.tabs.addTab(tab, "💾 Sao lưu Dữ liệu")

    def backup_data(self):
        os.makedirs("backups", exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        try:
            db = self.get_db(); cur = db.cursor()
            tables = ['students', 'proctors', 'classes', 'exam_results', 'violations']
            for table in tables:
                cur.execute(f"SELECT * FROM {table}")
                rows = cur.fetchall()
                if not rows: continue
                headers = [i[0] for i in cur.description]
                filepath = f"backups/{table}_backup_{timestamp}.csv"
                with open(filepath, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    writer.writerows(rows)
            db.close()
            self.log_action("Thực hiện sao lưu toàn bộ dữ liệu hệ thống")
            QMessageBox.information(self, "Thành công", f"Đã xuất dữ liệu ra thư mục 'backups'!")
        except Exception as e: QMessageBox.critical(self, "Lỗi", f"Lỗi sao lưu: {e}")

    # ================= ĐĂNG NHẬP VÀ KHỞI TẠO =================
    def process_login(self):
        aid, pw = self.inp_aid.text().strip(), self.inp_apw.text().strip()
        if self.admin_login_inflight:
            return
        if not aid or not pw:
            QMessageBox.warning(self, "Lỗi", "Vui lòng nhập ID và mật khẩu Quản trị.")
            return
        self._set_login_busy(True)
        self._run_background_task(
            lambda: self._fetch_admin_login_bundle(aid, pw),
            self._on_admin_login_finished,
            on_error=self._on_admin_login_error,
            on_finished=lambda: self._set_login_busy(False),
        )

if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = AdminDashboard(); win.show()
    sys.exit(app.exec_())